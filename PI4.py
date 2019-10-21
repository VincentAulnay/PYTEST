from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import xlwt
import xlrd
import time
from xlutils.copy import copy
import datetime
import datetime as dt
from tkinter import filedialog
from tkinter import *
from bs4 import BeautifulSoup
from urllib.request import urlopen
import os
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
import threading
import sys



chrome_options = webdriver.ChromeOptions()
prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)
#chrome_options.add_argument("-headless")
#chrome_options.add_argument("-disable-gpu")

print ('▀▄▀▄▀▄ STOPBNB ▄▀▄▀▄▀')


#-----EXCEL RESULT OPEN AND READ-----

#book = xlrd.open_workbook(path_RESULT.filename)
#wb=copy(book)
#sheet_write = wb.get_sheet(0)
#sheet_read = book.sheet_by_index(0)

wbx = load_workbook(path_RESULT.filename)
ws = wbx.active

#-------FIND COLUMN UPDATE------
up=0
k=1
while up==0:
	#V_up=sheet_read.cell(0,i).value
	V_up=ws.cell(row=1, column=k).value
	if V_up=='UPDATE_CALENDAR':
		up=1
	else:
		k=k+1
		
c_mouth=k+1


#-----RECUP INFO XPATH FROM EXCEL------
book_GMAIL = xlrd.open_workbook('/home/pi/Desktop/GMAIL_ACCOUNT.xls')
sheet_GMAIL = book_GMAIL.sheet_by_index(0)
ADRESS_GMAIL=sheet_GMAIL.cell(0,1).value
PSW_GMAIL=sheet_GMAIL.cell(1,1).value
RECEIVER=sheet_GMAIL.cell(2,1).value

#-------DATE DU JOUR-------
date = int(datetime.datetime.now().day)
month = int(datetime.datetime.now().month)
Hr=dt.datetime.now().hour

#------RECUP INFO CALANDAR------

def email(DIR2,NAMEFile,now):
	sender = ADRESS_GMAIL
	sender_password = PSW_GMAIL
	receivers = RECEIVER

	s = smtplib.SMTP('smtp.gmail.com', 587)
	s.starttls()
	s.login(sender, sender_password)
	msg = MIMEMultipart()
	msg['From'] = sender
	msg['To'] = receivers
	#msg['Subject'] = "Subject of the Mail- image -2"
	body = "Body_of_the_mail"
	msg.attach(MIMEText(body, 'plain'))
	msg['Subject'] = "STOP AIRBNB - extraction du - "+str(now)
	# path along with extension of file to be attachmented 
	filename = DIR2+NAMEFile+str(now)+".xlsx"
	attachmentment = open(filename, "rb")
	 
	# instance of MIMEBase and named as p
	attachment = MIMEBase('application', 'octet-stream')
	# To change the payload into encoded form
	attachment.set_payload((attachmentment).read())
	# encode into base64
	encoders.encode_base64(attachment)
	attachment.add_header('Content-Disposition', "attachmentment; filename= %s" % filename)
	# attachment the instance  to instance 'msg'
	msg.attach(attachment)
	text = msg.as_string()
	s.sendmail(sender, receivers, text)
	print('*** email sent ***') 
	time.sleep(10)
	del filename
	del attachmentment
	del attachment
	del text
	del msg
def whatmounth():
	month = int(datetime.datetime.now().month)
	global name_mois1
	global name_mois2
	global name_mois3
	global name_mois4
	global name_mois5
	if month==1:
		name_mois1='janvier 2020'
		name_mois2='février 2020'
		name_mois3='mars 2020'
		name_mois4='avril 2020'
		name_mois5='mai 2020'
	elif month==2:
		name_mois1='février 2020'
		name_mois2='mars 2020'
		name_mois3='avril 2020'
		name_mois4='mai 2020'
		name_mois5='juin 2020'
	elif month==3:
		name_mois1='mars 2020'
		name_mois2='avril 2020'
		name_mois3='mai 2020'
		name_mois4='juin 2020'
		name_mois5='juillet 2020'
	elif month==4:
		name_mois1='avril 2020'
		name_mois2='mai 2020'
		name_mois3='juin 2020'
		name_mois4='juillet 2020'
		name_mois5='août 2020'
	elif month==5:
		name_mois1='mai 2020'
		name_mois2='juin 2020'
		name_mois3='juillet 2020'
		name_mois4='août 2020'
		name_mois5='septembre 2020'
	elif month==6:
		name_mois1='juin 2020'
		name_mois2='juillet 2020'
		name_mois3='août 2020'
		name_mois4='septembre 2020'
		name_mois5='octobre 2020'
	elif month==7:
		name_mois1='juillet 2020'
		name_mois2='août 2020'
		name_mois3='septembre 2020'
		name_mois4='octobre 2020'
		name_mois5='novembre 2020'
	elif month==8:
		name_mois1='août 2020'
		name_mois2='septembre 2020'
		name_mois3='octobre 2020'
		name_mois4='novembre 2020'
		name_mois5='décembre 2020'
	elif month==9:
		name_mois1='septembre 2019'
		name_mois2='octobre 2019'
		name_mois3='novembre 2019'
		name_mois4='décembre 2019'
		name_mois5='janvier 2020'
	elif month==10:
		name_mois1='octobre 2019'
		name_mois2='novembre 2019'
		name_mois3='décembre 2019'
		name_mois4='janvier 2020'
		name_mois5='février 2020'
	elif month==11:
		name_mois1='novembre 2019'
		name_mois2='décembre 2019'
		name_mois3='janvier 2020'
		name_mois4='février 2020'
		name_mois5='mars 2020'
	elif month==12:
		name_mois1='décembre 2019'
		name_mois2='janvier 2020'
		name_mois3='fevrier 2020'
		name_mois4='mars 2020'
		name_mois5='avril 2020'
def MnumDay (Mmois):
	global MNumday
	if Mmois=='janvier':
		MNumday=31
	elif Mmois=='février':
		MNumday=28
	elif Mmois=='mars':
		MNumday=31	
	elif Mmois=='avril':
		MNumday=30
	elif Mmois=='mai':
		MNumday=31
	elif Mmois=='juin':
		MNumday=30
	elif Mmois=='juillet':
		MNumday=31
	elif Mmois=='août':
		MNumday=30
	elif Mmois=='septembre':
		MNumday=31
	elif Mmois=='octobre':
		MNumday=30
	elif Mmois=='novembre':
		MNumday=31
	elif Mmois=='décembre':
		MNumday=30
		
def A_Colonne_mois(name_mois,c):
#1- récupération book Result qui évolue au court du script
#2- compter le nombre de colonne
#3- déterminer si colonne == name_mois de airbnb
#4- si condition alors c_write=c pour définir la colonne où écrire
	global c_write
	global new_month
	book_mois = xlrd.open_workbook(path_RESULT.filename, on_demand = True)
	sheet_mois = book_mois.sheet_by_index(0)
	nc=sheet_mois.ncols
	book_mois.release_resources()
	del book_mois
	
	new_month=0
	
	find_month=0
	while find_month==0:
		this_month=ws.cell(row=1, column=c+1).value
		if this_month==name_mois:
			c_write=c+1
			break
		elif this_month==None:
			ws.cell(row=1, column=c+1).value = name_mois
			ws.cell(row=1, column=c+2).value = 'NB_COMMENT'
			ws.cell(row=1, column=c+3).value = 'DIF_COMMENT'
			ws.cell(row=1, column=c+4).value = 'NB_/A'
			ws.cell(row=1, column=c+5).value = 'NB_NO/A'
			ws.cell(row=1, column=c+6).value = 'SUM_NB'
			ws.cell(row=1, column=c+7).value = 'nJ_/A'
			ws.cell(row=1, column=c+8).value = 'nJ_NO/A'
			ws.cell(row=1, column=c+9).value = 'SUM_nJ'
			ws.cell(row=1, column=c+10).value = 'SUM_all_nJ/A'
			ws.cell(row=1, column=c+11).value = 'SUM_all_nJ'
			ws.cell(row=1, column=c+12).value = 'nb_/P'
			ws.cell(row=1, column=c+13).value = 'nJ_/D'
			ws.cell(row=1, column=c+14).value = 'TOTAL_J'
			c_write=c+1
			find_month=1
			new_month=1
			print ('plus une colonne')
			wbx.save(path_RESULT.filename)
			break
		else:
			c=c+1

def A_Statu_day2(date,c_write,page,j,g,ResAirbnb,new_mo,MNday,ONCOM):	
	int_timeday=int(date)
	month=soup.findAll('div', attrs={"class":u"_1lds9wb"})[g]
	i=0
	li=[]
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			the_tr= month.findAll('td', attrs={"class": "_z39f86g"})[i]
			div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
			intdiv=int(div)
			if intdiv>=int_timeday:
				li.append(intdiv)
			i=i+1
		except:
			break
	#print (li)
	try:
		if len(li)>0:
			ca=ws.cell(row=j, column=c_write).value
			#print(ca)
			#-------DATE DU JOUR-------
			date = int(datetime.datetime.now().day)
			month = int(datetime.datetime.now().month)
			toto=str(date)+'-'+str(month)
			if ca!=None:
				li_ca=ca.split(";")
			else:
				li_ca=[]

			lie=[]
			if li_ca!=[]:
				lenL=len(li_ca)
				h=0
				LB=[]
				while h!=lenL:
					LA=li_ca[h]
					LA=LA.split(':')
					del LA[0]
					LA=LA[0].split(',')
					lenLA=len(LA)
					g=0
					while g!=lenLA:
						intV=int(LA[g])
						LB.append(intV)
						g=g+1
					h=h+1
			
				lie=[elem for elem in li if elem not in LB ]
				if len(lie)!=0:
					#identification si nuitée est bloquée par préavis automatique
					preavis=''
					if len(lie)==1:
						dif=lie[0]-date
						preavis=''
						if dif==0 or dif==1 or dif==2 or dif==6:
							preavis='/P'
						elif dif<0:
							difP=MNday-date+lie[0]
							if difP==0 or difP==1 or difP==2 or difP==6:
								preavis='/P'
							
					t=ResAirbnb+preavis+toto+':'+str(lie)
					t=t.replace("[","")
					t=t.replace("]","")
					r=str(ca)+';    '+t
					#lenli=len(lie)+len(LB)
					#ws.cell(row=j, column=c_write+3).value=lenli
			else:
				t=ResAirbnb+toto+':'+str(li)
				t=t.replace("[","")
				t=t.replace("]","")
				r=t
				#print(r)
				#lenli=len(li)
				#ws.cell(row=j, column=c_write+3).value=lenli
			if r!='set()':
				print (r)
				ws.cell(row=j, column=c_write).value=r
	except:
		#print('rater 1')
		pass
	#COMMENTAIRE
	ONC=ONCOM
	if ONC==1:
		try:
			Bcomment=soup.find('button', attrs={"class": "_ff6jfq"})
			Scomment=Bcomment.find('span', attrs={"class": "_so3dpm2"}).text
			ws.cell(row=j, column=c_write+1).value=Scomment
		except:
			print('pas de commentaire')
			pass
	
def A_Statu_day4(c_write,j,ResAirbnb,new_mo):	
	month5=soup.find('div', attrs={"class":u"_kuxo8ai"})
	i=0
	li=[]
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			the_tr= month5.findAll('td', attrs={"class": "_z39f86g"})[i]
			div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
			intdiv=int(div)
			li.append(intdiv)
			i=i+1
		except:
			break
	try:
		if len(li)>0:
			ca=ws.cell(row=j, column=c_write).value
			#-------DATE DU JOUR-------
			date = int(datetime.datetime.now().day)
			month = int(datetime.datetime.now().month)
			toto=str(date)+'-'+str(month)
			if ca!=None:
				li_ca=ca.split(";")
			else:
				li_ca=[]

			lie=[]
			if li_ca!=[]:
				lenL=len(li_ca)
				h=0
				LB=[]
				while h!=lenL:
					LA=li_ca[h]
					LA=LA.split(':')
					del LA[0]
					LA=LA[0].split(',')
					lenLA=len(LA)
					g=0
					while g!=lenLA:
						intV=int(LA[g])
						LB.append(intV)
						g=g+1
					h=h+1
			
				lie=[elem for elem in li if elem not in LB ]
				if len(lie)!=0:
					t=ResAirbnb+toto+':'+str(lie)
					t=t.replace("[","")
					t=t.replace("]","")
					r=str(ca)+';    '+t
					#lenli=len(lie)+len(LB)
					#ws.cell(row=j, column=c_write+3).value=lenli
			else:
				t=ResAirbnb+toto+':'+str(li)
				t=t.replace("[","")
				t=t.replace("]","")
				r=t
				#lenli=len(li)
				#ws.cell(row=j, column=c_write+3).value=lenli
			if r!='set()':
				print (r)
				ws.cell(row=j, column=c_write).value=r
	except:
		pass

def A_Statu_day6(c_write,j,ResAirbnb,new_mo):	
	month5=soup2.find('div', attrs={"class":u"_kuxo8ai"})
	i=0
	li=[]
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			the_tr= month5.findAll('td', attrs={"class": "_z39f86g"})[i]
			div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
			intdiv=int(div)
			li.append(intdiv)
			i=i+1
		except:
			break
	try:
		if len(li)>0:
			ca=ws.cell(row=j, column=c_write).value
			#-------DATE DU JOUR-------
			date = int(datetime.datetime.now().day)
			month = int(datetime.datetime.now().month)
			toto=str(date)+'-'+str(month)
			if ca!=None:
				li_ca=ca.split(";")
			else:
				li_ca=[]

			lie=[]
			if li_ca!=[]:
				lenL=len(li_ca)
				h=0
				LB=[]
				while h!=lenL:
					LA=li_ca[h]
					LA=LA.split(':')
					del LA[0]
					LA=LA[0].split(',')
					lenLA=len(LA)
					g=0
					while g!=lenLA:
						intV=int(LA[g])
						LB.append(intV)
						g=g+1
					h=h+1
			
				lie=[elem for elem in li if elem not in LB ]
				if len(lie)!=0:
					t=ResAirbnb+toto+':'+str(lie)
					t=t.replace("[","")
					t=t.replace("]","")
					r=str(ca)+';    '+t
					#lenli=len(lie)+len(LB)
					#ws.cell(row=j, column=c_write+3).value=lenli
			else:
				t=ResAirbnb+toto+':'+str(li)
				t=t.replace("[","")
				t=t.replace("]","")
				r=t
				#lenli=len(li)
				#ws.cell(row=j, column=c_write+3).value=lenli
			if r!='set()':
				print (r)
				ws.cell(row=j, column=c_write).value=r
	except:
		pass
		
def A_Statu_day5(c_write,j,ResAirbnb,new_mo,g):	
	month5=soup2.findAll('div', attrs={"class":u"_1lds9wb"})[g]
	i=0
	li=[]
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			the_tr= month5.findAll('td', attrs={"class": "_z39f86g"})[i]
			div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
			intdiv=int(div)
			li.append(intdiv)
			i=i+1
		except:
			break
	try:
		if len(li)>0:
			ca=ws.cell(row=j, column=c_write).value
			#-------DATE DU JOUR-------
			date = int(datetime.datetime.now().day)
			month = int(datetime.datetime.now().month)
			toto=str(date)+'-'+str(month)
			if ca!=None:
				li_ca=ca.split(";")
			else:
				li_ca=[]

			lie=[]
			if li_ca!=[]:
				lenL=len(li_ca)
				h=0
				LB=[]
				while h!=lenL:
					LA=li_ca[h]
					LA=LA.split(':')
					del LA[0]
					LA=LA[0].split(',')
					lenLA=len(LA)
					g=0
					while g!=lenLA:
						intV=int(LA[g])
						LB.append(intV)
						g=g+1
					h=h+1
			
				lie=[elem for elem in li if elem not in LB ]
				if len(lie)!=0:
					t=ResAirbnb+toto+':'+str(lie)
					t=t.replace("[","")
					t=t.replace("]","")
					r=str(ca)+';    '+t
					#lenli=len(lie)+len(LB)
					#ws.cell(row=j, column=c_write+3).value=lenli
			else:
				t=ResAirbnb+toto+':'+str(li)
				t=t.replace("[","")
				t=t.replace("]","")
				r=t
				#lenli=len(li)
				#ws.cell(row=j, column=c_write+3).value=lenli
			if r!='set()':
				print (r)
				ws.cell(row=j, column=c_write).value=r
	except:
		pass
	

def COMPUTE_M1(name_mois1):
	Dif_c=1
	if Dif_c==1:
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up==name_mois1:
				up=1
			else:
				i=i+1
		#print('Cmois='+str(i))
		Cmois=i

		up=0
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NB_COMMENT':
				up=1
			else:
				i=i+1
		#print('Ccomment1='+str(i))
		Ccomment1=i

		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='DIF_COMMENT':
				up=1
			else:
				i=i+1
		#print('DIF_Comment='+str(i))
		DIF_Comment=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NB_/A':
				up=1
			else:
				i=i+1
		C_nbA=i

		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NB_NO/A':
				up=1
			else:
				i=i+1
		C_nbnoA=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SUM_NB':
				up=1
			else:
				i=i+1
		C_SUMnb=i

		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='nJ_/A':
				up=1
			else:
				i=i+1
		C_nJA=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='nJ_NO/A':
				up=1
			else:
				i=i+1
		C_NOnJA=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SUM_nJ':
				up=1
			else:
				i=i+1
		#print('SUM_nJ='+str(i))
		C_SUMnJ=i

		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SUM_all_nJ':
				up=1
			else:
				i=i+1
		C_SUM_all_nJ=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='nb_/P':
				up=1
			else:
				i=i+1
		C_nb_P=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='nJ_/D':
				up=1
			else:
				i=i+1
		C_nJD=i
		
		up=0
		i=Cmois
		try:
			while up==0:
				V_up=ws.cell(row=1, column=i).value
				if V_up=='NB_COMMENT':
					up=1
				else:
					i=i-1
			#print('Ccommont2='+str(i))
			Ccomment2=i
			NOC2=0
		except:
			NOC2=1
			#print ('NOC2=====1')
	c=2
	while c<=nrow:
		if NOC2==0:
			V1=ws.cell(row=c, column=Ccomment1).value
			V2=ws.cell(row=c, column=Ccomment2).value
			try:
				DIF=int(V1)-int(V2)
				#print('ANNONCE:'+str(c)+('   DIF:')+str(DIF))
				ws.cell(row=c, column=DIF_Comment).value=DIF
			except:
				pass
	#--------COUNT NB/A and NB NO/A---------
		STR_NBA=ws.cell(row=c, column=Cmois).value
		continu=1
		if STR_NBA==None:
			continu=0
		if continu==1:
			count_AP=0
			count_D=0
			count_P=0
			count=0
			count_NBA=0
			count_AP=STR_NBA.count('/A/P')
			count_NBA=STR_NBA.count('/A')
			real_NBA=count_NBA-count_AP
			count_P=STR_NBA.count('/P')
			count_D=STR_NBA.count('/D')
			count=STR_NBA.count(':')
			
			NBNOA=count-count_D-real_NBA-count_P-count_AP
			#print (('NB_NO/A ===')+str(NBNOA))
			ws.cell(row=c, column=C_nbA).value=real_NBA
			ws.cell(row=c, column=C_nbnoA).value=NBNOA
			ws.cell(row=c, column=C_nb_P).value=count_P
			write=int(NBNOA)+int(real_NBA)
			ws.cell(row=c, column=C_SUMnb).value=write
		#---------COUNT nJ ---------
			list=STR_NBA.split(';')
			B=['/P', '/D', '/A/P']
			blacklist = re.compile('|'.join([re.escape(word) for word in B]))
			newL=[word for word in list if not blacklist.search(word)]
			D=['/D']
			blacklistD = re.compile('|'.join([re.escape(wordD) for wordD in D]))
			newLforD=[wordD for wordD in list if blacklistD.search(wordD)]
			rd=0
			lenD=len(newLforD)
			nbD=0
			while rd<lenD:
				pnlD=newLforD[rd].split(':')
				del pnlD[0]
				pld=pnlD[0].split(',')
				nbD=nbD+len(pld)
				rd=rd+1
			ws.cell(row=c, column=C_nJD).value=nbD
			#[x for x in list if not x.startswith('/A/P') and not x.startswith('/D') and not x.startswith('/P')]
			#[x for x in list if not any(bad in x for bad in B)]
			#-----/A--------
			BA=['/A']
			blacklistA = re.compile('|'.join([re.escape(wordA) for wordA in BA]))
			newLforA=[wordA for wordA in newL if blacklistA.search(wordA)] #-------Creation list AVEC que les lot /A
			newLfornoA=[wordA for wordA in newL if not blacklistA.search(wordA)] #-------Creation list SANS les lot /A
			nAlen=len(newLforA)
			rr=0
			nbA=0
			try: #---Recuperation nJ dans les lot /A
				while rr<nAlen:
					pnlA=newLforA[rr].split(':')
					del pnlA[0]
					pla=pnlA[0].split(',')
					nbA=nbA+len(pla)
					rr=rr+1
			except:
				pass
			ws.cell(row=c, column=C_nJA).value=nbA
			nAlen=len(newLfornoA)
			rr=0
			NnoJA=0
			try: #---Recuperation nJ dans les lot SANS /A
				while rr<nAlen:
					pnlA=newLfornoA[rr].split(':')
					del pnlA[0]
					pla=pnlA[0].split(',')
					NnoJA=NnoJA+len(pla)
					rr=rr+1
			except:
				pass
			ws.cell(row=c, column=C_NOnJA).value=NnoJA
			write=int(nbA)+int(NnoJA)
			ws.cell(row=c, column=C_SUMnJ).value=write
		c=c+1
	wbx.save(path_RESULT.filename)

	
rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
rootdriver2 = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
rootdriver.set_window_size(2000, 1000)
rootdriver2.set_window_size(2000, 1000)
wait = WebDriverWait(rootdriver, 5)
wait2 = WebDriverWait(rootdriver2, 5)
f_xpathdate=0
w_month=0
c_month=0
fff=1
while w_month==0:
	run_month=whatmounth()
	print (name_mois1)
	print (name_mois2)
	print (name_mois3)
	print (name_mois4)
	print (name_mois5)
	w_month=1
while c_month==0:
#--mois1--
	print(name_mois1)
	Mname1=name_mois1.split(' ')
	MN1=Mname1[0]
	run_MN=MnumDay(MN1)
	MNday1=MNumday
	run_c=A_Colonne_mois(name_mois1,k)
	m1_write=c_write
	m1_newmonth=new_month
	print (m1_newmonth)
#--mois 2--
	print(name_mois2)
	Mname2=name_mois2.split(' ')
	MN2=Mname2[0]
	run_MN=MnumDay(MN2)
	MNday2=MNumday
	run_c=A_Colonne_mois(name_mois2,k)
	m2_write=c_write
	m2_newmonth=new_month
	print (m2_newmonth)
#--mois 3--
	print(name_mois3)
	Mname3=name_mois3.split(' ')
	MN3=Mname3[0]
	run_MN=MnumDay(MN3)
	MNday3=MNumday
	run_c=A_Colonne_mois(name_mois3,k)
	m3_write=c_write
	m3_newmonth=new_month
	print (m3_newmonth)
#--mois 4--
	print(name_mois4)
	Mname4=name_mois4.split(' ')
	MN4=Mname4[0]
	run_MN=MnumDay(MN4)
	MNday4=MNumday
	run_c=A_Colonne_mois(name_mois4,k)
	m4_write=c_write
	m4_newmonth=new_month
	print (m4_newmonth)
#--mois 5--
	print(name_mois5)
	Mname5=name_mois5.split(' ')
	MN5=Mname5[0]
	run_MN=MnumDay(MN5)
	MNday5=MNumday
	run_c=A_Colonne_mois(name_mois5,k)
	m5_write=c_write
	m5_newmonth=new_month
	print (m5_newmonth)
	time.sleep(5)
	c_month=1		

while f_xpathdate==0:
	h=ws.cell(row=fff, column=2).value
	print(h)
	if fff==5:
		f_mounth=1
		f_xpathdate=1
		end=0
		#run=emailfalde2()
	fff=fff+1
	try:
		rootdriver.get(h)
		#time.sleep(2)
		#html = rootdriver.page_source
		#soup = BeautifulSoup(html, 'html.parser')
		time.sleep(4)
		x_date = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_13m7kz7i']"))).text
		print("x date trouve")
		f_xpathdate=1
	except:
		if fff!=5:
			rootdriver.quit()
			rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
			rootdriver.set_window_size(1000, 1500)
			wait = WebDriverWait(rootdriver, 3)
	#-----OPEN GOOGLE CHROME and AIRBNB PAGE---------

nrow=ws.max_row
print('NROW'+str(nrow))
j=2
z=0
end=0
EE=0
Tr=0
C_mois=0
C_mois5=0
drive=0
date = int(datetime.datetime.now().day)
def f1(a):
	global hok
	print('F1')
	rootdriver.get(a)
	print('OK')
	hok=1
def f2(a):
	global hok2
	print('F2')
	rootdriver2.get(str(a)+'?check_in=2019-11-05&check_out=2019-11-06')
	print('OK2')
	hok2=1
h=ws.cell(row=j, column=2).value
rootdriver.get(h)
v_nextmounth='?check_in=2019-11-05&check_out=2019-11-06'
rootdriver2.get(str(h)+str(v_nextmounth))
time.sleep(2)
hok=1
hok2=1
while end==0:
	try:
		while j<=nrow:
			print('------'+str(j-1)+'------'+str(h))
			if j+1<=nrow:
				h=ws.cell(row=j+1, column=2).value
			
			if h==None:
				j=j+1
				print('h=None')
			elif 'airbnb' in h:
				ResAirbnb=''
				V_up=ws.cell(row=j, column=k).value
				e=0
				while hok!=1:
					time.sleep(1)
					e=e+1
					if e==10:
						hok=1
				ee=0
				while hok2!=1:
					time.sleep(1)
					ee=ee+1
					if ee==10:
						hok2=1
				hok=0
				hok2=0
				time.sleep(6)
				html = rootdriver.page_source
				soup = BeautifulSoup(html, 'html.parser')
				time.sleep(3)
				html2 = rootdriver2.page_source
				soup2 = BeautifulSoup(html2, 'html.parser')
				time.sleep(3)
				threading.Thread(target=f1, args=(h,)).start()
				threading.Thread(target=f2, args=(h,)).start()
				time.sleep(1)
				try:
				#-----RECUPERATION CALANDAR MOIS 1--------
					print('le mois 1 est '+name_mois1)
					#threading.Thread(target=A_Statu_day2, args=(date,m1_write,1,j,0,ResAirbnb,m1_newmonth,500,1,)).start()
					run_day=A_Statu_day2(date,m1_write,1,j,0,ResAirbnb,m1_newmonth,500,1)
				except:
					pass
				try:
				#-----RECUPERATION CALANDAR MOIS 2--------
					print('le mois 2 est '+name_mois2)
					#threading.Thread(target=A_Statu_day2, args=(1,m2_write,2,j,1,ResAirbnb,m2_newmonth,MNday1,0,)).start()
					run_day=A_Statu_day2(1,m2_write,2,j,1,ResAirbnb,m2_newmonth,MNday1,0)
				except:
					pass
				try:
				#-----RECUPERATION CALANDAR MOIS 3--------
					print('le mois 3 est '+name_mois3)
					#threading.Thread(target=A_Statu_day4, args=(m3_write,j,RA4,m3_newmonth,)).start()
					run_resday=A_Statu_day4(m3_write,j,ResAirbnb,m3_newmonth)
				except:
					#print('PAS DE MOIS 3')
					pass
			#-----MOIS 4-5 -----
				try:
					print('le mois 4 est '+name_mois4)
					#threading.Thread(target=A_Statu_day5, args=(m4_write,j,ResAirbnb,m4_newmonth,0,)).start()
					run_day=A_Statu_day5(m4_write,j,ResAirbnb,m4_newmonth,1)
				except:
					pass
					#-----RECUPERATION CALANDAR MOIS 5--------
				try:
					print('le mois 5 est '+name_mois5)
					#threading.Thread(target=A_Statu_day5, args=(m5_write,j,ResAirbnb,m5_newmonth,1,)).start()
					RA4=ResAirbnb
					if date==1:
						RA4='/D'
						run_day=A_Statu_day6(m5_write,j,ResAirbnb,m5_newmonth)
				except:
					pass
				if (j/10).is_integer():
					print('A2')
					wbx.save(path_RESULT.filename)
				j=j+1
			elif 'abritel' in h:
				j=j+1
			else:
				j=j+1
		
		end=1
		wbx.save(path_RESULT.filename)
		now = str(datetime.datetime.now())[:19]
		now = now.replace(":","_")
		Tr=date
		print ('_______    ___    ___     ___')
		print ('|      |   |  |   |  \    |  |')
		print ('|  |__     |  |   |   \   |  |')
		print ('|     |    |  |   |    \  |  |')
		print ('|  |       |  |   |  |\ \ |  |')
		print ('|  |       |  |   |  | \ \|  |')
		print ('|__|       |__|   |__|  \____|')
		wbx = load_workbook(path_RESULT.filename)
		ws = wbx.active
		COMPUTE_M1(name_mois1)
		COMPUTE_M1(name_mois2)
		#COMPUTE_M1(name_mois3)
		#COMPUTE_M1(name_mois4)
		#COMPUTE_M1(name_mois5)
		wbx.save(DIR2+NAMEFile+str(now)+".xlsx")
		#run=email(DIR2,NAMEFile,now)
		rootdriver.quit()
		rootdriver2.quit()
		wbx.close()
	except:
		try:
			rootdriver.quit()
			rootdriver2.quit()
		except:
			pass
		# EXCEPT si Chrome se ferme tout seul, ici il va le réouvrir et relancer la boucle d'extraction
		rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
		rootdriver2 = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
		rootdriver.set_window_size(1000, 1500)
		rootdriver2.set_window_size(1000, 1500)
		wait = WebDriverWait(rootdriver, 3)
		wait2 = WebDriverWait(rootdriver2, 3)


#print('FIN')