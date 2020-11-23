import xlwt
import xlrd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from xlutils.copy import copy
from decimal import Decimal
from decimal import *
import decimal
#import xlwings as xw
#from xlwings.constants import DeleteShiftDirection
from xlrd import open_workbook,XL_CELL_TEXT
import datetime
from datetime import date
import datetime as dt
from tkinter import filedialog
from tkinter import *
import os
import re
import json
from urllib.request import urlopen
import pip
import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from shapely.geometry import Point
from shapely.geometry.polygon import Polygon
import datetime
import datetime as dt

#--------SELECTION DU FICHIER--------
path_RESULT = Tk()
Label1 = Label(path_RESULT, text = "Select File", fg = 'red')
Label1.pack()
path_RESULT.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select File",filetypes = (("Excel file","*.xlsx"),("all files","*.*")))
print (path_RESULT.filename)
NAMEFile=os.path.splitext(os.path.basename(path_RESULT.filename))[0]
print(NAMEFile)
DIR=os.path.dirname(path_RESULT.filename)
DIR2=DIR+'/'
print(os.path.dirname(path_RESULT.filename))

#now = str(datetime.date.now())[:19]
#now = now.replace(":","_")
now = str(date.today())
print(now)

Deccontext = Context(prec=10, rounding=ROUND_HALF_DOWN)
setcontext(Deccontext)

chrome_options = webdriver.ChromeOptions()
#prefs = {"profile.managed_default_content_settings.images": 2}
#chrome_options.add_experimental_option("prefs", prefs)

#on récupère la value dans le excel ORIGINE
#32460


wb = load_workbook(path_RESULT.filename)
ws=wb.active
nrow=ws.max_row


searchcolumn=1
if searchcolumn==1:
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NUMERO':
				up=1
			else:
				i=i+1
		cNUMERO=i
		print('cNUMERO')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='TITRE':
				up=1
			else:
				i=i+1
		cTITLE=i
		print('cTITLE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ANNONCE':
				up=1
			else:
				i=i+1
		cANNONCE=i
		print('cANNONCE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NAME_HOTE':
				up=1
			else:
				i=i+1
		cNAME_HOTE=i
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ANCIENNETE':
				up=1
			else:
				i=i+1
		cANCIENNETE=i
		print('cANCIENNETE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='HOTE':
				up=1
			else:
				i=i+1
		cHOTE=i
		print('cHOTE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='PRICE':
				up=1
			else:
				i=i+1
		cPRICE=i
		print('cPRICE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COMMENTAIRE':
				up=1
			else:
				i=i+1
		cCOMMENT=i
		print('cCOMMENTAIRE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='TYPE_LOGEMENT':
				up=1
			else:
				i=i+1
		cTYPE_LOGEMENT=i
		print('cTYPE_LOGEMENT')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='VOYAGEUR':
				up=1
			else:
				i=i+1
		cVOYAGEUR=i
		print('cVOYAGEUR')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='CHAMBRE':
				up=1
			else:
				i=i+1
		cCHAMBRE=i
		print('cCHAMBRE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SdB':
				up=1
			else:
				i=i+1
		cSdB=i
		print('cSdB')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='LITS':
				up=1
			else:
				i=i+1
		cLITS=i
		print('cLITS')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='VILLE':
				up=1
			else:
				i=i+1
		cVILLE=i
		print('cVILLE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='lat':
				up=1
			else:
				i=i+1
		clat=i
		print('clat')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='lon':
				up=1
			else:
				i=i+1
		clon=i
		print('clon')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SUPERHOTE':
				up=1
			else:
				i=i+1
		cSUPERHOTE=i
		print('cSUPERHOTE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COMMENT_PROFIL':
				up=1
			else:
				i=i+1
		cCOMMENT_PROFIL=i
		print('cCOMMENT_PROFIL')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ID_VERIF':
				up=1
			else:
				i=i+1
		cID_VERIF=i
		print('cID_VERIF')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NB_ANNONCE':
				up=1
			else:
				i=i+1
		cNB_ANNONCE=i
		print('cNB_ANNONCE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='PROPRETE':
				up=1
			else:
				i=i+1
		cPROPRETE=i
		print('cPROPRETE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='PRECISION':
				up=1
			else:
				i=i+1
		cPRECISION=i
		print('cPRECISION')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COMMUNICATION':
				up=1
			else:
				i=i+1
		cCOMMUNICATION=i
		print('cCOMMUNICATION')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EMPLACEMENT':
				up=1
			else:
				i=i+1
		cEMPLACEMENT=i
		print('cEMPLACEMENT')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ARRIVEE':
				up=1
			else:
				i=i+1
		cARRIVEE=i
		print('cARRIVEE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='QUALITY_PRICE':
				up=1
			else:
				i=i+1
		cQUALITY_PRICE=i
		print('cQUALITY_PRICE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='REGISTER':
				up=1
			else:
				i=i+1
		cREGISTER=i
		print('cREGISTER')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='TAUX_REPONSE':
				up=1
			else:
				i=i+1
		cTAUX_REPONSE=i
		print('cTAUX_REPONSE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='DELAI_REPONSE':
				up=1
			else:
				i=i+1
		cDELAI_REPONSE=i
		print('cDELAI_REPONSE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='CHECK_IN':
				up=1
			else:
				i=i+1
		cCHECK_IN=i
		print('cCHECK_IN')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='CHECK_OUT':
				up=1
			else:
				i=i+1
		cCHECK_OUT=i
		print('cCHECK_OUT')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='FUMEUR':
				up=1
			else:
				i=i+1
		cFUMEUR=i
		print('cFUMEUR')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ENFANT':
				up=1
			else:
				i=i+1
		cENFANT=i
		print('cENFANT')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SERRURE':
				up=1
			else:
				i=i+1
		cSERRURE=i
		print('cSERRURE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ANIMAUX':
				up=1
			else:
				i=i+1
		cANIMAUX=i
		print('cANIMAUX')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='CAUTION':
				up=1
			else:
				i=i+1
		cCAUTION=i
		print('cCAUTION')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='FUMEE':
				up=1
			else:
				i=i+1
		cFUMEE=i
		print('cFUMEE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='MONOXYDE':
				up=1
			else:
				i=i+1
		cMONOXYDE=i
		print('cMONOXYDE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='DISTANCIATION_SOCIAL':
				up=1
			else:
				i=i+1
		cDISTANCIATION_SOCIAL=i
		print('cDISTANCIATION_SOCIAL')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='LANGUE':
				up=1
			else:
				i=i+1
		cLANGUE=i
		print('cLANGUE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_PROFIL':
				up=1
			else:
				i=i+1
		cIMAGE_PROFIL=i
		print('cIMAGE_PROFIL')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_1':
				up=1
			else:
				i=i+1
		cIMAGE_1=i
		print('cIMAGE_1')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_2':
				up=1
			else:
				i=i+1
		cIMAGE_2=i
		print('cIMAGE_2')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_3':
				up=1
			else:
				i=i+1
		cIMAGE_3=i
		print('cIMAGE_3')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_4':
				up=1
			else:
				i=i+1
		cIMAGE_4=i
		print('cIMAGE_4')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_5':
				up=1
			else:
				i=i+1
		cIMAGE_5=i
		print('cIMAGE_5')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_URL1':
				up=1
			else:
				i=i+1
		cCOHOTE_URL1=i
		print('cCOHOTE_URL1')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_NAME1':
				up=1
			else:
				i=i+1
		cCOHOTE_NAME1=i
		print('cCOHOTE_NAME1')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_IMAGE1':
				up=1
			else:
				i=i+1
		cCOHOTE_IMAGE1=i
		print('cCOHOTE_IMAGE1')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_URL2':
				up=1
			else:
				i=i+1
		cCOHOTE_URL2=i
		print('cCOHOTE_URL2')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_NAME2':
				up=1
			else:
				i=i+1
		cCOHOTE_NAME2=i
		print('cCOHOTE_NAME2')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_IMAGE2':
				up=1
			else:
				i=i+1
		cCOHOTE_IMAGE2=i
		print('cCOHOTE_IMAGE2')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NB_COHOTE':
				up=1
			else:
				i=i+1
		cNB_COHOTE=i
		print('cNB_COHOTE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='FETE':
				up=1
			else:
				i=i+1
		cFETE=i
		print('cFETE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ACTIVE_YES/NO':
				up=1
			else:
				i=i+1
		cACTIVE=i
		print('cACTIVE')

driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
#driver = webdriver.Chrome(chrome_options=chrome_options)
driver.set_window_size(1500, 2000)

#c = ligne 2 du xls resultant
c=2
wait2 = WebDriverWait(driver, 2)
wait3 = WebDriverWait(driver, 3)	


while c<=nrow:
	print (str(c)+'/'+str(nrow))
	h=ws.cell(row=c, column=cANNONCE).value
	numero=ws.cell(row=c, column=1).value
	#print (h)
	#do=sheet_read.cell(i,0).value
	if numero is None:
		driver.get(h)
		time.sleep(3)
		f_ele=0
		while f_ele<=3:
			try:
				#ele=driver.find_element_by_xpath("//div[@class='_1cvivhm']")
				ele=driver.find_element_by_xpath("//div[@class='_cg8a3u']")
				driver.execute_script("arguments[0].scrollIntoView(true);", ele)
				#driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
				time.sleep(3)
				driver.execute_script("window.scrollBy(0,-800);")
				#driver.execute_script("window.scrollBy(0,500);")
				f_ele=6
				time.sleep(1)
			except:
				#driver.execute_script("window.scrollBy(0,1000);")
				f_ele=f_ele+1
				time.sleep(2)
		#driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	#PROFILE
		html = driver.page_source
		time.sleep(2)
		soup = BeautifulSoup(html, 'html.parser')
		time.sleep(1)
		try:
			#GPS
			try:
				tp_c=soup.find('div', attrs={"class": "gm-style"})
				tt=tp_c.find('div', attrs={"style": "margin-left: 5px; margin-right: 5px; z-index: 1000000; position: absolute; left: 0px; bottom: 0px;"})
				tt1=tt.find('a', attrs={"target": "_blank"})
				tt2=tt1['href']
				#----------Create translation table----------
				table = str.maketrans('=&', '++')
				result_gps = tt2.translate(table)
				split_gps=result_gps.split("+")
				#https://www.google.com/maps?ll+46.23657,3.92341+z=14&t=m&hl=fr&gl=FR&mapclient=apiv3
				#https://maps.google.com/maps?ll=49.19054,-2.11715&z=14&t=m&hl=fr&gl=FR&mapclient=apiv3
				coor=split_gps[1]
				long_lat=coor.split(',')
				#--------------Write results--------------
				#print(long_lat[0])
				#print(long_lat[1])
				#point=Point(float(long_lat[1]),float(long_lat[0]))
				#question=polygon.contains(point)
				question=True
				#print(question)
				#sheet.write(c, 12, long_lat[0])
				#sheet.write(c, 13, long_lat[1])
				#ws.cell(row=c+1, column=13+1).value = long_lat[0]
				#ws.cell(row=c+1, column=14+1).value = long_lat[1]
			except:
				question=True
				print('NO GPS')
			if question is True:
				#sheet.write(c, 12, long_lat[0])
				#sheet.write(c, 13, long_lat[1])
				try:
					ws.cell(row=c, column=cANNONCE).value = h
					ws.cell(row=c, column=clat).value = long_lat[0]
					ws.cell(row=c, column=clon).value = long_lat[1]
				except:
					ee=1
			#TITLE
				try:
					div1=soup.find('div', attrs={"class": "_mbmcsn"})
					ws.cell(row=c, column=cTITLE).value = div1.h1.text
				except:
					#print('NO TITLE')
					aaa=1
			#URL HOTE
				try:
					div=soup.findAll('a', attrs={"class": "_105023be"})[-1]
					div1=div['href']  #.attrs['href']
					ws.cell(row=c, column=cHOTE).value = "https://www.airbnb.fr"+str(div1)
				except:
					#print('NO PROFILE')
					aaa=1
			#COMMENTAIRE
				COMMENT='NO COMMENT'
				#run_price=extract("//span[@class='_wfad8t']",6,COMMENT,c,YN_comment)
				try:
					p_c=[]
					tp_c=soup.findAll('span', attrs={"class": "_bq6krt"})[1].text
					p_c=tp_c.replace("(","")
					cc=p_c.replace(")","")
					try:
						pp=cc.split(' ')
						cc=pp[0]
						ws.cell(row=c, column=cCOMMENT).value = cc
					except:
						pass
					ws.cell(row=c, column=cCOMMENT).value = cc
					#print ("COMMENT ===")
					#print(cc)
					#p_c=tp_c.split("(")
					#print('ici1')
					#table_c = p_c[1].replace(")"," ")
					#print (table_c)
					#ws.cell(row=c+1, column=6+1).value = table_c
				except:
					#print('NOCOMMENT')
					aaa=1
			#VOYAGEUR
				try:
					the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
					tt=the_tr.find_all('div')[1]
					tt1=tt.find_all('span')[0]
					tt2=tt1.text
					p_tp=tt2.split(" ")
					ws.cell(row=c, column=cVOYAGEUR).value = p_tp[0]
				except:
					#print('NO VOYAGER')
					aaa=1

			#LITS
				try:
					the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
					tt=the_tr.find_all('div')[1]
					tt1=tt.find_all('span')[4]
					tt2=tt1.text
					p_tp=tt2.split(" ")
					ws.cell(row=c, column=cLITS).value = p_tp[0]
				except:
					#print('NO LIT')
					aaa=1
			#SdB
				try:
					the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
					tt=the_tr.find_all('div')[1]
					tt1=tt.find_all('span')[6]
					tt2=tt1.text
					p_tp=tt2.split(" ")
					ws.cell(row=c, column=cSdB).value = p_tp[0]
				except:
					#print('NO SdB')
					aaa=1
			#CHAMBRE
				try:
					the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
					tt=the_tr.find_all('div')[1]
					tt1=tt.find_all('span')[2]
					tt2=tt1.text
					p_tp=tt2.split(" ")
					ws.cell(row=c, column=cCHAMBRE).value = p_tp[0]
				except:
					#print('NO CHAMBRE')
					aaa=1
			#VILLE
				try:
					tp_c=soup.find('a', attrs={"class": "_5twioja"}).text
					ws.cell(row=c, column=cVILLE).value = tp_c
				except:
					#print('NO VILLE')
					aaa=1


			#NAME_HOTE
				try:
					tp_c=soup.find('div', attrs={"class": "_f47qa6"})
					tt=tp_c.find('div', attrs={"class": "_svr7sj"})
					tt1=tt.h2.get_text()
					pp=tt1.split('par ')
					ws.cell(row=c, column=cNAME_HOTE).value = pp[1]
				except:
					#print ('NO_NAME')
					aaa=1
			#TYPE_HOME
				try:
					the_tr= soup.find('div', attrs={"class": "_1b3ij9t"}).text
					pp=the_tr.split('.')
					ws.cell(row=c, column=cTYPE_LOGEMENT).value = pp[0]
				except:
					try:
						the_tr= soup.find('div', attrs={"class": "_xcsyj0"}).text
						pp=the_tr.split('.')
						ws.cell(row=c, column=cTYPE_LOGEMENT).value = pp[0]
					except:
						#print('NOTYPE')
						aaa=1

			#ANCIENNETE
				try:
					tp_c=soup.find('div', attrs={"class": "_f47qa6"})
					tt=tp_c.find('div', attrs={"class": "_svr7sj"})
					tt1=tt.div.get_text()
					ws.cell(row=c, column=cANCIENNETE).value = tt1
				except:
					#print ('NOOLD')
					aaa=1

			#SUPER HOTE
				try:
					#the_tr= soup.find('span', text=re.compile(r'\bSuperhost\b'),attrs = {'aria-hidden' : 'false'})
					tp_c=soup.find('div', attrs={"class": "_1ft6jxp"}).text
					ws.cell(row=c, column=cSUPERHOTE).value = 'X'
				except:
					#print('no superhote')
					aaa=1
			#COMMENT PROFIL
				try:
					the_tr= soup.findAll('span', attrs = {'class' : '_pog3hg'})[0]
					ccc=the_tr.text
					pp=ccc.split('c')
					cc=pp[0]
					#div2=the_tr.findNextSibling('div')
					#print(the_tr.section.span.div.span.text)
					if cc=='Identité':
						cc=0
					ws.cell(row=c, column=cCOMMENT_PROFIL).value = cc
					#print(div2.text)
				except:
					#print('No Comment profil')
					aaa=1
			#IDENTIFIE CHECK
				try:
					the_tr= soup.find('span', text=re.compile(r"\bIdentité vérifiée\b"))
					ws.cell(row=c, column=cID_VERIF).value = 'YES'
					#print(div2.text)
				except:
					ws.cell(row=c, column=cID_VERIF).value = 'NO'
					#print('No CHECK ID')
					aaa=1
			#CO HOTE
				try:
					the_tr= soup.find('ul', attrs = {'class' : '_1omtyzc'})
					the_tr1= the_tr.findAll('li', attrs = {'class' : '_108byt5'})[0]
					tt11= the_tr1.find('a', attrs = {'target' : '_blank'})
					tt12= the_tr1.find('span', attrs = {'class' : '_1kfl0pr'})
					tt13= the_tr1.find('img', attrs = {'class' : '_6tbg2q'})
					div1=tt11['href']  #.attrs['href']
					the_tr2= the_tr.findAll('li', attrs = {'class' : '_108byt5'})[1]
					tt21= the_tr2.find('a', attrs = {'target' : '_blank'})
					tt22= the_tr2.find('span', attrs = {'class' : '_1kfl0pr'})
					tt23= the_tr2.find('img', attrs = {'class' : '_6tbg2q'})
					div2=tt21['href']  #.attrs['href']
					ws.cell(row=c, column=cCOHOTE_URL1).value = "https://www.airbnb.fr"+str(div1)
					ws.cell(row=c, column=cCOHOTE_NAME1).value = tt12.text
					ws.cell(row=c, column=cCOHOTE_IMAGE1).value = tt13['src']
					ws.cell(row=c, column=cCOHOTE_URL2).value = "https://www.airbnb.fr"+str(div2)
					ws.cell(row=c, column=cCOHOTE_NAME2).value = tt22.text
					ws.cell(row=c, column=cCOHOTE_IMAGE2).value = tt23['src']
					ws.cell(row=c, column=cNB_COHOTE).value = 2
				except:
					try:
						the_tr= soup.find('ul', attrs = {'class' : '_1omtyzc'})
						the_tr1= the_tr.find('li', attrs = {'class' : '_108byt5'})
						tt= the_tr1.find('a', attrs = {'target' : '_blank'})
						tt2= the_tr1.find('span', attrs = {'class' : '_1kfl0pr'})
						tt3= the_tr1.find('img', attrs = {'class' : '_6tbg2q'})
						div1=tt['href']  #.attrs['href']
						ws.cell(row=c, column=cCOHOTE_URL1).value = "https://www.airbnb.fr"+str(div1)
						ws.cell(row=c, column=cCOHOTE_NAME1).value = tt2.text
						ws.cell(row=c, column=cCOHOTE_IMAGE1).value = tt3['src']
						ws.cell(row=c, column=cNB_COHOTE).value = 1
					except:
						ws.cell(row=c, column=cNB_COHOTE).value = 0
						#print('no co hote')
						aaa=1
			#PROPRETE
				try:
					tt= soup.findAll('span', attrs={"class": "_4oybiu"})[0]
					#print(tt.text)
					ws.cell(row=c, column=cPROPRETE).value = tt.text
				except:
					#print('no proprete')
					aaa=1
			#PRECISION
				try:
					tt= soup.findAll('span', attrs={"class": "_4oybiu"})[1]
					#print(tt.text)
					ws.cell(row=c, column=cPRECISION).value = tt.text
				except:
					#print('no Precision')
					aaa=1
			#COMMUNICATION
				try:
					tt= soup.findAll('span', attrs={"class": "_4oybiu"})[2]
					#print(tt.text)
					ws.cell(row=c, column=cCOMMUNICATION).value = tt.text
				except:
					#print('no communication')
					aaa=1
			#EMPLACEMENT
				try:
					tt= soup.findAll('span', attrs={"class": "_4oybiu"})[3]
					#print(tt.text)
					ws.cell(row=c, column=cEMPLACEMENT).value = tt.text
				except:
					#print('no emplacement')
					aaa=1
			#ARRIVEE
				try:
					tt= soup.findAll('span', attrs={"class": "_4oybiu"})[4]
					#print(tt.text)
					ws.cell(row=c, column=cARRIVEE).value = tt.text
				except:
					#print('no arrivee')
					aaa=1
			#QUALITY PRICE
				try:
					tt= soup.findAll('span', attrs={"class": "_4oybiu"})[5]
					#print(tt.text)
					ws.cell(row=c, column=cQUALITY_PRICE).value = tt.text
				except:
					#print('no price quality')
					aaa=1
		#N° ENREGISTREMENT
				try:
					the_tr= soup.find('li', text=re.compile(r'\bNuméro\b'), attrs = {'class' : '_1q2lt74'})
					pp=the_tr.text
					#print(pp)
					sp=pp.split(' ')
					ws.cell(row=c, column=cREGISTER).value = sp[-1]
				except:
					a=1
					#print('no N° enregistrement')
					aaa=1
		#TAUX REPONSE
				try:
					the_tr=soup.find('li', text=re.compile(r'\bTaux\b'))
					pp=the_tr.text
					pp=pp.replace(" ","")
					#print(pp)
					sp=pp.split(':')
					#print(sp[-1])
					ws.cell(row=c, column=cTAUX_REPONSE).value = sp[-1]
				except:
					#print('no taux réponse')
					aaa=1
		#DELAI REPONSE
				try:
					the_tr=soup.find('li', text=re.compile(r'\bDélai\b'))
					pp=the_tr.text
					sp=pp.split(':')
					ws.cell(row=c, column=cDELAI_REPONSE).value = sp[-1]
				except:
					#print('no DELAI REPONSE')
					aaa=1
		#DURING SEJOUR
				try:
					the_tr=soup.findAll('div', attrs={"class": "_1byskwn"})[-1]
					try:
						tt= the_tr.find('span', text=re.compile(r'\bArrivée\b'))
						ws.cell(row=c, column=cCHECK_IN).value = tt.text
					except:
						#print('no ARRIVE')
						aaa=1
					try:
						tt= the_tr.find('span', text=re.compile(r'\bDépart\b'))
						ws.cell(row=c, column=cCHECK_OUT).value = tt.text
					except:
						#print('no DEPART')
						aaa=1
					try:
						tt= the_tr.find('span', text=re.compile(r'\bNon fumeur\b'))
						ws.cell(row=c, column=cFUMEUR).value = tt.text
					except:
						#print('no FUMEUR')
						aaa=1
					try:
						tt= the_tr.find('span', text=re.compile(r'\bNe convient pas aux\b'))
						ws.cell(row=c, column=cENFANT).value = tt.text
					except:
						#print('no CHILD')
						aaa=1
					try:
						tt= the_tr.find('span', text=re.compile(r"\bArrivée autonome\b"))
						ws.cell(row=c, column=cSERRURE).value = tt.text
					except:
						#print('no AUTOMATIC')
						aaa=1
					try:
						tt= the_tr.find('span', text=re.compile(r"\bPas d'animaux\b"))
						ws.cell(row=c, column=cANIMAUX).value = tt.text
					except:
						try:
							tt= the_tr.find('span', text=re.compile(r"\bAnimaux de compagnie\b"))
							ws.cell(row=c, column=cANIMAUX).value = tt.text
						except:
							#print('no ANIMAL')
							aaa=1
					try:
						tt= the_tr.find('span', text=re.compile(r"\bCaution\b"))
						ws.cell(row=c, column=cCAUTION).value = tt.text
					except:
						#print('no Caution')
						aaa=1
					try:
						tt= the_tr.find('span', text=re.compile(r"\bDétecteur de fumée\b"))
						ws.cell(row=c, column=cFUMEE).value = tt.text
					except:
						#print('no detecteur fumee')
						aaa=1
					try:
						tt= the_tr.find('span', text=re.compile(r"\bDétecteur de monoxyde de carbone\b"))
						ws.cell(row=c, column=cMONOXYDE).value = tt.text
					except:
						#print('no detecteur monoxyde')
						aaa=1
					try:
						tt= the_tr.find('span', text=re.compile(r"\bPas de fête ni de soirée\b"))
						ws.cell(row=c, column=cFETE).value = tt.text
					except:
						#print('no detecteur monoxyde')
						aaa=1
					try:
						tt= the_tr.find('span', text=re.compile(r"\bmatière de distanciation sociale\b"))
						ws.cell(row=c, column=cDISTANCIATION_SOCIAL).value = 'Y'
					except:
						ws.cell(row=c, column=cDISTANCIATION_SOCIAL).value = 'N'
						#print('no distanciation sociale')
				except:
					#print('no INSIDE RULE')
					aaa=1
		#LANGUE
				try:
					the_tr= soup.find('li', text=re.compile(r'\bLangues\b'))
					#print(the_tr)
					pp=the_tr.text
					#print(pp)
					sp=pp.split(':')
					ws.cell(row=c, column=cLANGUE).value = sp[-1]
				except:
					#print('no LANGUAGE')
					aaa=1
		#IMAGE
				try:
					the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[0]
					#print(the_tr)
					tt=the_tr['src']
					ws.cell(row=c, column=cIMAGE_1).value = tt
				except:
					try:
						the_tr= soup.find('img', attrs={"class": "_6tbg2q"})
						#print(the_tr)
						tt=the_tr['data-original-uri']
						ws.cell(row=c, column=cIMAGE_1).value = tt
					except:
						#print('no IMAGE 0')
						aaa=1
				try:
					the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[1]
					#print(the_tr)
					tt=the_tr['src']
					ws.cell(row=c, column=cIMAGE_2).value = tt
				except:
					try:
						the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[1]
						#print(the_tr)
						tt=the_tr['data-original-uri']
						ws.cell(row=c, column=cIMAGE_2).value = tt
					except:
						#print('no IMAGE 1')
						aaa=1
				try:
					the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[2]
					#print(the_tr)
					tt=the_tr['src']
					ws.cell(row=c, column=cIMAGE_3).value = tt
				except:
					try:
						the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[2]
						#print(the_tr)
						tt=the_tr['data-original-uri']
						ws.cell(row=c, column=cIMAGE_3).value = tt
					except:
						#print('no IMAGE 2')
						aaa=1
				try:
					the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[3]
					#print(the_tr)
					tt=the_tr['src']
					ws.cell(row=c, column=cIMAGE_4).value = tt
				except:
					try:
						the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[3]
						#print(the_tr)
						tt=the_tr['data-original-uri']
						ws.cell(row=c, column=cIMAGE_4).value = tt
					except:
						#print('no IMAGE 3')
						aaa=1
				try:
					the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[4]
					#print(the_tr)
					tt=the_tr['src']
					#print(tt)
					ws.cell(row=c, column=cIMAGE_5).value = tt
				except:
					try:
						the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[4]
						#print(the_tr)
						tt=the_tr['data-original-uri']
						ws.cell(row=c, column=cIMAGE_5).value = tt
					except:
						#print('no IMAGE 4')
						aaa=1
		#IMAGE_HOTE
				try:
					the_tr= soup.find('div', attrs={"class": "_5kripx"})
					t= the_tr.find('img', attrs={"class": "_6tbg2q"})
					tt=t['src']
					ws.cell(row=c, column=cIMAGE_PROFIL).value = tt
				except:
					#print('no IMAGE_HOTE')
					aaa=1
				ws.cell(row=c, column=cACTIVE).value = 'YES'
				if (c/200).is_integer():
					wb.save(path_RESULT.filename)
					time.sleep(5)
				if (c/2000).is_integer():
					driver.quit()
					time.sleep(5)
					driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
					#driver = webdriver.Chrome(chrome_options=chrome_options)
					driver.set_window_size(1500, 2000)
					wait = WebDriverWait(driver, 3)
					wait2 = WebDriverWait(driver, 2)
					wait3 = WebDriverWait(driver, 3)
					time.sleep(5)
					driver.get(h)
					time.sleep(5)
					wb.save(DIR2+NAMEFile+str(c)+".xlsx")
#------------------------
		except:
			try:
				driver.quit()
				time.sleep(5)
			except:
				time.sleep(5)
			driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
			#driver = webdriver.Chrome(chrome_options=chrome_options)
			driver.set_window_size(1500, 2000)
			wait = WebDriverWait(driver, 3)
	c=c+1
wb.save(path_RESULT.filename)
print ('_______    ___    ___     ___')
print ('|      |   |  |   |  \    |  |')
print ('|  |__     |  |   |   \   |  |')
print ('|     |    |  |   |    \  |  |')
print ('|  |       |  |   |  |\ \ |  |')
print ('|  |       |  |   |  | \ \|  |')
print ('|__|       |__|   |__|  \____|')
try:
	driver.quit()
except:
	print('fin')
