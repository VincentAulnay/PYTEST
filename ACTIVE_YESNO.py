import xlwt
import xlrd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from xlutils.copy import copy
from decimal import Decimal
from decimal import *
import decimal
import datetime
from datetime import date
import datetime as dt
from tkinter import filedialog
from tkinter import *
import os
import re
import json
from urllib.request import urlopen
import pip
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import datetime
import datetime as dt
import threading
import sys

Deccontext = Context(prec=10, rounding=ROUND_HALF_DOWN)
setcontext(Deccontext)

chrome_options = webdriver.ChromeOptions()
prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)

#on récupère la value dans le excel ORIGINE
#32460


wb = load_workbook(path_RESULT.filename)
ws=wb.active
nrow=ws.max_row


searchcolumn=1
if searchcolumn==1:
		up=0
		i=1
		#while up==0:
		#	V_up=ws.cell(row=1, column=i).value
		#	if V_up=='NUMERO':
		#		up=1
		#	else:
		#		i=i+1
		#cNUMERO=i
		print('cNUMERO')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='TITRE':
				up=1
			else:
				i=i+1
		cTITLE=i
		print('cTITLE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ANNONCE':
				up=1
			else:
				i=i+1
		cANNONCE=i
		print('cANNONCE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NAME_HOTE':
				up=1
			else:
				i=i+1
		cNAME_HOTE=i
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ANCIENNETE':
				up=1
			else:
				i=i+1
		cANCIENNETE=i
		print('cANCIENNETE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='HOTE':
				up=1
			else:
				i=i+1
		cHOTE=i
		print('cHOTE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='PRICE':
				up=1
			else:
				i=i+1
		cPRICE=i
		print('cPRICE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COMMENTAIRE':
				up=1
			else:
				i=i+1
		cCOMMENT=i
		print('cCOMMENTAIRE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='TYPE_LOGEMENT':
				up=1
			else:
				i=i+1
		cTYPE_LOGEMENT=i
		print('cTYPE_LOGEMENT')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='VOYAGEUR':
				up=1
			else:
				i=i+1
		cVOYAGEUR=i
		print('cVOYAGEUR')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='CHAMBRE':
				up=1
			else:
				i=i+1
		cCHAMBRE=i
		print('cCHAMBRE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SdB':
				up=1
			else:
				i=i+1
		cSdB=i
		print('cSdB')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='LITS':
				up=1
			else:
				i=i+1
		cLITS=i
		print('cLITS')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='VILLE':
				up=1
			else:
				i=i+1
		cVILLE=i
		print('cVILLE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='lat':
				up=1
			else:
				i=i+1
		clat=i
		print('clat')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='lon':
				up=1
			else:
				i=i+1
		clon=i
		print('clon')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SUPERHOTE':
				up=1
			else:
				i=i+1
		cSUPERHOTE=i
		print('cSUPERHOTE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COMMENT_PROFIL':
				up=1
			else:
				i=i+1
		cCOMMENT_PROFIL=i
		print('cCOMMENT_PROFIL')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ID_VERIF':
				up=1
			else:
				i=i+1
		cID_VERIF=i
		print('cID_VERIF')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NB_ANNONCE':
				up=1
			else:
				i=i+1
		cNB_ANNONCE=i
		print('cNB_ANNONCE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='PROPRETE':
				up=1
			else:
				i=i+1
		cPROPRETE=i
		print('cPROPRETE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='PRECISION':
				up=1
			else:
				i=i+1
		cPRECISION=i
		print('cPRECISION')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COMMUNICATION':
				up=1
			else:
				i=i+1
		cCOMMUNICATION=i
		print('cCOMMUNICATION')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EMPLACEMENT':
				up=1
			else:
				i=i+1
		cEMPLACEMENT=i
		print('cEMPLACEMENT')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ARRIVEE':
				up=1
			else:
				i=i+1
		cARRIVEE=i
		print('cARRIVEE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='QUALITY_PRICE':
				up=1
			else:
				i=i+1
		cQUALITY_PRICE=i
		print('cQUALITY_PRICE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='REGISTER':
				up=1
			else:
				i=i+1
		cREGISTER=i
		print('cREGISTER')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='TAUX_REPONSE':
				up=1
			else:
				i=i+1
		cTAUX_REPONSE=i
		print('cTAUX_REPONSE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='DELAI_REPONSE':
				up=1
			else:
				i=i+1
		cDELAI_REPONSE=i
		print('cDELAI_REPONSE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='CHECK_IN':
				up=1
			else:
				i=i+1
		cCHECK_IN=i
		print('cCHECK_IN')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='CHECK_OUT':
				up=1
			else:
				i=i+1
		cCHECK_OUT=i
		print('cCHECK_OUT')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='FUMEUR':
				up=1
			else:
				i=i+1
		cFUMEUR=i
		print('cFUMEUR')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ENFANT':
				up=1
			else:
				i=i+1
		cENFANT=i
		print('cENFANT')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SERRURE':
				up=1
			else:
				i=i+1
		cSERRURE=i
		print('cSERRURE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ANIMAUX':
				up=1
			else:
				i=i+1
		cANIMAUX=i
		print('cANIMAUX')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='CAUTION':
				up=1
			else:
				i=i+1
		cCAUTION=i
		print('cCAUTION')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='FUMEE':
				up=1
			else:
				i=i+1
		cFUMEE=i
		print('cFUMEE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='MONOXYDE':
				up=1
			else:
				i=i+1
		cMONOXYDE=i
		print('cMONOXYDE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='DISTANCIATION_SOCIAL':
				up=1
			else:
				i=i+1
		cDISTANCIATION_SOCIAL=i
		print('cDISTANCIATION_SOCIAL')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='LANGUE':
				up=1
			else:
				i=i+1
		cLANGUE=i
		print('cLANGUE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_PROFIL':
				up=1
			else:
				i=i+1
		cIMAGE_PROFIL=i
		print('cIMAGE_PROFIL')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_1':
				up=1
			else:
				i=i+1
		cIMAGE_1=i
		print('cIMAGE_1')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_2':
				up=1
			else:
				i=i+1
		cIMAGE_2=i
		print('cIMAGE_2')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_3':
				up=1
			else:
				i=i+1
		cIMAGE_3=i
		print('cIMAGE_3')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_4':
				up=1
			else:
				i=i+1
		cIMAGE_4=i
		print('cIMAGE_4')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='IMAGE_5':
				up=1
			else:
				i=i+1
		cIMAGE_5=i
		print('cIMAGE_5')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_URL1':
				up=1
			else:
				i=i+1
		cCOHOTE_URL1=i
		print('cCOHOTE_URL1')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_NAME1':
				up=1
			else:
				i=i+1
		cCOHOTE_NAME1=i
		print('cCOHOTE_NAME1')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_IMAGE1':
				up=1
			else:
				i=i+1
		cCOHOTE_IMAGE1=i
		print('cCOHOTE_IMAGE1')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_URL2':
				up=1
			else:
				i=i+1
		cCOHOTE_URL2=i
		print('cCOHOTE_URL2')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_NAME2':
				up=1
			else:
				i=i+1
		cCOHOTE_NAME2=i
		print('cCOHOTE_NAME2')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='COHOTE_IMAGE2':
				up=1
			else:
				i=i+1
		cCOHOTE_IMAGE2=i
		print('cCOHOTE_IMAGE2')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NB_COHOTE':
				up=1
			else:
				i=i+1
		cNB_COHOTE=i
		print('cNB_COHOTE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='FETE':
				up=1
			else:
				i=i+1
		cFETE=i
		print('cFETE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ACTIVE_YES/NO':
				up=1
			else:
				i=i+1
		cACTIVE=i
		print('cACTIVE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ENTREPRISE':
				up=1
			else:
				i=i+1
		cENTREPRISE=i
		print('cENTREPRISE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SOCIALWASHING':
				up=1
			else:
				i=i+1
		cSOCIALWASHING=i
		print('cSOCIALWASHING')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='ADRESSE':
				up=1
			else:
				i=i+1
		cADRESS=i
		print('cADRESS')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_SALLEDEBAIN':
				up=1
			else:
				i=i+1
		cEQU_SALLEDEBAIN=i
		print('cEQU_SALLEDEBAIN')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_CHAMBRELINGE':
				up=1
			else:
				i=i+1
		cEQU_CHAMBRELINGE=i
		print('cEQU_CHAMBRELINGE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_DIVERTISSEMENT':
				up=1
			else:
				i=i+1
		cEQU_DIVERTISSEMENT=i
		print('cEQU_DIVERTISSEMENT')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_FAMILLE':
				up=1
			else:
				i=i+1
		cEQU_FAMILLE=i
		print('cEQU_FAMILLE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_CHAUFFAGECLIME':
				up=1
			else:
				i=i+1
		cEQU_CHAUFFAGECLIME=i
		print('cEQU_CHAUFFAGECLIME')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_SECURITE':
				up=1
			else:
				i=i+1
		cEQU_SECURITE=i
		print('cEQU_SECURITE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_INTERNETBUREAU':
				up=1
			else:
				i=i+1
		cEQU_INTERNETBUREAU=i
		print('cEQU_INTERNETBUREAU')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_CUISINE':
				up=1
			else:
				i=i+1
		cEQU_CUISINE=i
		print('cEQU_CUISINE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_PARKINGINSTALLATION':
				up=1
			else:
				i=i+1
		cEQU_PARKINGINSTALLATION=i
		print('cEQU_PARKINGINSTALLATION')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_CARACTEMPLACEMENT':
				up=1
			else:
				i=i+1
		cEQU_CARACTEMPLACEMENT=i
		print('cEQU_CARACTEMPLACEMENT')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_EXTERIEUR':
				up=1
			else:
				i=i+1
		cEQU_EXTERIEUR=i
		print('cEQU_EXTERIEUR')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_SERVICE':
				up=1
			else:
				i=i+1
		cEQU_SERVICE=i
		print('cEQU_SERVICE')
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='EQU_NONINCLUS':
				up=1
			else:
				i=i+1
		cEQU_NONINCLUS=i
		print('cEQU_NONINCLUS')

		
driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
#driver = webdriver.Chrome(chrome_options=chrome_options)
driver.set_window_size(1500, 2000)


c=2
while c<=nrow:
	print (str(c)+'/'+str(nrow))
	h=ws.cell(row=c, column=cANNONCE).value
	#print (h)
	#do=sheet_read.cell(i,0).value
	driver.get(h)
	time.sleep(2)
	try:
		html = driver.page_source
		soup = BeautifulSoup(html, 'html.parser')
    time.sleep(1)
		the_tr= soup.findAll('td', attrs={'aria-label':re.compile(r'\bnon\b')})[1]
		print('yes')
		ws.cell(row=c, column=cACTIVE).value="YES"

	except:
		ws.cell(row=c, column=cACTIVE).value="NO"
		pass

	c=c+1
wb.save(path_RESULT.filename)
print ('_______    ___    ___     ___')
print ('|      |   |  |   |  \    |  |')
print ('|  |__     |  |   |   \   |  |')
print ('|     |    |  |   |    \  |  |')
print ('|  |       |  |   |  |\ \ |  |')
print ('|  |       |  |   |  | \ \|  |')
print ('|__|       |__|   |__|  \____|')
try:
	driver.quit()
except:
	print('fin')

