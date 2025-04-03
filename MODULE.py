print('start')
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import xlwt
import xlrd
import time
from xlutils.copy import copy
import datetime
import datetime as dt
from tkinter import filedialog
from tkinter import *
from bs4 import BeautifulSoup
from urllib.request import urlopen
import os
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
import threading
import sys



chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("window-size=2000,1000")
#chrome_options.add_argument("-headless")
#chrome_options.add_argument("-disable-gpu")

chrome_options_description = webdriver.ChromeOptions()
prefs_description = {"profile.managed_default_content_settings.images": 2}
chrome_options_description.add_experimental_option("prefs", prefs_description)

print ('▀▄▀▄▀▄ STOPBNB ▄▀▄▀▄▀')

now = str(datetime.datetime.now())[:19]
now = now.replace(":","_")
print(now)
print('tes')
#-----EXCEL RESULT OPEN AND READ-----

import re
import json
import csv
from google.oauth2 import service_account
import pygsheets
import pandas as pd

#wbx = load_workbook(path_RESULT.filename)
#ws = wbx.active
print('ici1')
if name_rpi=='Rpi3':
	client = pygsheets.authorize(service_account_file='/home/vincent/Desktop/rpi1-378418-06f4d82571c9.json')
elif name_rpi=='Rpi1':
	client = pygsheets.authorize(service_account_file='/home/vincent/Desktop/rpi1-378709-ede595e84cf3.json')
elif name_rpi=='Rpi2':
	client = pygsheets.authorize(service_account_file='/home/vincent/Desktop/raspbian-364809-be26e1ee6573.json')
elif name_rpi=='Rpi4':
	client = pygsheets.authorize(service_account_file='/home/vincent/Desktop/rpi4-404313-beba1f1e0e8f.json')
#try:
#	client = pygsheets.authorize(service_account_file='/home/vincent/Desktop/raspbian-364809-be26e1ee6573.json')
#except:
#	client = pygsheets.authorize(service_account_file='/home/pi/Desktop/raspbian-364809-be26e1ee6573.json')
print('ici1')
#spreadsheet_url = "https://docs.google.com/spreadsheets/d/1vx34zctZXc2eQSFFe4I7zY6bjKJz9MtO7pgAIaQix4c/edit?usp=sharing"
#spreadsheet_url = "https://docs.google.com/spreadsheets/d/1ACSlRUHdqn9ExIM2M-18VGHBoo8RaxAfxkbvKCd1ylw/edit?usp=sharing"
#spreadsheet_url = "https://docs.google.com/spreadsheets/d/1QJS5Vl_V6b-Tah8BpuWoUttKd1xRh-MLk8306qbXjJg/edit?usp=sharing"

#1foRAOdxPydwyz5ju4nGsPw-L9196j8MwnFzisDgZJmo
#1D9V6zS87cjcMJUFDVqFi1IktqbOR7NyEV2q3a0DtFYM
#1kAYBZN4NNASkL24DBaAEav-21xo1ii3o0CSchS6qN0s
#1GauTmUNmCiv3B-psbn8oZvn23A6rh5GNB6L600VM3rA
#1si2nENFfVIx73f-o00E0rrv0t4dyrn92fKoikutyOIo

#url_rpi="https://docs.google.com/spreadsheets/d/14fiETLENGjJU3LMIybT-LAKw2DgXnol8ZAHgn46FkPs/edit?usp=sharing"
#id_rpi="14fiETLENGjJU3LMIybT-LAKw2DgXnol8ZAHgn46FkPs"
#name_rpi="testoct"

spreadsheet_url = url_rpi

print('ici2')	


#sheet_data = client.sheet.get('1QJS5Vl_V6b-Tah8BpuWoUttKd1xRh-MLk8306qbXjJg')
sheet_data = client.sheet.get(id_rpi)

print('ici3')
#sheet = client.open('MODULE')
sheet = client.open(name_rpi)
print('ici4')
ws = sheet.worksheet_by_title('Sheet1')
print('sheet')
#print(ws)
#-------FIND COLUMN UPDATE------
cTITLE=2
cANNONCE=3
cNAME_HOTE=4
cANCIENNETE=5
cHOTE=6
cPRICE=7
cCOMMENT=8
cTYPE_LOGEMENT=9
cVOYAGEUR=10
cCHAMBRE=11
cSdB=12
cLITS=13
CODE_INSEE=14
CODE_POSTALE=15
Commune=16
Departement=17
Region=18
CIRCONSCRIPTION=19
cVILLE=20
clat=21
clon=22
cSUPERHOTE=23
cCOMMENT_PROFIL=24
cNB_ANNONCE=25
cREGISTER=26
cCHECK_IN=27
cCHECK_OUT=28
cENFANT=29
cSERRURE=30
cANIMAUX=31
cCAUTION=32
cIMAGE_PROFIL=33
cIMAGE_1=34
cIMAGE_2=35
cIMAGE_3=36
cIMAGE_4=37
cIMAGE_5=38
cCOHOTE_URL1=39
cCOHOTE_NAME1=40
cCOHOTE_IMAGE1=41
cNB_COHOTE=42
cCOHOTE_URL2=43
cCOHOTE_NAME2=44
cCOHOTE_IMAGE2=45
cENTREPRISE=46
cACTIVE=47


driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
driver.set_window_size(2000, 1000)
driver.get('https://www.google.com/')
driver.get('chrome://settings/')
driver.execute_script('chrome.settingsPrivate.setDefaultZoom(0.5);')
driver.implicitly_wait(10)
wait = WebDriverWait(driver, 5)

driver_description = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver', options=chrome_options_description)
#driver = webdriver.Chrome(options=chrome_options)
driver_description.set_window_size(200, 400)

c=2
wait2 = WebDriverWait(driver, 5)
wait3 = WebDriverWait(driver, 5)
wait = WebDriverWait(driver, 5)
wait_description = WebDriverWait(driver, 3)
def scrap(h):
	driver.get(h)
	#time.sleep(2)
def scrap_description(h,c):
	driver_description.get(h+'?modal=DESCRIPTION')
	time.sleep(8)
	#try:
	#	clos_translate = wait_description.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Fermer']")))
	#	clos_translate.click()
	#except:
	#	aaa=1
	
	html_description = driver_description.page_source
	soup_description = BeautifulSoup(html_description, 'html.parser')
	time.sleep(1)

	try:
		tp_c=soup_description.find_all('div', attrs={"class": "_gt7myn"})[-1].h2
		if tp_c.text == "Numéro d'enregistrement":
			h2tag=tp_c.parent
			divtag=h2tag.parent
			value_nc=divtag.span
			ws.cell((c, cREGISTER)).value = value_nc.text
	except:
		aaa=1
def GSwrite(c):
	print('=======test écriture title')
	ws.cell((c, 2)).value = 'test'
	print('=======print ok')

def SCRAP_detail(c):
	cTITLE=2
	cANNONCE=3
	cNAME_HOTE=4
	cANCIENNETE=5
	cHOTE=6
	cPRICE=7
	cCOMMENT=8
	cTYPE_LOGEMENT=9
	cVOYAGEUR=10
	cCHAMBRE=11
	cSdB=12
	cLITS=13
	CODE_INSEE=14
	CODE_POSTALE=15
	Commune=16
	Departement=17
	Region=18
	CIRCONSCRIPTION=19
	cVILLE=20
	clat=21
	clon=22
	cSUPERHOTE=23
	cCOMMENT_PROFIL=24
	cNB_ANNONCE=25
	cREGISTER=26
	cCHECK_IN=27
	cCHECK_OUT=28
	cENFANT=29
	cSERRURE=30
	cANIMAUX=31
	cCAUTION=32
	cIMAGE_PROFIL=33
	cIMAGE_1=34
	cIMAGE_2=35
	cIMAGE_3=36
	cIMAGE_4=37
	cIMAGE_5=38
	cCOHOTE_URL1=39
	cCOHOTE_NAME1=40
	cCOHOTE_IMAGE1=41
	cNB_COHOTE=42
	cCOHOTE_URL2=43
	cCOHOTE_NAME2=44
	cCOHOTE_IMAGE2=45
	cENTREPRISE=46
	cACTIVE=47
	try:
	    FTitle = soup.find('div', attrs={"data-plugin-in-point-id": "TITLE_DEFAULT"})
	    Flogement = soup.find('div', attrs={"data-plugin-in-point-id": "OVERVIEW_DEFAULT_V2"})
	    FProfile = soup.find('div', attrs={"data-plugin-in-point-id": "HOST_PROFILE_DEFAULT"})
	except:
	    aaa=1
	try:
	    FTitle = soup.find('div', attrs={"data-plugin-in-point-id": "TITLE_DEFAULT"})
	    Flogement = soup.find('div', attrs={"data-plugin-in-point-id": "OVERVIEW_DEFAULT_V2"})
	    FProfile = soup.find('div', attrs={"data-plugin-in-point-id": "MEET_YOUR_HOST"})
	    FPolicies = soup.find('div', attrs={"data-plugin-in-point-id": "POLICIES_DEFAULT"})
	except:
	    aaa=1
	try:
	    FProfile = soup.find('div', attrs={"data-plugin-in-point-id": "MEET_YOUR_HOST"})
	    Vprofil=1
	except:
	    try:
	        FProfile = soup.find('div', attrs={"data-plugin-in-point-id": "MEET_YOUR_HOST"})
	        Vprofil=2
	    except:
	        aaa=1
	try:
		FHero = soup.find('div', attrs={"data-plugin-in-point-id": "HERO_DEFAULT"})
	except:
		aaa=1
	#print('start bs4')
	try:
	    #GPS
	    try:
	        tp_c=soup.find('div', attrs={"class": "gm-style"})
	        tt=tp_c.find('div', attrs={"style": "margin: 0px 5px; z-index: 1000000; position: absolute; left: 0px; bottom: 0px;"})
	        tt1=tt.find('a', attrs={"target": "_blank"})
	        tt2=tt1['href']
	        #----------Create translation table----------
	        table = str.maketrans('=&', '++')
	        result_gps = tt2.translate(table)
	        split_gps=result_gps.split("+")
	        coor=split_gps[1]
	        long_lat=coor.split(',')
	        question=True
	    except:
	        question=False
	        print('NO GPS')
	    if question is True:
	        try:
	            ws.cell((c, clat)).value = long_lat[0]
	            ws.cell((c, clon)).value = long_lat[1]
	        except:
	            ee=1
	    #TITLE
	        try:
	            div1=FTitle.find('div', attrs={"_1czgyoo"})
	            ws.cell((c, cTITLE)).value = div1.h1.text
	        except:
	            #print('NO TITLE')
	            aaa=1
	    #TYPE_HOME
	        try:
	            the_tr= Flogement.find('h2')
	            ttt=the_tr.text
	            type_home=ttt.split(' - ')
	            #print(pp)
	            ws.cell((c, cTYPE_LOGEMENT)).value=type_home[0]
	        except:
	            try:
	                the_tr= soup.find('div', attrs={"class": "toieuka dir dir-ltr"})
	                ttt=the_tr.find('h1').text
	                type_home=ttt.split(' - ')
	                ws.cell((c, cTYPE_LOGEMENT)).value=type_home[0]
	            except:
	                #print('NOTYPE')
	                aaa=1
	    #URL HOTE
	        try:
	            if Vprofil==1:
	                div=FProfile.find('a')
	                div1=div['href']
	            elif Vprofil==2:
	                div=FProfile.find_all('a')[0]
	                div1=div['href']
	            #div2=div.find('a')
	            ws.cell((c, cHOTE)).value = "https://www.airbnb.fr"+str(div1)
	            #print("URLHOT1"+str(div1))
	        except:
	            print('NO PROFILE')
	        if "Chambre" in type_home[0]:
	            try:
	                div=soup.find('div', attrs={'class': 'c1u4hpjh dir dir-ltr'})
	                div2=div.find('a')
	                div1=div2['href']
	                ws.cell((c, cHOTE)).value = "https://www.airbnb.fr"+str(div1)
	            except:
	                aaa=1
	    #COMMENTAIRE
	        try:
	            tp_c=soup.find('div', attrs={'class': 'r16onr0j atm_c8_vvn7el atm_g3_k2d186 atm_fr_1vi102y atm_gq_myb0kj atm_vv_qvpr2i atm_c8_sz6sci__14195v1 atm_g3_17zsb9a__14195v1 atm_fr_kzfbxz__14195v1 atm_gq_idpfg4__14195v1 dir dir-ltr'}).text
	            ws.cell(row=c, column=cCOMMENT).value = tp_c
	        except:
	            try:
	                tp_c=Flogement.find('a').text
	                pp=tp_c.split(' ')
	                cc=pp[0]
	                ws.cell((c, cCOMMENT)).value = cc
	            except:
	                aaa=1
	    #VOYAGEUR
	        try:
	            tt=Flogement.find_all('li')[0]
	            #tt1=tt.find_all('span')[0]
	            tt2=tt.text
	            p_tp=tt2.split(" ")
	            ws.cell((c, cVOYAGEUR)).value = p_tp[0]
	            #print("V="+str(p_tp[0]))
	        except:
	            #print('NO VOYAGER')
	            aaa=1
	
	    #LITS
	        try:
	            tt=Flogement.find_all('li')[2].text
	            #tt1=tt.find_all('span')[2]
	            #print(tt)
	            #tt2=tt.find("span", class_="pen26si dir dir-ltr").find_next(text=True)
	            p_tp=tt.split(" ")
	            ws.cell((c, cLITS)).value = p_tp[2]
	            #print("L="+str(p_tp[0]))
	        except:
	            print('NO LIT')
	            aaa=1
	    #SdB
	        try:
	            tt=Flogement.find_all('li')[3].text
	            p_tp=tt.split(" ")
	            ws.cell((c, cSdB)).value = p_tp[2]
	            #print("B="+str(p_tp[0]))
	        except:
	            #print('NO SdB')
	            aaa=1
	    #CHAMBRE
	        try:
	            tt=Flogement.find_all('li')[1].text
	            p_tp=tt.split(" ")
	            ws.cell((c, cCHAMBRE)).value = p_tp[2]
	            #print("C="+str(p_tp[0]))
	        except:
	            #print('NO CHAMBRE')
	            aaa=1
	    #VILLE
	        try:
	            tp_c=soup.find('div', attrs={"class": "s1qk96pm atm_gq_p5ox87 dir dir-ltr"}).text
	            ws.cell((c, cVILLE)).value = tp_c
	            #print(tp_c)
	        except:
	            try:
	                tp_c=soup.find('div', attrs={"class": "_152qbzi"}).text
	                ws.cell((c, cVILLE)).value = tp_c
	                #print(tp_c)
	            except:
	                try:
	                    tp_c = soup.find('div', attrs={"data-plugin-in-point-id": "LOCATION_DEFAULT"}).h3.text
	                    ws.cell((c, cVILLE)).value = tp_c
	                except:
	                    print('NO VILLE')
	                    aaa=1
	    #NAME_HOTE
	        try:
	            tp_c=FProfile.find('span', attrs={'class': 't1gpcl1t atm_w4_16rzvi6 atm_9s_1o8liyq atm_gi_idpfg4 dir dir-ltr'}).text
	            pp=tp_c
	            ws.cell((c, cNAME_HOTE)).value = pp
	        except:
	            print ('NO_NAME')
	        if "Chambre" in type_home[0]:
	            try:
	                div=soup.find('button', attrs={'class': '_1j5azqwp l1ovpqvx dir dir-ltr'})
	                div1=div['aria-label']
	                ws.cell((c, cNAME_HOTE)).value =div1
	            except:
	                aaa=1
	    #ANCIENNETE
	
	    #SUPER HOTE
	        try:
	            tp_c=soup.find('div', attrs={"class": "bkwpcc1 atm_mk_stnw88 atm_6i_12gsa0d atm_n3_myb0kj dir dir-ltr"})
	            if tp_c is not None:
	                ws.cell((c, cSUPERHOTE)).value = 'X'
	                #print (tp_c)
	        except:
	            aaa=1
	    #COMMENT PROFIL
	
	        try:
	            #tp_c=FProfile.find('ul', attrs={"class": "tq6hspd h1aqtv1m dir dir-ltr"})
	            sp=FProfile.find('span', text=re.compile(r'\bévaluations\b')).text
	            pp=sp.split('é')
	            cc=pp[0]
	            if cc=='Identité vérifiée':
	                cc=0
	            ws.cell((c, cCOMMENT_PROFIL)).value = cc
	        except:
	            #print('No Comment profil')
	            aaa=1
	
	#DURING SEJOUR
	        try:
	            tt= FPolicies.find('span', text=re.compile(r'\bArrivée\b'))
	            ws.cell((c, cCHECK_IN)).value = tt.text
	            #print(tt.text)
	        except:
	            #print('no ARRIVE')
	            aaa=1
	        try:
	            tt= FPolicies.find('span', text=re.compile(r'\bDépart\b'))
	            ws.cell((c, cCHECK_OUT)).value = tt.text
	        except:
	            #print('no DEPART')
	            aaa=1
	        try:
	            tt= FPolicies.find('span', text=re.compile(r'\bNon fumeur\b'))
	            ws.cell((c, cFUMEUR)).value = tt.text
	        except:
	            #print('no FUMEUR')
	            aaa=1
	        try:
	            tt= FPolicies.find('span', text=re.compile(r'\bNe convient pas aux\b'))
	            ws.cell((c, cENFANT)).value = tt.text
	        except:
	            #print('no CHILD')
	            aaa=1
	        try:
	            tt= FPolicies.find('span', text=re.compile(r"\bArrivée autonome\b"))
	            ws.cell((c, cSERRURE)).value = tt.text
	        except:
	            #print('no AUTOMATIC')
	            aaa=1
	        try:
	            tt= FPolicies.find('span', text=re.compile(r"\bPas d'animaux\b"))
	            ws.cell((c, cANIMAUX)).value = tt.text
	        except:
	            try:
	                tt= FPolicies.find('span', text=re.compile(r"\bAnimaux de compagnie\b"))
	                ws.cell((c, cANIMAUX)).value = tt.text
	            except:
	                #print('no ANIMAL')
	                aaa=1
	        try:
	            tt= FPolicies.find('span', text=re.compile(r"\bDétecteur de fumée\b"))
	            ws.cell((c, cFUMEE)).value = tt.text
	        except:
	            #print('no detecteur fumee')
	            aaa=1
	        try:
	            tt= FPolicies.find('span', text=re.compile(r"\bDétecteur de monoxyde de carbone\b"))
	            ws.cell((c, cMONOXYDE)).value = tt.text
	        except:
	            #print('no detecteur monoxyde')
	            aaa=1
	        try:
	            tt= FPolicies.find('span', text=re.compile(r"\bPas de fête ni de soirée\b"))
	            ws.cell((c, cFETE)).value = tt.text
	        except:
	            #print('no detecteur monoxyde')
	            aaa=1

	#ENTREPRISE
	        try:
	            the_tr= FProfile.find('button', text=re.compile(r"\bentreprise\b"))
	            #print(the_tr)
	            if the_tr is not None:
	                ws.cell((c, cENTREPRISE)).value = 'YES'
	        except:
	            try:
	                the_tr= FProfile.find('li', text=re.compile(r"\bProfessionnel\b"))
	                #print(the_tr)
	                if the_tr is not None:
	                    ws.cell((c, cENTREPRISE)).value = 'YES'
	            except:
	                ws.cell((c, cENTREPRISE)).value = 'NO'
	
	#CO HOTE
	        try:
	            FCohote = FProfile.find('ul', attrs={'class': 'ato18ul atm_84_ave25a atm_9s_1txwivl atm_au_qxlwhf atm_gb_glywfm atm_gq_idpfg4 atm_h3_idpfg4 atm_l8_idpfg4 atm_n5_ave25a dir dir-ltr'})
	            try:
	                url_cohote1 = FCohote.findAll('a')[0]
	                url_cohote2 = FCohote.findAll('a')[1]
	                #url_cohote1 = FCohote.find('a')
	                nam_cohote1 = FCohote.findAll('span')[0]
	                ws.cell((c, cCOHOTE_URL1)).value = "https://www.airbnb.fr"+str(url_cohote1['href'])
	                ws.cell((c, cCOHOTE_NAME1)).value = nam_cohote1.text
	                ws.cell((c, cNB_COHOTE)).value = 1
	                #url_cohote2 = FCohote.find('a')
	                nam_cohote2 = FCohote.findAll('span')[1]
	                ws.cell((c, cCOHOTE_URL2)).value = "https://www.airbnb.fr"+str(url_cohote2['href'])
	                ws.cell((c, cCOHOTE_NAME2)).value = nam_cohote2.text
	                ws.cell((c, cNB_COHOTE)).value = 2
	            except:
	                url_cohote = FCohote.find('a')
	                nam_cohote = FCohote.find('span')
	                #print(url_cohote)
	                #print(nam_cohote.text)
	                ws.cell((c, cCOHOTE_URL1)).value = "https://www.airbnb.fr"+str(url_cohote['href'])
	                ws.cell((c, cCOHOTE_NAME1)).value = nam_cohote.text
	                ws.cell((c, cNB_COHOTE)).value = 1
	
	        except:
	            aaa=1
	#------------------------
	except:
	    pass


fm=2
fff=0

nrow=100000
h=ws.cell((c, cANNONCE)).value
threading.Thread(target=scrap, args=(h,)).start()
threading.Thread(target=scrap_description, args=(h,c,)).start()
time.sleep(8)
while c<=nrow:
	print (c)
	if (c/1000).is_integer():
		driver.quit()
		driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
		driver.set_window_size(2000, 1000)
		driver.get('https://www.google.com/')
		driver.get('chrome://settings/')
		driver.execute_script('chrome.settingsPrivate.setDefaultZoom(0.5);')
		driver.implicitly_wait(10)
		wait = WebDriverWait(driver, 5)
		wait2 = WebDriverWait(driver, 5)
		wait3 = WebDriverWait(driver, 5)
		wait = WebDriverWait(driver, 2)
		time.sleep(2)
		driver.get(h)
		time.sleep(20)
		scrap_ok=1

	try:
		time.sleep(1.5)
		if (c/20).is_integer():
			time.sleep(15)
			next_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Fermer']")))
			next_calendar.click()
		elif c==2:
			time.sleep(15)
			next_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Fermer']")))
			next_calendar.click()
	except:
		aaa=1
	try:
		time.sleep(15)
		driver.execute_script("window.scrollBy(0,2000);")
		time.sleep(4)
		driver.execute_script("window.scrollBy(0,2000);")
		time.sleep(4)
		driver.execute_script("window.scrollBy(0,1100);")
		time.sleep(7)
		html = driver.page_source
		soup = BeautifulSoup(html, 'html.parser')
		h=ws.cell((c+1, cANNONCE)).value
		threading.Thread(target=scrap, args=(h,)).start()
		time.sleep(1)
		threading.Thread(target=scrap_description, args=(h,c+1,)).start()
		threading.Thread(target=SCRAP_detail, args=(c,)).start()
	except:
		aaa=1
	c=c+1
print ('_______    ___    ___     ___')
print ('|      |   |  |   |  \    |  |')
print ('|  |__     |  |   |   \   |  |')
print ('|     |    |  |   |    \  |  |')
print ('|  |       |  |   |  |\ \ |  |')
print ('|  |       |  |   |  | \ \|  |')
print ('|__|       |__|   |__|  \____|')
