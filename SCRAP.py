from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import xlwt
import xlrd
import time
from xlutils.copy import copy
import re
import json
from urllib.request import urlopen
from openpyxl import load_workbook
from tkinter import filedialog
from tkinter import *
import os
from bs4 import BeautifulSoup
import threading
import sys


chrome_options = webdriver.ChromeOptions()
#prefs = {"profile.managed_default_content_settings.images": 2}
#chrome_options.add_experimental_option("prefs", prefs)

#driver.set_window_position(0, 0)
print ('▀▄▀▄▀▄ STOP AIRBNB ▄▀▄▀▄▀')

#--------SELECTION DU FICHIER AVEC LES NOUVELLES ANNONCES--------

wbx = load_workbook(path_RESULT.filename)
ws = wbx.active
#INITIALISATION EXCEL


#Récupération des URL depuis le xls
#list_URL=sheet_read.col_values(1)
list_URL=[]
for col in ws['B']:
    list_URL.append(col.value)

del list_URL[0]

#Récupérer les XPATH

YN_title='YES'
YN_profil='YES'
YN_name='YES'
YN_price='NO'
YN_comment='YES'
YN_voyageur='YES'
YN_chamber='YES'
YN_bed='YES'
YN_gps='YES'
YN_type='YES'
YN_old1='YES'
YN_old2='YES'
YN_ville='YES'
YN_super='YES'
YN_SdB='YES'

#OUVERTURE DES PAGES CHROME
#driver = webdriver.Chrome(chrome_options=chrome_options)
driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver')
#driver = webdriver.Chrome()
driver.set_window_size(800, 800)
time.sleep(2)
#c = ligne 2 du xls resultant
c=2
wait2 = WebDriverWait(driver, 2)
wait3 = WebDriverWait(driver, 3)	

i=1

def extract(XP, nb, type, c, YP):
	wait = WebDriverWait(driver, 2)
	if YP=='YES':
		try:
			tp = wait.until(EC.presence_of_element_located((By.XPATH, XP))).text
			#sheet.write(c, nb, tp)
			ws.cell(row=c+1, column=nb+1).value = tp
			print (tp)
		except:
			print(type)
			pass
end=0
while end==0:
	try:
		hh=ws.cell(row=c, column=2).value
		print(hh)
		if hh==None:
			print('END')
			end=1
		else:
			print (c)
			print (hh)
			#do=sheet_read.cell(i,0).value
			do=ws.cell(row=c, column=1).value
			if do==None:
				driver.get(hh)
				time.sleep(7)
				f_ele=0
				while f_ele<=3:
					try:
						ele=driver.find_element_by_xpath("//div[@class='_384m8u']")
						driver.execute_script("arguments[0].scrollIntoView(true);", ele)
						driver.execute_script("window.scrollBy(0,-100);")
						#driver.execute_script("window.scrollBy(0,500);")
						f_ele=6
						time.sleep(5)
					except:
						f_ele=f_ele+1
						time.sleep(1)
			#PROFILE
				html = driver.page_source
				time.sleep(2)
				soup = BeautifulSoup(html, 'html.parser')
				time.sleep(3)
				try:
				#TITLE
					try:
						div1=soup.find('div', attrs={"class": "_mbmcsn"})
						ws.cell(row=c, column=1).value = div1.h1.text
					except:
						print('NO TITLE')
				#URL HOTE
					try:
						div=soup.findAll('a', attrs={"class": "_105023be"})[-1]
						div1=div['href']  #.attrs['href']
						ws.cell(row=c, column=5).value = "https://www.airbnb.fr"+str(div1)
					except:
						print('NO PROFILE')
				#COMMENTAIRE
					COMMENT='NO COMMENT'
					#run_price=extract("//span[@class='_wfad8t']",6,COMMENT,c,YN_comment)
					try:
						p_c=[]
						tp_c=soup.findAll('span', attrs={"class": "_bq6krt"})[1].text
						p_c=tp_c.replace("(","")
						cc=p_c.replace(")","")
						try:
							pp=cc.split(' ')
							cc=pp[1]
							ws.cell(row=c, column=20).value = pp[0]
						except:
							pass
						ws.cell(row=c, column=7).value = cc
						#print ("COMMENT ===")
						print(cc)
						#p_c=tp_c.split("(")
						#print('ici1')
						#table_c = p_c[1].replace(")"," ")
						#print (table_c)
						#ws.cell(row=c+1, column=6+1).value = table_c
					except:
						print('NOCOMMENT')
				#VOYAGEUR
					if YN_voyageur=='YES':
						try:
							the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
							tt=the_tr.find_all('div')[1]
							tt1=tt.find_all('span')[0]
							tt2=tt1.text
							p_tp=tt2.split(" ")
							ws.cell(row=c, column=9).value = p_tp[0]
						except:
							print('NO VOYAGER')

				#LITS
					if YN_bed=='YES':
						try:
							the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
							tt=the_tr.find_all('div')[1]
							tt1=tt.find_all('span')[4]
							tt2=tt1.text
							p_tp=tt2.split(" ")
							ws.cell(row=c, column=12).value = p_tp[0]
						except:
							print('NO LIT')
				#SdB
					if YN_SdB=='YES':
						try:
							the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
							tt=the_tr.find_all('div')[1]
							tt1=tt.find_all('span')[6]
							tt2=tt1.text
							p_tp=tt2.split(" ")
							ws.cell(row=c, column=11).value = p_tp[0]
						except:
							print('NO SdB')
				#CHAMBRE
					if YN_chamber=='YES':
						try:
							the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
							tt=the_tr.find_all('div')[1]
							tt1=tt.find_all('span')[2]
							tt2=tt1.text
							p_tp=tt2.split(" ")
							ws.cell(row=c, column=10).value = p_tp[0]
						except:
							print('NO CHAMBRE')
				#VILLE
					try:
						tp_c=soup.find('a', attrs={"class": "_5twioja"}).text
						ws.cell(row=c, column=13).value = tp_c
					except:
						print('NO VILLE')

				#GPS
					try:
						tp_c=soup.find('div', attrs={"class": "gm-style"})
						tt=tp_c.find('div', attrs={"style": "margin-left: 5px; margin-right: 5px; z-index: 1000000; position: absolute; left: 0px; bottom: 0px;"})
						tt1=tt.find('a', attrs={"target": "_blank"})
						tt2=tt1['href']
						#----------Create translation table----------
						table = str.maketrans('=&', '++')
						result_gps = tt2.translate(table)
						split_gps=result_gps.split("+")
						#https://www.google.com/maps?ll+46.23657,3.92341+z=14&t=m&hl=fr&gl=FR&mapclient=apiv3
						#https://maps.google.com/maps?ll=49.19054,-2.11715&z=14&t=m&hl=fr&gl=FR&mapclient=apiv3
						coor=split_gps[1]
						long_lat=coor.split(',')
						#--------------Write results--------------
						print(long_lat[0])
						print(long_lat[1])
						#sheet.write(c, 12, long_lat[0])
						#sheet.write(c, 13, long_lat[1])
						ws.cell(row=c, column=14).value = long_lat[0]
						ws.cell(row=c, column=15).value = long_lat[1]
					except:
						print('NO GPS')
				#NAME_HOTE
					try:
						tp_c=soup.find('div', attrs={"class": "_f47qa6"})
						tt=tp_c.find('div', attrs={"class": "_svr7sj"})
						tt1=tt.h2.get_text()
						pp=tt1.split('par ')
						ws.cell(row=c, column=3).value = pp[1]
					except:
						print ('NO_NAME')
				#TYPE_HOME
					try:
						the_tr= soup.find('div', attrs={"class": "_1b3ij9t"}).text
						pp=the_tr.split('.')
						ws.cell(row=c, column=8).value = pp[0]
					except:
						try:
							the_tr= soup.find('div', attrs={"class": "_xcsyj0"}).text
							pp=the_tr.split('.')
							ws.cell(row=c, column=8).value = pp[0]
						except:
							print('NOTYPE')
					
				#ANCIENNETE
					try:
						tp_c=soup.find('div', attrs={"class": "_f47qa6"})
						tt=tp_c.find('div', attrs={"class": "_svr7sj"})
						tt1=tt.div.get_text()
						ws.cell(row=c, column=4).value = tt1
					except:
						print ('NOOLD')

				#SUPER HOTE
					try:
						#the_tr= soup.find('span', text=re.compile(r'\bSuperhost\b'),attrs = {'aria-hidden' : 'false'})
						tp_c=soup.find('div', attrs={"class": "_1ft6jxp"}).text
						ws.cell(row=c, column=16).value = 'X'
					except:
						print('no superhote')
				#AUTONOME
					try:
						the_tr= soup.find('div', attrs = {'class' : '_vd6w38n'})
						#div2=the_tr.findNextSibling('div')
						#print(the_tr.section.span.div.span.text)
						ws.cell(row=c, column=17).value = the_tr.section.span.div.span.text
						#print(div2.text)
					except:
						print('no auto')
				#CHILDREN
					try:
						the_tr= soup.find('span', text=re.compile(r'\bNe convient pas aux enfants ni aux bébés\b'))
						ws.cell(row=c, column=18).value = the_tr.text
					except:
						print('no child')
				#ANNULATION
					try:
						the_tr= soup.find('div', text=re.compile(r'\bAnnulation gratuite\b'),attrs = {'class' : '_1qsawv5'})
						div2=the_tr.findNextSibling('div')
						ws.cell(row=c, column=19).value = div2.text
					except:
						print('no child')

			#/////NOTATION////
				#PROPRETE
					try:
						the_tr= soup.findAll('div', attrs={"class": "_1s11ltsf"})[0]
						tt=the_tr.find('div', attrs={"class": "_7pay"})
						#print(tt.span.text)
						ws.cell(row=c, column=21).value = tt.span.text
					except:
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[0]
							#print(tt.text)
							ws.cell(row=c, column=21).value = tt.text
						except:
							print('no proprete')
				#PRECISION
					try:
						the_tr= soup.findAll('div', attrs={"class": "_1s11ltsf"})[1]
						tt=the_tr.find('div', attrs={"class": "_7pay"})
						#print(tt.span.text)
						ws.cell(row=c, column=22).value = tt.span.text
					except:
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[1]
							#print(tt.text)
							ws.cell(row=c, column=22).value = tt.text
						except:
							print('no Precision')
				#COMMUNICATION
					try:
						the_tr= soup.findAll('div', attrs={"class": "_1s11ltsf"})[2]
						tt=the_tr.find('div', attrs={"class": "_7pay"})
						#print(tt.span.text)
						ws.cell(row=c, column=23).value = tt.span.text
					except:
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[2]
							#print(tt.text)
							ws.cell(row=c, column=23).value = tt.text
						except:
							print('no communication')
				#EMPLACEMENT
					try:
						the_tr= soup.findAll('div', attrs={"class": "_1s11ltsf"})[3]
						tt=the_tr.find('div', attrs={"class": "_7pay"})
						#print(tt.span.text)
						ws.cell(row=c, column=24).value = tt.span.text
					except:
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[3]
							#print(tt.text)
							ws.cell(row=c, column=24).value = tt.text
						except:
							print('no emplacement')
				#ARRIVEE
					try:
						the_tr= soup.findAll('div', attrs={"class": "_1s11ltsf"})[4]
						tt=the_tr.find('div', attrs={"class": "_7pay"})
						#print(tt.span.text)
						ws.cell(row=c, column=25).value = tt.span.text
					except:
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[4]
							#print(tt.text)
							ws.cell(row=c, column=25).value = tt.text
						except:
							print('no arrivee')
				#QUALITY PRICE
					try:
						the_tr= soup.findAll('div', attrs={"class": "_1s11ltsf"})[5]
						tt=the_tr.find('div', attrs={"class": "_7pay"})
						#print(tt.span.text)
						ws.cell(row=c, column=26).value = tt.span.text
					except:
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[5]
							#print(tt.text)
							ws.cell(row=c, column=26).value = tt.text
						except:
							print('no price quality')
			#N° ENREGISTREMENT
					try:
						the_tr= soup.find('li', text=re.compile(r'\bNuméro\b'), attrs = {'class' : '_1q2lt74'})
						pp=the_tr.text
						#print(pp)
						sp=pp.split(' ')
						ws.cell(row=c, column=27).value = sp[-1]
					except:
						a=1
						#print('no N° enregistrement')				
			#TAUX REPONSE
					try:
						the_tr=soup.find('li', text=re.compile(r'\bTaux\b'))
						pp=the_tr.text
						pp=pp.replace(" ","")
						#print(pp)
						sp=pp.split(':')
						#print(sp[-1])
						ws.cell(row=c, column=28).value = sp[-1]
					except:
						print('no taux réponse')
				#DELAI REPONSE
					try:
						the_tr=soup.find('li', text=re.compile(r'\bDélai\b'))
						pp=the_tr.text
						sp=pp.split(':')
						ws.cell(row=c, column=29).value = sp[-1]
					except:
						print('no DELAI REPONSE')
			#DURING SEJOUR
					try:
						the_tr=soup.find('div', attrs = {'class' : '_uz1jgk'})
						tt= the_tr.findAll('div', attrs = {'class' : '_eeq7h0'})[0]
						ttt=tt.span.text
						ws.cell(row=c, column=30).value = ttt
					except:
						print('no DURING SEJOUR')
					try:
						the_tr=soup.findAll('div', attrs={"class": "_1byskwn"})[-1]
						try:
							tt= the_tr.find('span', text=re.compile(r'\bArrivée\b'))
							ws.cell(row=c, column=31).value = tt.text
						except:
							print('no ARRIVE')
						try:
							tt= the_tr.find('span', text=re.compile(r'\bDépart\b'))
							ws.cell(row=c, column=32).value = tt.text
						except:
							print('no DEPART')
						try:
							tt= the_tr.find('span', text=re.compile(r'\bNon fumeur\b'))
							ws.cell(row=c, column=33).value = tt.text
						except:
							print('no FUMEUR')
						try:
							tt= the_tr.find('span', text=re.compile(r'\bNe convient pas aux\b'))
							ws.cell(row=c, column=34).value = tt.text
						except:
							print('no CHILD')
						try:
							tt= the_tr.find('span', text=re.compile(r"\bArrivée autonome\b"))
							ws.cell(row=c, column=35).value = tt.text
						except:
							print('no AUTOMATIC')
						try:
							tt= the_tr.find('span', text=re.compile(r"\bPas d'animaux\b"))
							ws.cell(row=c, column=36).value = tt.text
						except:
							try:
								tt= the_tr.find('span', text=re.compile(r"\bAnimaux de compagnie\b"))
								ws.cell(row=c, column=36).value = tt.text
							except:
								print('no ANIMAL')
						try:
							tt= the_tr.find('span', text=re.compile(r"\bCaution\b"))
							ws.cell(row=c, column=37).value = tt.text
						except:
							print('no Caution')
						try:
							tt= the_tr.find('span', text=re.compile(r"\bDétecteur de fumée\b"))
							ws.cell(row=c, column=38).value = tt.text
						except:
							print('no detecteur fumee')
						try:
							tt= the_tr.find('span', text=re.compile(r"\bDétecteur de monoxyde de carbone\b"))
							ws.cell(row=c, column=39).value = tt.text
						except:
							print('no detecteur monoxyde')
						try:
							tt= the_tr.find('span', text=re.compile(r"\bPas de fête ni de soirée\b"))
							ws.cell(row=c, column=40).value = tt.text
						except:
							print('no detecteur monoxyde')
					except:
						print('no INSIDE RULE')
			#LANGUE
					try:
						the_tr= soup.find('li', text=re.compile(r'\bLangues\b'))
						#print(the_tr)
						pp=the_tr.text
						#print(pp)
						sp=pp.split(':')
						ws.cell(row=c, column=41).value = sp[-1]
					except:
						print('no LANGUAGE')
			#IMAGE
					try:
						the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[0]
						#print(the_tr)
						tt=the_tr['src']
						ws.cell(row=c, column=42).value = tt
					except:
						try:
							the_tr= soup.find('img', attrs={"class": "_6tbg2q"})
							#print(the_tr)
							tt=the_tr['data-original-uri']
							ws.cell(row=c, column=42).value = tt
						except:
							print('no IMAGE 0')
					try:
						the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[1]
						#print(the_tr)
						tt=the_tr['src']
						ws.cell(row=c, column=44).value = tt
					except:
						try:
							the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[1]
							#print(the_tr)
							tt=the_tr['data-original-uri']
							ws.cell(row=c, column=44).value = tt
						except:
							print('no IMAGE 1')
					try:
						the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[2]
						#print(the_tr)
						tt=the_tr['src']
						ws.cell(row=c, column=45).value = tt
					except:
						try:
							the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[2]
							#print(the_tr)
							tt=the_tr['data-original-uri']
							ws.cell(row=c, column=45).value = tt
						except:
							print('no IMAGE 2')
					try:
						the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[3]
						#print(the_tr)
						tt=the_tr['src']
						ws.cell(row=c, column=46).value = tt
					except:
						try:
							the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[3]
							#print(the_tr)
							tt=the_tr['data-original-uri']
							ws.cell(row=c, column=46).value = tt
						except:
							print('no IMAGE 3')
					try:
						the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[4]
						#print(the_tr)
						tt=the_tr['src']
						print(tt)
						ws.cell(row=c, column=47).value = tt
					except:
						try:
							the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[4]
							#print(the_tr)
							tt=the_tr['data-original-uri']
							ws.cell(row=c, column=47).value = tt
						except:
							print('no IMAGE 4')
			#IMAGE_HOTE
					try:
						the_tr= soup.find('div', attrs={"class": "_5kripx"})
						t= the_tr.find('img', attrs={"class": "_6tbg2q"})
						tt=t['src']
						ws.cell(row=c, column=43).value = tt
					except:
						print('no IMAGE_HOTE')
					if (c/200).is_integer():
						wbx.save(path_RESULT.filename)
					if (c/1000).is_integer():
						wbx.save(DIR2+NAMEFile+str(c)+".xlsx")
		#------------------------
				except:
					try:
						#page sans annonce, je dois trouver quelque soit le type
						wait3.until(EC.presence_of_element_located((By.XPATH, "//h3[@class='_jmmm34f']")))
					except:
						# rien tourvé précédent, donc c'est que je suis sur mauvais design, clos et reopen chrome
						driver.quit()
						f_xpathdate=0
						fff=0
						fm=2
						driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver')
						#driver = webdriver.Chrome()
						driver.set_window_size(800, 800)
						wait3 = WebDriverWait(driver, 3)
						while f_xpathdate==0:
							h=ws.cell(row=fm, column=2).value
							print(h)
							if fff==10:
								f_mounth=1
								f_xpathdate=1
								end=0
							fff=fff+1
							try:
								driver.get(h)
								time.sleep(4)
								x_date = wait3.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_13m7kz7i']"))).text
								print("x date trouve")
								f_xpathdate=1
							except:
								if fff!=10:
									driver.quit()
									#driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver')
									driver = webdriver.Chrome()
									driver.set_window_size(800, 800)
									wait3 = WebDriverWait(driver, 3)

			c=c+1
					

	except:
		print("END")
		# EXCEPT si Chrome se ferme tout seul, ici il va le réouvrir et relancer la boucle d'extraction
		#driver = webdriver.Chrome()
		#driver.set_window_size(800, 1500)
		wbx.save(path_RESULT.filename)
wbx.save(path_RESULT.filename)
print ('_______    ___    ___     ___')
print ('|      |   |  |   |  \    |  |')
print ('|  |__     |  |   |   \   |  |')
print ('|     |    |  |   |    \  |  |')
print ('|  |       |  |   |  |\ \ |  |')
print ('|  |       |  |   |  | \ \|  |')
print ('|__|       |__|   |__|  \____|')
print ('Votre fichier RESULT est à présent terminé, renommer le et concerver le.')
print ('Vous pouvez maintenant:')
print ('  1- exécuter le .exe N°4 qui calcule le nombre de logement détenus par hôte')
print ('  2- exétuter le .exe N°5 qui lui calcule les nuitées de chaque annonces, à exécuter hebdomadairement')
print ('  3- N°6 qui est à utiliser si votre fichier RESULT contient plus de 2000 annonces, il va découper le fichier en lot de moins de 2000 annonces si vous voulez l importer sur GoogleMap')
print ('  4- N°7 vous permet de comparer le fichier RESULT avec un autre fichier de liste d annonce que vous possédez déjà, il va alors ajouter dans votre autre fichier les annonces manquante')
end=1
