import xlwt
import xlrd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from xlutils.copy import copy
from decimal import Decimal
from decimal import *
import decimal
#import xlwings as xw
#from xlwings.constants import DeleteShiftDirection
from xlrd import open_workbook,XL_CELL_TEXT
import datetime
from datetime import date
import datetime as dt
from tkinter import filedialog
from tkinter import *
import os
import re
import json
from urllib.request import urlopen
import pip
import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from shapely.geometry import Point
from shapely.geometry.polygon import Polygon
import datetime
import datetime as dt

#--------SELECTION DU FICHIER AVEC LES NOUVELLES ANNONCES--------

#now = str(datetime.date.now())[:19]
#now = now.replace(":","_")
now = str(date.today())
print(now)

Deccontext = Context(prec=10, rounding=ROUND_HALF_DOWN)
setcontext(Deccontext)

chrome_options = webdriver.ChromeOptions()
#prefs = {"profile.managed_default_content_settings.images": 2}
#chrome_options.add_experimental_option("prefs", prefs)

#on récupère la value dans le excel ORIGINE
#32460

workbook = xlwt.Workbook()
sheet_write = workbook.add_sheet('URL_PAGE')
sheet_write.write(0, 0, 'A')
sheet_write.write(0, 1, 'B')
sheet_write.write(0, 2, 'C')
sheet_write.write(0, 3, 'D')
sheet_write.write(0, 4, 'E')
sheet_write.write(0, 5, 'F')
sheet_write.write(1, 5, 'b')
sheet_write.write(0, 6, 'G')

fromINSEE=0
if fromINSEE==1:
	wb2 = xlrd.open_workbook('INSEE_FRANCE_XML.xls')
	sheet1 = wb2.sheet_by_index(0)
	v=0
	i=1
	ville='78158'   #LECHESNAY
	#ville='78646'  #VERSAILLES
	#ville='75102'
	#ville="64483"
	#ville="29046"
	#ISSY
	#ville=92040
	#st-malo
	#ville=35288
	#SABLE-OLONNE
	#ville=85194
	tv=str(ville)
	print(tv)
	while v==0:
		naming=sheet1.cell_value(i,0)
		try:
			tt=int(naming)
			t=str(tt)
		except:
			t=str(naming)
		if t==tv:
			v=1
			GPS_ORIGIN=sheet1.cell_value(i,6)
		i=i+1

	#on supprime les space
	GPS_sans_space=GPS_ORIGIN.replace(" ","")

	# split sur "["
	split_gps=GPS_sans_space.split("[")

	#on supprime le 1er element de la liste car il est vide
	split_gps.remove(split_gps[0])
	split_gps.remove(split_gps[0])
	split_gps.remove(split_gps[0])
	#longeur de la liste
	T_gps=len(split_gps)
	print (T_gps)

	#ligne excel initialisé à 0
	c=0
	l_max=[]
	l_min=[]
	for h in split_gps:
		crochet=h.split("]")
		double_gps=crochet[0].split(",")
		c=c+1
		l_max.append(double_gps[-1])
		l_min.append(double_gps[0])
	print(max(l_max)) #ne_lat3 N1
	print(min(l_min)) #sw_lat1 S3
	print(min(l_max)) #ne_lng2 w2
	print(max(l_min)) #sw_lng4 e4
elif fromINSEE==2:
	#Poly=[(2.922176, 48.928696), (2.626191, 48.935857), (2.628215, 48.783178), (2.953307, 48.785616), (2.922176, 48.928696)]
	#Poly=[(-1.7022587,46.556284), (-1.8665397,46.45638)]
	#Poly=[(-0.566461976479953, 44.54605534503501), (-0.601676214501757, 44.55285985236557), (-0.615502791796292, 44.58296692120781), (-0.609008046685042, 44.60364546556298), (-0.619038811778664, 44.61095042862514), (-0.6619057650858861, 44.6087626175587), (-0.694386280213617, 44.64370391840443), (-0.729966737111434, 44.67257549413891), (-0.721742144539515, 44.68114439646803), (-0.65662732785245, 44.68355975040168), (-0.66042725398573, 44.68993207437811), (-0.672078913543385, 44.69481824801064), (-0.64955519919815, 44.72100596663734), (-0.638113761690622, 44.7272407973585), (-0.635703284981555, 44.73024005567959), (-0.645128260959859, 44.73169655392201), (-0.638302988360768, 44.73661663347337), (-0.633985326942008, 44.73593830796983), (-0.632718880232667, 44.74217148515195), (-0.616642192434454, 44.74918679010368), (-0.607622410104385, 44.75723173048317), (-0.590136906201729, 44.76270268236247), (-0.585187455624579, 44.75312966997373), (-0.571356501419781, 44.74405964965744), (-0.562255748275608, 44.74192854372193), (-0.5658157884666259, 44.75200675728984), (-0.540660545025235, 44.76185305655517), (-0.53462259120344, 44.76721886057388), (-0.510893830266989, 44.76909999690599), (-0.5111915025439741, 44.76495405568519), (-0.511578560950593, 44.74583642722698), (-0.501591689732834, 44.73999758637085), (-0.483650238920068, 44.73612976395617), (-0.458048243215631, 44.72667717584467), (-0.448416777206444, 44.72194919998991), (-0.438030215293798, 44.70545692757102), (-0.436840151970893, 44.70417822912138), (-0.433892184161965, 44.70266572453448), (-0.436871442073003, 44.68875876065009), (-0.457059178510567, 44.66861994643612), (-0.441642519281372, 44.65480890650544), (-0.459694058384267, 44.63011315020232), (-0.488561495415893, 44.6269617627626), (-0.488719701862386, 44.61728753564925), (-0.488809427352535, 44.61241855986734), (-0.46708540617104, 44.60544291991868), (-0.481051634606019, 44.59159488660275), (-0.48790043592681, 44.57530270334252), (-0.511297994605287, 44.57001029560205), (-0.54576155958516, 44.56977378565707), (-0.550574420737808, 44.55971544008108), (-0.566461976479953, 44.54605534503501)]
	#perimeter=[-0.566461976479953,44.54605534503501;-0.601676214501757,44.55285985236557;-0.615502791796292,44.58296692120781;-0.609008046685042,44.60364546556298;-0.619038811778664,44.61095042862514;-0.6619057650858861,44.6087626175587;-0.694386280213617,44.64370391840443;-0.729966737111434,44.67257549413891;-0.721742144539515,44.68114439646803;-0.65662732785245,44.68355975040168;-0.66042725398573,44.68993207437811;-0.672078913543385,44.69481824801064;-0.64955519919815,44.72100596663734;-0.638113761690622,44.7272407973585;-0.635703284981555,44.73024005567959;-0.645128260959859,44.73169655392201;-0.638302988360768,44.73661663347337;-0.633985326942008,44.73593830796983;-0.632718880232667,44.74217148515195;-0.616642192434454,44.74918679010368;-0.607622410104385,44.75723173048317;-0.590136906201729,44.76270268236247;-0.585187455624579,44.75312966997373;-0.571356501419781,44.74405964965744;-0.562255748275608,44.74192854372193;-0.5658157884666259,44.75200675728984;-0.540660545025235,44.76185305655517;-0.53462259120344,44.76721886057388;-0.510893830266989,44.76909999690599;-0.5111915025439741,44.76495405568519;-0.511578560950593,44.74583642722698;-0.501591689732834,44.73999758637085;-0.483650238920068,44.73612976395617;-0.458048243215631,44.72667717584467;-0.448416777206444,44.72194919998991;-0.438030215293798,44.70545692757102;-0.436840151970893,44.70417822912138;-0.433892184161965,44.70266572453448;-0.436871442073003,44.68875876065009;-0.457059178510567,44.66861994643612;-0.441642519281372,44.65480890650544;-0.459694058384267,44.63011315020232;-0.488561495415893,44.6269617627626;-0.488719701862386,44.61728753564925;-0.488809427352535,44.61241855986734;-0.46708540617104,44.60544291991868;-0.481051634606019,44.59159488660275;-0.48790043592681,44.57530270334252;-0.511297994605287,44.57001029560205;-0.54576155958516,44.56977378565707;-0.550574420737808,44.55971544008108;-0.566461976479953,44.54605534503501]
	#perimeter=[-0.566461976479953,44.54605534503501; -0.601676214501757,44.55285985236557; -0.615502791796292,44.58296692120781; -0.609008046685042,44.60364546556298; -0.619038811778664,44.61095042862514; -0.6619057650858861,44.6087626175587; -0.694386280213617,44.64370391840443; -0.729966737111434,44.67257549413891; -0.721742144539515,44.68114439646803; -0.65662732785245,44.68355975040168; -0.66042725398573,44.68993207437811; -0.672078913543385,44.69481824801064; -0.64955519919815,44.72100596663734; -0.638113761690622,44.7272407973585; -0.635703284981555,44.73024005567959; -0.645128260959859,44.73169655392201; -0.638302988360768,44.73661663347337; -0.633985326942008,44.73593830796983; -0.632718880232667,44.74217148515195; -0.616642192434454,44.74918679010368; -0.607622410104385,44.75723173048317; -0.590136906201729,44.76270268236247; -0.585187455624579,44.75312966997373; -0.571356501419781,44.74405964965744; -0.562255748275608,44.74192854372193; -0.5658157884666259,44.75200675728984; -0.540660545025235,44.76185305655517; -0.53462259120344,44.76721886057388; -0.510893830266989,44.76909999690599; -0.5111915025439741,44.76495405568519; -0.511578560950593,44.74583642722698; -0.501591689732834,44.73999758637085; -0.483650238920068,44.73612976395617; -0.458048243215631,44.72667717584467; -0.448416777206444,44.72194919998991; -0.438030215293798,44.70545692757102; -0.436840151970893,44.70417822912138; -0.433892184161965,44.70266572453448; -0.436871442073003,44.68875876065009; -0.457059178510567,44.66861994643612; -0.441642519281372,44.65480890650544; -0.459694058384267,44.63011315020232; -0.488561495415893,44.6269617627626; -0.488719701862386,44.61728753564925; -0.488809427352535,44.61241855986734; -0.46708540617104,44.60544291991868; -0.481051634606019,44.59159488660275; -0.48790043592681,44.57530270334252; -0.511297994605287,44.57001029560205; -0.54576155958516,44.56977378565707; -0.550574420737808,44.55971544008108; -0.566461976479953,44.54605534503501]
	#perimeter=[(-0.566461976479953, 44.54605534503501), (-0.601676214501757, 44.55285985236557), (-0.615502791796292, 44.58296692120781), (-0.609008046685042, 44.60364546556298), (-0.619038811778664, 44.61095042862514), (-0.6619057650858861, 44.6087626175587), (-0.694386280213617, 44.64370391840443), (-0.729966737111434, 44.67257549413891), (-0.721742144539515, 44.68114439646803), (-0.65662732785245, 44.68355975040168), (-0.66042725398573, 44.68993207437811), (-0.672078913543385, 44.69481824801064), (-0.64955519919815, 44.72100596663734), (-0.638113761690622, 44.7272407973585), (-0.635703284981555, 44.73024005567959), (-0.645128260959859, 44.73169655392201), (-0.638302988360768, 44.73661663347337), (-0.633985326942008, 44.73593830796983), (-0.632718880232667, 44.74217148515195), (-0.616642192434454, 44.74918679010368), (-0.607622410104385, 44.75723173048317), (-0.590136906201729, 44.76270268236247), (-0.585187455624579, 44.75312966997373), (-0.571356501419781, 44.74405964965744), (-0.562255748275608, 44.74192854372193), (-0.5658157884666259, 44.75200675728984), (-0.540660545025235, 44.76185305655517), (-0.53462259120344, 44.76721886057388), (-0.510893830266989, 44.76909999690599), (-0.5111915025439741, 44.76495405568519), (-0.511578560950593, 44.74583642722698), (-0.501591689732834, 44.73999758637085), (-0.483650238920068, 44.73612976395617), (-0.458048243215631, 44.72667717584467), (-0.448416777206444, 44.72194919998991), (-0.438030215293798, 44.70545692757102), (-0.436840151970893, 44.70417822912138), (-0.433892184161965, 44.70266572453448), (-0.436871442073003, 44.68875876065009), (-0.457059178510567, 44.66861994643612), (-0.441642519281372, 44.65480890650544), (-0.459694058384267, 44.63011315020232), (-0.488561495415893, 44.6269617627626), (-0.488719701862386, 44.61728753564925), (-0.488809427352535, 44.61241855986734), (-0.46708540617104, 44.60544291991868), (-0.481051634606019, 44.59159488660275), (-0.48790043592681, 44.57530270334252), (-0.511297994605287, 44.57001029560205), (-0.54576155958516, 44.56977378565707), (-0.550574420737808, 44.55971544008108), (-0.566461976479953, 44.54605534503501)]
	#GPS_ORIGIN='{"type": "Polygon", "coordinates": [[[-0.566461976479953, 44.54605534503501], [-0.601676214501757, 44.55285985236557], [-0.615502791796292, 44.58296692120781], [-0.609008046685042, 44.60364546556298], [-0.619038811778664, 44.61095042862514], [-0.6619057650858861, 44.6087626175587], [-0.694386280213617, 44.64370391840443], [-0.729966737111434, 44.67257549413891], [-0.721742144539515, 44.68114439646803], [-0.65662732785245, 44.68355975040168], [-0.66042725398573, 44.68993207437811], [-0.672078913543385, 44.69481824801064], [-0.64955519919815, 44.72100596663734], [-0.638113761690622, 44.7272407973585], [-0.635703284981555, 44.73024005567959], [-0.645128260959859, 44.73169655392201], [-0.638302988360768, 44.73661663347337], [-0.633985326942008, 44.73593830796983], [-0.632718880232667, 44.74217148515195], [-0.616642192434454, 44.74918679010368], [-0.607622410104385, 44.75723173048317], [-0.590136906201729, 44.76270268236247], [-0.585187455624579, 44.75312966997373], [-0.571356501419781, 44.74405964965744], [-0.562255748275608, 44.74192854372193], [-0.5658157884666259, 44.75200675728984], [-0.540660545025235, 44.76185305655517], [-0.53462259120344, 44.76721886057388], [-0.510893830266989, 44.76909999690599], [-0.5111915025439741, 44.76495405568519], [-0.511578560950593, 44.74583642722698], [-0.501591689732834, 44.73999758637085], [-0.483650238920068, 44.73612976395617], [-0.458048243215631, 44.72667717584467], [-0.448416777206444, 44.72194919998991], [-0.438030215293798, 44.70545692757102], [-0.436840151970893, 44.70417822912138], [-0.433892184161965, 44.70266572453448], [-0.436871442073003, 44.68875876065009], [-0.457059178510567, 44.66861994643612], [-0.441642519281372, 44.65480890650544], [-0.459694058384267, 44.63011315020232], [-0.488561495415893, 44.6269617627626], [-0.488719701862386, 44.61728753564925], [-0.488809427352535, 44.61241855986734], [-0.46708540617104, 44.60544291991868], [-0.481051634606019, 44.59159488660275], [-0.48790043592681, 44.57530270334252], [-0.511297994605287, 44.57001029560205], [-0.54576155958516, 44.56977378565707], [-0.550574420737808, 44.55971544008108], [-0.566461976479953, 44.54605534503501]]]}'
	#GPS_ORIGIN='{"type": "Polygon", "coordinates": [[[2.922176, 48.928696], [2.626191, 48.935857], [2.628215, 48.783178], [2.953307, 48.785616], [2.922176, 48.928696]]]}'
	#olonne
	#GPS_ORIGIN='{"type": "Polygon", "coordinates": [[[-1.7022587,46.556284], [-1.7022587,46.45638], [-1.8665397,46.45638], [-1.8665397,46.556284]]]}'
	#angers
	#GPS_ORIGIN='{"type": "Polygon", "coordinates": [[[-0.3343564,47.5825096], [-0.3343564,47.267604], [-0.9328407,47.267604], [-0.9328407,47.5825096]]]}'
	GPS_ORIGIN='{"type": "Polygon", "coordinates": [[[5.9192223,50.240395], [5.9192223,48.917916], [3.583670,48.917916], [3.583670,50.240395]]]}'
	
	#on supprime les space
	GPS_sans_space=GPS_ORIGIN.replace(" ","")

	# split sur "["
	split_gps=GPS_sans_space.split("[")

	#on supprime le 1er element de la liste car il est vide
	split_gps.remove(split_gps[0])
	split_gps.remove(split_gps[0])
	split_gps.remove(split_gps[0])
	#longeur de la liste
	T_gps=len(split_gps)
	print (T_gps)

	#ligne excel initialisé à 0
	c=0
	l_max=[]
	l_min=[]
	for h in split_gps:
		crochet=h.split("]")
		double_gps=crochet[0].split(",")
		c=c+1
		l_max.append(double_gps[-1])
		l_min.append(double_gps[0])
	print(max(l_max)) #ne_lat N
	print(max(l_min)) #ne_lng E
	print(min(l_max)) #sw_lat S
	print(min(l_min)) #sw_lng W
else:
	if zone =='ZJ':
		l_max=[48.108438, 46.170246]
		l_min=[2.1249163, -0.8501037]
#2.1249163	48.108438	-0.8501037	46.170246
	elif zone =='ZA':
		l_max=[49.269973, 48.108438]
		l_min=[3.6975063, 1.4208963]
#3.6975063	49.269973	1.4208963	48.108438
	elif zone =='ZB':
		l_max=[50.262211, 49.269973]
		l_min=[1.4771713, -0.1394017]
#1.4771713	50.262211	-0.1394017	49.269973
	elif zone =='ZC':
		l_max=[51.557589, 49.269973]
		l_min=[5.3282123, 1.4771713]
#5.3282123	51.557589	1.4771713	49.269973
	elif zone =='ZD':
		l_max=[49.907681, 47.396602]
		l_min=[8.4895203, 4.7506603]
#8.4895203	49.907681	4.7506603	47.396602
	elif zone =='ZE':
		l_max=[49.269973, 48.108438]
		l_min=[5.1227783, 3.6975063]
#5.1227783	49.269973	3.6975063	48.108438
	elif zone =='ZF':
		l_max=[49.269973, 48.108438]
		l_min=[1.4208963, -0.8501037]
#1.4208963	49.269973	-0.8501037	48.108438
	elif zone =='ZG':
		l_max=[49.269973, 47.38346]
		l_min=[-0.8501037, -2.6680587]
#-0.8501037	49.269973	-2.6680587	47.38346
	elif zone =='ZH':
		l_max=[49.026844, 46.894108]
		l_min=[-2.6680587, -5.6799817]
#-2.6680587	49.026844	-5.6799817	46.894108
	elif zone =='ZI':
		l_max=[47.38346, 46.170246]
		l_min=[-0.8501037, -2.8552827]
#-0.8501037	47.38346	-2.8552827	46.170246
	elif zone =='ZK':
		l_max=[48.108438, 46.170246]
		l_min=[8.4895203, 2.1249163]
#8.4895203	48.108438	2.1249163	46.170246
	elif zone =='ZL':
		l_max=[47.396602, 46.170246]
		l_min=[7.2510053, 4.7506603]
#7.2510053	47.396602	4.7506603	46.170246
	elif zone =='ZM':
		l_max=[47.396602, 44.277209]
		l_min=[7.4344513, 3.8511853]
#7.4344513	47.396602	3.8511853	44.277209
	elif zone =='ZN':
		l_max=[47.396602, 44.277209]
		l_min=[3.8511853, 0.7378083]
#3.8511853	47.396602	0.7378083	44.277209
	elif zone =='ZO':
		l_max=[47.396602, 44.555503]
		l_min=[0.7378083, -1.9987367]
#0.7378083	47.396602	-1.9987367	44.555503
	elif zone =='ZP':
		l_max=[44.555503, 42.674653]
		l_min=[0.7378083, -1.821258]
#0.7378083	44.555503	-1.821258	42.674653
	elif zone =='ZQ':
		l_max=[44.277209, 42.316054]
		l_min=[3.5806873, -1.821258]
#3.5806873	44.277209	-1.821258	42.316054
	elif zone =='ZR':
		l_max=[44.277209, 42.920252]
		l_min=[6.0728653, 3.5806873]
#6.0728653	44.277209	3.5806873	42.920252
	elif zone =='ZS':
		l_max=[44.277209, 42.920252]
		l_min=[7.7303183, 6.0728653]
#7.7303183	44.277209	6.0728653	42.920252
	elif zone =='ZT':
		l_max=[50.262211, 49.269973]
		l_min=[-0.1394017, -2.1716267]
#-0.1394017	50.262211	-2.1716267	49.269973
	
	
	
	print(max(l_max)) #ne_lat N
	print(max(l_min)) #ne_lng E
	print(min(l_max)) #sw_lat S
	print(min(l_min)) #sw_lng W

#l_max=[48.863344374598334, 48.869863809746434, 48.870630685567946]
#l_min=[2.350834505477619, 2.327877416924118, 2.347826239446091]
#https://www.airbnb.fr/s/dinard/homes?refinement_paths%5B%5D=%2Fhomes&query=dinard&search_type=unknown&ne_lat=48.6435799942263&ne_lng=-2.0336458067321246&sw_lat=48.623414896166715&sw_lng=-2.074629960479683&zoom=15&search_by_map=true&place_id=ChIJVVqqTcCADkgR48BvH6Pvhd8&s_tag=Y6jz9TTh
#https://www.airbnb.fr/s/blois/homes?refinement_paths%5B%5D=%2Fhomes&current_tab_id=home_tab&selected_tab_id=home_tab&source=mc_search_bar&click_referer=t%3ASEE_ALL%7Csid%3Aba7e066c-7524-4d15-8ebc-24f988bad2e4%7Cst%3ALANDING_PAGE_MARQUEE&screen_size=large&zoom=14&search_by_map=true&sw_lat=47.56476024379833&sw_lng=1.2938918252564235&ne_lat=47.611939549842894&ne_lng=1.3847865243531032&search_type=unknown&adults=1
#https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat=47.54160887370493&sw_lng=1.355879995800566&ne_lat=47.620619985628096&ne_lng=1.254120619005437&zoom=14&search_by_map=true&map_toggle=true
#url_start='https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&current_tab_id=home_tab&selected_tab_id=home_tab&allow_override%5B%5D=&sw_lat='+str(min(l_max))+'&sw_lng='+str(min(l_min))+'&ne_lat='+str(max(l_max))+'&ne_lng='+str(max(l_min))
#url_start='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat='+str(min(l_max))+'&sw_lng='+str(min(l_min))+'&ne_lat='+str(max(l_max))+'&ne_lng='+str(max(l_min))+'&zoom=14&search_by_map=true&map_toggle=true'
#ttt
typ=3

if typ==1:
	url_start='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=unknown&ne_lat='+str(max(l_max))+'&ne_lng='+str(max(l_min))+'& sw_lat='+str(min(l_max))+'&sw_lng=-'+str(min(l_min))+'&zoom=17&search_by_map=true'
	sheet_write.write(1, 0, url_start)
	sheet_write.write(1, 1, max(l_max))#3 N
	sheet_write.write(1, 2, max(l_min))#4 E
	sheet_write.write(1, 3, min(l_max))#1 S
	sheet_write.write(1, 4, min(l_min))#2 W
	url_start='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=unknown&ne_lat='+str(max(l_max))+'&ne_lng='+str(max(l_min))+'& sw_lat='+str(min(l_max))+'&sw_lng=-'+str(min(l_min))+'&zoom=17&search_by_map=true'

elif typ==2:
	url_start='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=unknown&ne_lat='+str(max(l_max))+'&ne_lng='+str(min(l_min))+'& sw_lat='+str(min(l_max))+'&sw_lng=-'+str(max(l_min))+'&zoom=17&search_by_map=true'
	sheet_write.write(1, 0, url_start)
	sheet_write.write(1, 1, max(l_max))#3 N
	sheet_write.write(1, 2, min(l_min))#4 E
	sheet_write.write(1, 3, min(l_max))#1 S
	sheet_write.write(1, 4, max(l_min))#2 W

else:
	url_start='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=unknown&ne_lat='+str(l_max[0])+'&ne_lng='+str(l_min[0])+'& sw_lat='+str(l_max[1])+'&sw_lng=-'+str(l_min[1])+'&zoom=17&search_by_map=true'
	sheet_write.write(1, 0, url_start)
	sheet_write.write(1, 1, l_max[0])#3 N
	sheet_write.write(1, 2, l_min[0])#4 E
	sheet_write.write(1, 3, l_max[1])#1 S
	sheet_write.write(1, 4, l_min[1])#2 W
	
print(url_start)
workbook.save('myFile'+str(now)+'.xls')

book = xlrd.open_workbook('myFile'+str(now)+'.xls')
copy_book =copy(book)
sheet_write=copy_book.get_sheet(0)
r=2
nrow=2
h=1

#driver = webdriver.Chrome(chrome_options=chrome_options)
driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
driver.set_window_size(1500, 2000)
wait = WebDriverWait(driver, 3)
time.sleep(5)
x_p="//div[@class='_1h559tl']"
while h < nrow:

	print (h)
	book_dup = xlrd.open_workbook('myFile'+str(now)+'.xls')
	sheet_read = book_dup.sheet_by_index(0)
	URL=sheet_read.cell(h,0).value
	print (URL)
	driver.get(URL)
	time.sleep(3)
	wait = WebDriverWait(driver, 3)
	locations='aucune annonce'
	try:
		#print ('ici')
		locations = wait.until(EC.presence_of_element_located((By.XPATH, x_p))).text
		#print (locations)
		n_loc=locations.split(" ")
		#print (n_loc)
		#nb_loc=n_loc[4].replace("+","")
		n_loc.reverse()
		nb_loc=n_loc[1]
		f_loc=float(nb_loc)
		print (nb_loc)
		sheet_write.write(h, 6, nb_loc)
		copy_book.save('myFile'+str(now)+'.xls')
	except:
		pass
	#print (locations)
	if locations=='aucune annonce':
		sheet_write.write(h, 5, 'O')
		sheet_write.write(h, 6, 0)
		copy_book.save('myFile'+str(now)+'.xls')
		f_loc=0
	if f_loc>=300:
	#MARQUAGE X SUR URL 300+
		sheet_write.write(h, 5, "X")
	#FORMULE POUR DIVISION 4 CELLS
		n_t=sheet_read.cell(h,1).value
		e_t=sheet_read.cell(h,2).value
		s_t=sheet_read.cell(h,3).value
		w_t=sheet_read.cell(h,4).value
		#print (n_t)
		#print (e_t)
		#print (s_t)
		#print (w_t)
		div_lat=(Decimal(n_t)-Decimal(s_t))/2
		div_long=(Decimal(e_t)-Decimal(w_t))/2
		#print (div_lat)
		#print (div_long)
	#DIVISION 1
		n_g1=Decimal(n_t)-(div_lat)
		e_g1=Decimal(e_t)-(div_long)
		s_g1=Decimal(s_t)
		w_g1=Decimal(w_t)
		n_t1=str(n_g1)
		e_t1=str(e_g1)
		s_t1=str(s_g1)
		w_t1=str(w_g1)
		sheet_write.write(r, 1, n_t1)
		sheet_write.write(r, 2, e_t1)
		sheet_write.write(r, 3, s_t1)
		sheet_write.write(r, 4, w_t1)
	#DIVISION 2
		r=r+1
		n_g2=Decimal(n_t)
		e_g2=Decimal(w_t)+(div_long)
		s_g2=Decimal(s_t)+(div_lat)
		w_g2=Decimal(w_t)
		n_t2=str(n_g2)
		e_t2=str(e_g2)
		s_t2=str(s_g2)
		w_t2=str(w_g2)
		sheet_write.write(r, 1, n_t2)
		sheet_write.write(r, 2, e_t2)
		sheet_write.write(r, 3, s_t2)
		sheet_write.write(r, 4, w_t2)
	#DIVISION 3
		r=r+1
		n_g3=Decimal(n_t)-(div_lat)
		e_g3=Decimal(e_t)
		s_g3=Decimal(s_t)
		w_g3=Decimal(w_t)+(div_long)
		n_t3=str(n_g3)
		e_t3=str(e_g3)
		s_t3=str(s_g3)
		w_t3=str(w_g3)
		sheet_write.write(r, 1, n_t3)
		sheet_write.write(r, 2, e_t3)
		sheet_write.write(r, 3, s_t3)
		sheet_write.write(r, 4, w_t3)
	#DIVISION 4
		r=r+1
		n_g4=Decimal(n_t)
		e_g4=Decimal(e_t)
		s_g4=Decimal(s_t)+(div_lat)
		w_g4=Decimal(w_t)+(div_long)
		n_t4=str(n_g4)
		e_t4=str(e_g4)
		s_t4=str(s_g4)
		w_t4=str(w_g4)
		sheet_write.write(r, 1, n_t4)
		sheet_write.write(r, 2, e_t4)
		sheet_write.write(r, 3, s_t4)
		sheet_write.write(r, 4, w_t4)
	#CREATION URL DES DIVISIONS
		#url1='https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat='+str(s_g1)+'&sw_lng='+str(w_g1)+'&ne_lat='+str(n_g1)+'&ne_lng='+str(e_g1)+'&zoom=14&search_by_map=true&map_toggle=true'
		#url2='https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat='+str(s_g2)+'&sw_lng='+str(w_g2)+'&ne_lat='+str(n_g2)+'&ne_lng='+str(e_g2)+'&zoom=14&search_by_map=true&map_toggle=true'
		#url3='https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat='+str(s_g3)+'&sw_lng='+str(w_g3)+'&ne_lat='+str(n_g3)+'&ne_lng='+str(e_g3)+'&zoom=14&search_by_map=true&map_toggle=true'
		#url4='https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat='+str(s_g4)+'&sw_lng='+str(w_g4)+'&ne_lat='+str(n_g4)+'&ne_lng='+str(e_g4)+'&zoom=14&search_by_map=true&map_toggle=true'
		url1='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=pagination&ne_lat='+str(n_g1)+'&ne_lng='+str(e_g1)+'&sw_lat='+str(s_g1)+'&sw_lng='+str(w_g1)+'&zoom=17&search_by_map=true'
		url2='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=pagination&ne_lat='+str(n_g2)+'&ne_lng='+str(e_g2)+'&sw_lat='+str(s_g2)+'&sw_lng='+str(w_g2)+'&zoom=17&search_by_map=true'
		url3='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=pagination&ne_lat='+str(n_g3)+'&ne_lng='+str(e_g3)+'&sw_lat='+str(s_g3)+'&sw_lng='+str(w_g3)+'&zoom=17&search_by_map=true'
		url4='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=pagination&ne_lat='+str(n_g4)+'&ne_lng='+str(e_g4)+'&sw_lat='+str(s_g4)+'&sw_lng='+str(w_g4)+'&zoom=17&search_by_map=true'

		sheet_write.write((r-3), 0, url1)
		sheet_write.write((r-2), 0, url2)
		sheet_write.write((r-1), 0, url3)
		sheet_write.write(r, 0, url4)
	#SAVE EXCEL
		copy_book.save('myFile'+str(now)+'.xls')
		nrow=nrow+4
		r=r+1
	h=h+1

print ('_______    ___    ___     ___')
print ('|      |   |  |   |  \    |  |')
print ('|  |__     |  |   |   \   |  |')
print ('|     |    |  |   |    \  |  |')
print ('|  |       |  |   |  |\ \ |  |')
print ('|  |       |  |   |  | \ \|  |')
print ('|__|       |__|   |__|  \____|')
print ('le découpage des zones est terminé, vous pouvez exécuter le .exe N°2 qui va extraire la liste des URL de la totalité des annonces de votre zone.')

xl = pd.ExcelFile('myFile'+str(now)+'.xls')
df = xl.parse("URL_PAGE")
#SUPPRIME les lignes contenant X dans la colonne F
df2 = df.loc[df['F'] != 'X']
list_URL=df2['A'].tolist()
#print (list_URL)
#print (len(list_URL))
time.sleep(1)
#OUVERTURE DES PAGES CHROME
#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
wait = WebDriverWait(driver, 10)
wait2 = WebDriverWait(driver, 5)

#c = ligne 2 du xls
c=1

test='OK'
z=0
while test=='KO':
	driver.get(list_URL[0])
	time.sleep(2)
	try:
		link = wait.until(EC.presence_of_element_located((By.XPATH, '//span[@class="_qlq27g"]['+'1'+']/a')))
		url=link.get_attribute("href")
		test='OK'
	except:
		driver.quit()
		driver = webdriver.Chrome(chrome_options=chrome_options)
		wait = WebDriverWait(driver, 5)
		wait2 = WebDriverWait(driver, 5)
		z=z+1
	if z==10:
		test='OK'

dec=1
list_ann=[]


for h in list_URL:
	print ("DECOMPTE = "+str(dec)+"<<<<<<<<")
	dec=dec+1
	driver.get(h)
	time.sleep(3)
	html = driver.page_source
	soup = BeautifulSoup(html, 'html.parser')
	j=1
	time.sleep(2)
	print(len(list_ann))
	while (j<=17):
		i=0
		try:
			while (i<=20):
				try:
					try:
						#print(c)
					#URL ANNONCE
						#link = wait.until(EC.presence_of_element_located((By.XPATH, XPATH_ann1+str(i)+XPATH_ann2)))
						#link = wait.until(EC.presence_of_element_located((By.XPATH, "//span[@class='_qlq27g']["+str(i)+"]/a")))
						the_tr= soup.findAll('div', attrs={"class": "_8ssblpx"})[i]
						div=the_tr.find('a').attrs
						#link = rootdriver.find_element_by_xpath(XPATH_ann1+str(i)+XPATH_ann2)
						link=div['href']
						#url=link.get_attribute("href")
						v_url=link.split("?")
						#print (v_url[0])
						i=i+1
						
						#sheet_write2.write(c, 1, 'https://www.airbnb.fr'+str(v_url[0]))
						#sheet_write2.cell(row=c+1, column=2).value = 'https://www.airbnb.fr'+str(v_url[0])
						list_ann.append('https://www.airbnb.fr'+str(v_url[0]))
					except:
						#print('NO LINK')
						zzz=1
						#print(str(i)+"--------------"+str(c))
						break
					c=c+1
					#wb.save('RESULT.xls')
					
				except:
					print("dernière annonce")
					
					break
			#print(list_ann)
			nextpage='0'
			#nextpage = wait2.until(EC.presence_of_element_located((By.XPATH, "//li[@class='_r4n1gzb']/a")))
			
			nextpage = wait2.until(EC.presence_of_element_located((By.XPATH, "//a[@aria-label='Suivant']")))

			#nextpage=rootdriver.find_element_by_xpath(XPATH_Next)
			a_nextpage=nextpage.get_attribute("href")
			label=nextpage.get_attribute("aria-label")
			#print (label)
			#if label is None:
			if label=='Suivant':
			#le bouton NextPage n'a pas d'attribut Label, quand il n'y a plus de bouton nextpage, l'attribut Label est présent dans //li[@class='_b8vexar']/a
				driver.get(a_nextpage)
				time.sleep(1)
				html = driver.page_source
				soup = BeautifulSoup(html, 'html.parser')
				time.sleep(1)
				j=j+1
			else:
			#Si label pas NULL, arreter la boucle While j
				j=20
				print ("---------URL suivant-----------")
		except:
			print("FIN")
			#rootdriver.close()
			#driver.close()
			break
	else:
		continue
#print(list_ann)
#workbook2.save('RESULT_AIRBNB.xls')
print ('_______    ___    ___     ___')
print ('|      |   |  |   |  \    |  |')
print ('|  |__     |  |   |   \   |  |')
print ('|     |    |  |   |    \  |  |')
print ('|  |       |  |   |  |\ \ |  |')
print ('|  |       |  |   |  | \ \|  |')
print ('|__|       |__|   |__|  \____|')
print('Exécuter à présent le exe N°3 pour obtenir le détail des annonces ici extraites')

wb = openpyxl.Workbook()
ws = wb.active
#--------------------------------------------------------------------------------------------------------------
#-------------------CREATE LIST URL
#-------------------

c=2
for g in list_ann:
	ws.cell(row=c, column=3).value = g
	c=c+1
wb.save('RESULT_AIRBNB.xlsx')



try:
	driver.quit()
except:
	print('fin')
