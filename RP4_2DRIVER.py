from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import xlwt
import xlrd
import time
from xlutils.copy import copy
import datetime
import datetime as dt
from tkinter import filedialog
from tkinter import *
from bs4 import BeautifulSoup
from urllib.request import urlopen
import os
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
import threading
import sys


chrome_options = webdriver.ChromeOptions()
prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)
#chrome_options.add_argument("-headless")
#chrome_options.add_argument("-disable-gpu")
print ('▀▄▀▄▀▄ STOPBNB ▄▀▄▀▄▀')

#-----EXCEL RESULT OPEN AND READ-----

wbx = load_workbook(path_RESULT.filename)
ws = wbx.active

#-------FIND COLUMN UPDATE------
up=0
k=1
while up==0:
	#V_up=sheet_read.cell(0,i).value
	V_up=ws.cell(row=1, column=k).value
	if V_up=='UPDATE_CALENDAR':
		up=1
	else:
		k=k+1
		
c_mouth=k+1

V_mouth=ws.cell(row=1, column=c_mouth).value
if V_mouth!='3/5_mois':
	ws.insert_cols(c_mouth)
	ws.cell(row=1, column=c_mouth).value = '3/5_mois'

#-----RECUP INFO XPATH FROM EXCEL------
book_GMAIL = xlrd.open_workbook('/home/pi/Desktop/GMAIL_ACCOUNT.xls')
sheet_GMAIL = book_GMAIL.sheet_by_index(0)
ADRESS_GMAIL=sheet_GMAIL.cell(0,1).value
PSW_GMAIL=sheet_GMAIL.cell(1,1).value
RECEIVER=sheet_GMAIL.cell(2,1).value

#-------DATE DU JOUR-------
date = int(datetime.datetime.now().day)
month = int(datetime.datetime.now().month)
Hr=dt.datetime.now().hour

#------RECUP INFO CALANDAR------

def email(DIR2,NAMEFile,now):
	sender = ADRESS_GMAIL
	sender_password = PSW_GMAIL
	receivers = RECEIVER

	s = smtplib.SMTP('smtp.gmail.com', 587)
	s.starttls()
	s.login(sender, sender_password)
	msg = MIMEMultipart()
	msg['From'] = sender
	msg['To'] = receivers
	#msg['Subject'] = "Subject of the Mail- image -2"
	body = "Body_of_the_mail"
	msg.attach(MIMEText(body, 'plain'))
	msg['Subject'] = "STOP AIRBNB - extraction du - "+str(now)
	# path along with extension of file to be attachmented 
	filename = DIR2+NAMEFile+str(now)+".xlsx"
	attachmentment = open(filename, "rb")
	 
	# instance of MIMEBase and named as p
	attachment = MIMEBase('application', 'octet-stream')
	# To change the payload into encoded form
	attachment.set_payload((attachmentment).read())
	# encode into base64
	encoders.encode_base64(attachment)
	attachment.add_header('Content-Disposition', "attachmentment; filename= %s" % filename)
	# attachment the instance  to instance 'msg'
	msg.attach(attachment)
	text = msg.as_string()
	s.sendmail(sender, receivers, text)
	print('*** email sent ***') 
	time.sleep(10)
	del filename
	del attachmentment
	del attachment
	del text
	del msg

def emailfalde():
	sender = ADRESS_GMAIL
	sender_password = PSW_GMAIL
	receivers = 'vincent.aulnay@gmail.com'

	s = smtplib.SMTP('smtp.gmail.com', 587)
	s.starttls()
	s.login(sender, sender_password)
	text = "echec de capture xpath mois"
	s.sendmail(sender, receivers, text)
	print('*** email sent ***') 
	time.sleep(10)
	del text

def emailfalde2():
	sender = ADRESS_GMAIL
	sender_password = PSW_GMAIL
	receivers = 'vincent.aulnay@gmail.com'

	s = smtplib.SMTP('smtp.gmail.com', 587)
	s.starttls()
	s.login(sender, sender_password)
	text = "echec de xpathdate"
	s.sendmail(sender, receivers, text)
	print('*** email sent ***') 
	time.sleep(10)
	del text

def whatmounth():
	month = int(datetime.datetime.now().month)
	global name_mois1
	global name_mois2
	global name_mois3
	global name_mois4
	global name_mois5
	if month==1:
		name_mois1='janvier 2020'
		name_mois2='février 2020'
		name_mois3='mars 2020'
		name_mois4='avril 2020'
		name_mois5='mai 2020'
	elif month==2:
		name_mois1='février 2020'
		name_mois2='mars 2020'
		name_mois3='avril 2020'
		name_mois4='mai 2020'
		name_mois5='juin 2020'
	elif month==3:
		name_mois1='mars 2020'
		name_mois2='avril 2020'
		name_mois3='mai 2020'
		name_mois4='juin 2020'
		name_mois5='juillet 2020'
	elif month==4:
		name_mois1='avril 2020'
		name_mois2='mai 2020'
		name_mois3='juin 2020'
		name_mois4='juillet 2020'
		name_mois5='août 2020'
	elif month==5:
		name_mois1='mai 2020'
		name_mois2='juin 2020'
		name_mois3='juillet 2020'
		name_mois4='août 2020'
		name_mois5='septembre 2020'
	elif month==6:
		name_mois1='juin 2020'
		name_mois2='juillet 2020'
		name_mois3='août 2020'
		name_mois4='septembre 2020'
		name_mois5='octobre 2020'
	elif month==7:
		name_mois1='juillet 2020'
		name_mois2='août 2020'
		name_mois3='septembre 2020'
		name_mois4='octobre 2020'
		name_mois5='novembre 2020'
	elif month==8:
		name_mois1='août 2020'
		name_mois2='septembre 2020'
		name_mois3='octobre 2020'
		name_mois4='novembre 2020'
		name_mois5='décembre 2020'
	elif month==9:
		name_mois1='septembre 2019'
		name_mois2='octobre 2019'
		name_mois3='novembre 2019'
		name_mois4='décembre 2019'
		name_mois5='janvier 2020'
	elif month==10:
		name_mois1='octobre 2019'
		name_mois2='novembre 2019'
		name_mois3='décembre 2019'
		name_mois4='janvier 2020'
		name_mois5='février 2020'
	elif month==11:
		name_mois1='novembre 2019'
		name_mois2='décembre 2019'
		name_mois3='janvier 2020'
		name_mois4='février 2020'
		name_mois5='mars 2020'
	elif month==12:
		name_mois1='décembre 2019'
		name_mois2='janvier 2020'
		name_mois3='fevrier 2020'
		name_mois4='mars 2020'
		name_mois5='avril 2020'

def MnumDay (Mmois):
	global MNumday
	if Mmois=='janvier':
		MNumday=31
	elif Mmois=='février':
		MNumday=28
	elif Mmois=='mars':
		MNumday=31	
	elif Mmois=='avril':
		MNumday=30
	elif Mmois=='mai':
		MNumday=31
	elif Mmois=='juin':
		MNumday=30
	elif Mmois=='juillet':
		MNumday=31
	elif Mmois=='août':
		MNumday=30
	elif Mmois=='septembre':
		MNumday=31
	elif Mmois=='octobre':
		MNumday=30
	elif Mmois=='novembre':
		MNumday=31
	elif Mmois=='décembre':
		MNumday=30
		
def A_Colonne_mois(name_mois,c):
#1- récupération book Result qui évolue au court du script
#2- compter le nombre de colonne
#3- déterminer si colonne == name_mois de airbnb
#4- si condition alors c_write=c pour définir la colonne où écrire
	global c_write
	global new_month
	book_mois = xlrd.open_workbook(path_RESULT.filename, on_demand = True)
	sheet_mois = book_mois.sheet_by_index(0)
	nc=sheet_mois.ncols
	book_mois.release_resources()
	del book_mois
	
	new_month=0
	
	find_month=0
	while find_month==0:
		this_month=ws.cell(row=1, column=c+1).value
		if this_month==name_mois:
			c_write=c+1
			c_stat='rien'
			c_stat=ws.cell(row=1, column=c+2).value
			if c_stat!="STATE":
				ws.insert_cols(c+2)
				ws.cell(row=1, column=c+2).value = 'STATE'
				wbx.save(path_RESULT.filename)
			break
		elif this_month==None:
			ws.cell(row=1, column=c+1).value = name_mois
			ws.cell(row=1, column=c+2).value = 'STATE'
			ws.cell(row=1, column=c+3).value = 'NB_COMMENT'
			ws.cell(row=1, column=c+4).value = 'DIF_COMMENT'
			ws.cell(row=1, column=c+5).value = 'NB_/A'
			ws.cell(row=1, column=c+6).value = 'NB_NO/A'
			ws.cell(row=1, column=c+7).value = 'SUM_NB'
			ws.cell(row=1, column=c+8).value = 'nJ_/A'
			ws.cell(row=1, column=c+9).value = 'nJ_NO/A'
			ws.cell(row=1, column=c+10).value = 'SUM_nJ'
			ws.cell(row=1, column=c+11).value = 'SUM_all_nJ/A'
			ws.cell(row=1, column=c+12).value = 'SUM_all_nJ'
			ws.cell(row=1, column=c+13).value = 'nb_/P'
			ws.cell(row=1, column=c+14).value = 'nJ_/D'
			ws.cell(row=1, column=c+15).value = 'TOTAL_J'
			c_write=c+1
			find_month=1
			new_month=1
			print ('plus une colonne')
			wbx.save(path_RESULT.filename)
			break
		else:
			c=c+1

def A_Statu_PLUS(date,c_write,page,j,g,ResAirbnb,new_mo,MNday,ONCOM):	
	int_timeday=int(date)
	i=0
	li=[]
	#SEMAINE
	tr=1
	#JOUR
	td=1
	#x_ = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_14676s3']/div[2]/div/div["+str(page)+"]//tr[2]/td[6]/span/div/div/div"))).text
	#print (x_)
	ResAirbnb='/R'
	if new_mo==1:
		ResAirbnb='/D'
	while (tr <=5):
		while(td<=7):
			XP_day ="//div[@class='_14676s3']/div[2]/div/div["+str(page)+"]//tr["+str(tr)+"]/td["+str(td)+"]"
			link = rootdriver.find_element_by_xpath(XP_day)
			class_att=link.get_attribute("class")
			if len(class_att) !=0:
				if class_att=='_z39f86g':
					XP_date="//div[@class='_14676s3']/div[2]/div/div["+str(page)+"]//tr["+str(tr)+"]/td["+str(td)+"]/span/div/div/div"
					NM_date = wait.until(EC.presence_of_element_located((By.XPATH, XP_date))).text
					int_date=int(NM_date)
					if int_date>=int_timeday:
						li.append(int_date)
						#print (li)
					else:
						a=1
				else:
					a=1
			else:
				a=1
			td=td+1
		tr=tr+1
		td=1
	back_li=ws.cell(row=j, column=c_write+1).value
	if back_li!=None:
		back_li=back_li.replace("[","")
		back_li=back_li.replace("]","")
		back_li=back_li.split(",")
		i=0
		bl=[]
		while i!=len(back_li):
			ivb=int(back_li[i])
			if ivb>=int_timeday:
				bl.append(ivb)
			i=i+1
		back_li=bl
		#print ("back_li="+str(back_li))
	else:
		back_li=[]
	ws.cell(row=j, column=c_write+1).value = str(li)
	#print(li)
	c_added=[]
	c_remove=[]
	c_added=[elem for elem in li if elem not in back_li ]
	c_remove=[elem for elem in back_li if elem not in li ]
	#print(c_added)
	#print(c_remove)
	date = int(datetime.datetime.now().day)
	month = int(datetime.datetime.now().month)
	toto=str(date)+'-'+str(month)
	t_add='vide'
	t_rem='vide'
	t_wri='vide'
	if len(c_added)>0:
		if len(c_added)==1:
			dif=c_added[0]-date
			if dif==0 or dif==1 or dif==2 or dif==6:
				ResAirbnb='/P'
			elif dif<0:
				difP=MNday-date+lie[0]
				if difP==0 or difP==1 or difP==2 or difP==6:
					ResAirbnb='/P'		
		t_add=ResAirbnb+toto+':'+str(c_added)
		t_add=t_add.replace("[","")
		t_add=t_add.replace("]","")
		#print (t_add)
	if len(c_remove)>0:
		t_rem='/L'+toto+':'+str(c_remove)
		t_rem=t_rem.replace("[","")
		t_rem=t_rem.replace("]","")
		#print(t_rem)
	ca=ws.cell(row=j, column=c_write).value
	if ca==None:
		if t_add!='vide':
			t_wri=str(t_add)
	else:
		if t_add!='vide':
			if t_rem!='vide':
				t_wri=str(t_add)+';    '+str(t_rem)
			else:
				t_wri=str(t_add)
		else:
			if t_rem!='vide':
				t_wri=str(t_rem)
		if t_wri!='vide':
			t_wri=str(ca)+';    '+t_wri
	if t_wri!='vide':
		print(t_wri)
		ws.cell(row=j, column=c_write).value=t_wri
	#COMMENTAIRE
	ONC=ONCOM
	if ONC==1:
		try:
			#//span[@class='_so3dpm2']
			#Bcomment=soup.find('button', attrs={"class": "_ff6jfq"})
			#Scomment=Bcomment.find('span', attrs={"class": "_so3dpm2"}).text
			Lcomment=[]
			Icomment=0
			Scomment=soup.find('span', attrs={"class": "_1oftqjo1"}).text
			Scomment=Scomment.replace("(","")
			Scomment=Scomment.replace(")","")
			Scomment=Scomment.replace(" ","")
			Lcomment=Scomment.split("c")
			Icomment=int(Lcomment[0])
			ws.cell(row=j, column=c_write+2).value=S=Icomment
		except:
			pass

def A_Statu_PLUS2(c_write,j,ResAirbnb,new_mo,page):	
	i=0
	li=[]
	#SEMAINE
	tr=1
	#JOUR
	td=1
	if new_mo==1:
		ResAirbnb='/D'
	while (tr <=5):
		while(td<=7):
			XP_day ="//div[@class='_14676s3']/div[2]/div/div["+str(page)+"]//tr["+str(tr)+"]/td["+str(td)+"]"
			link = rootdriver.find_element_by_xpath(XP_day)
			class_att=link.get_attribute("class")
			if len(class_att) !=0:
				if class_att=='_z39f86g':
					XP_date="//div[@class='_14676s3']/div[2]/div/div["+str(page)+"]//tr["+str(tr)+"]/td["+str(td)+"]/span/div/div/div"
					NM_date = wait.until(EC.presence_of_element_located((By.XPATH, XP_date))).text
					int_date=int(NM_date)
					li.append(int_date)
					#print (li)
				else:
					a=1
			else:
				a=1
			td=td+1
		tr=tr+1
		td=1
	back_li=ws.cell(row=j, column=c_write+1).value
	if back_li!=None:
		back_li=back_li.replace("[","")
		back_li=back_li.replace("]","")
		back_li=back_li.split(",")
		i=0
		bl=[]
		while i!=len(back_li):
			ivb=int(back_li[i])
			bl.append(ivb)
			i=i+1
		back_li=bl
		#print ("back_li="+str(back_li))
	else:
		back_li=[]
	ws.cell(row=j, column=c_write+1).value = str(li)
	#print(li)
	c_added=[]
	c_remove=[]
	c_added=[elem for elem in li if elem not in back_li ]
	c_remove=[elem for elem in back_li if elem not in li ]
	#print(c_added)
	#print(c_remove)
	date = int(datetime.datetime.now().day)
	month = int(datetime.datetime.now().month)
	toto=str(date)+'-'+str(month)
	t_add='vide'
	t_rem='vide'
	t_wri='vide'
	if len(c_added)>0:		
		t_add=ResAirbnb+toto+':'+str(c_added)
		t_add=t_add.replace("[","")
		t_add=t_add.replace("]","")
		print (t_add)
	if len(c_remove)>0:
		t_rem='/L'+toto+':'+str(c_remove)
		t_rem=t_rem.replace("[","")
		t_rem=t_rem.replace("]","")
		print(t_rem)
	ca=ws.cell(row=j, column=c_write).value
	if ca==None:
		if t_add!='vide':
			t_wri=str(t_add)
	else:
		if t_add!='vide':
			if t_rem!='vide':
				t_wri=str(t_add)+';    '+str(t_rem)
			else:
				t_wri=str(t_add)
		else:
			if t_rem!='vide':
				t_wri=str(t_rem)
		if t_wri!='vide':
			t_wri=str(ca)+';    '+t_wri
	if t_wri!='vide':
		#print(t_wri)
		ws.cell(row=j, column=c_write).value=t_wri
			
def A_Statu_day2(date,c_write,page,j,g,ResAirbnb,new_mo,MNday,ONCOM):	
	int_timeday=int(date)
	month=soup.findAll('div', attrs={"class":u"_1lds9wb"})[g]
	i=0
	li=[]
	ResAirbnb='/R'
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			the_tr= month.findAll('td', attrs={"class": "_z39f86g"})[i]
			div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
			#_1lds9wb
			intdiv=int(div)
			if intdiv>=int_timeday:
				li.append(intdiv)
			i=i+1
		except:
			break
	back_li=ws.cell(row=j, column=c_write+1).value
	if back_li!=None:
		back_li=back_li.replace("[","")
		back_li=back_li.replace("]","")
		back_li=back_li.split(",")
		i=0
		bl=[]
		while i!=len(back_li):
			ivb=int(back_li[i])
			if ivb>=int_timeday:
				bl.append(ivb)
			i=i+1
		back_li=bl
		#print ("back_li="+str(back_li))
	else:
		back_li=[]
	ws.cell(row=j, column=c_write+1).value = str(li)
	#print(li)
	c_added=[]
	c_remove=[]
	c_added=[elem for elem in li if elem not in back_li ]
	c_remove=[elem for elem in back_li if elem not in li ]
	#print(c_added)
	#print(c_remove)
	date = int(datetime.datetime.now().day)
	month = int(datetime.datetime.now().month)
	toto=str(date)+'-'+str(month)
	t_add='vide'
	t_rem='vide'
	t_wri='vide'
	if len(c_added)>0:
		if len(c_added)==1:
			dif=c_added[0]-date
			if dif==0 or dif==1 or dif==2 or dif==6:
				ResAirbnb='/P'
			elif dif<0:
				difP=MNday-date+lie[0]
				if difP==0 or difP==1 or difP==2 or difP==6:
					ResAirbnb='/P'		
		t_add=ResAirbnb+toto+':'+str(c_added)
		t_add=t_add.replace("[","")
		t_add=t_add.replace("]","")
		#print (t_add)
	if c_remove!=['']:
		if c_remove!=[]:
			t_rem='/L'+toto+':'+str(c_remove)
			t_rem=t_rem.replace("[","")
			t_rem=t_rem.replace("]","")
			#print(t_rem)
	ca=ws.cell(row=j, column=c_write).value
	if ca==None:
		if t_add!='vide':
			t_wri=str(t_add)
	else:
		if t_add!='vide':
			if t_rem!='vide':
				t_wri=str(t_add)+';    '+str(t_rem)
			else:
				t_wri=str(t_add)
		else:
			if t_rem!='vide':
				t_wri=str(t_rem)
		if t_wri!='vide':
			t_wri=str(ca)+';    '+t_wri
	if t_wri!='vide':
		print(t_wri)
		ws.cell(row=j, column=c_write).value=t_wri

	#COMMENTAIRE
	ONC=ONCOM
	if ONC==1:
		try:
			#//span[@class='_so3dpm2']
			#Bcomment=soup.find('button', attrs={"class": "_ff6jfq"})
			#Scomment=Bcomment.find('span', attrs={"class": "_so3dpm2"}).text
			Lcomment=[]
			Scomment=soup.find('span', attrs={"class": "_1plk0jz1"}).text
			Scomment=Scomment.replace("(","")
			Scomment=Scomment.replace(")","")
			Scomment=Scomment.replace(" ","")
			Lcomment=Scomment.split("c")
			Icomment=int(Lcomment[0])
			ws.cell(row=j, column=c_write+2).value=S=Icomment
		except:
			pass

def A_Statu_day4(c_write,j,ResAirbnb,new_mo):
	month5=soup.find('div', attrs={"class":u"_kuxo8ai"})
	i=0
	li=[]
	ResAirbnb='/R'
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			the_tr= month5.findAll('td', attrs={"class": "_z39f86g"})[i]
			div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
			intdiv=int(div)
			li.append(intdiv)
			i=i+1
		except:
			break
	#print(li)
	back_li=ws.cell(row=j, column=c_write+1).value
	if back_li!=None:
		back_li=back_li.replace("[","")
		back_li=back_li.replace("]","")
		back_li=back_li.split(",")
		i=0
		bl=[]
		while i!=len(back_li):
			ivb=int(back_li[i])
			bl.append(ivb)
			i=i+1
		back_li=bl
		#print ("back_li="+str(back_li))
	else:
		back_li=[]
	ws.cell(row=j, column=c_write+1).value = str(li)
	#print(li)
	c_added=[]
	c_remove=[]
	c_added=[elem for elem in li if elem not in back_li ]
	c_remove=[elem for elem in back_li if elem not in li ]
	#print(c_added)
	#print(c_remove)
	date = int(datetime.datetime.now().day)
	month = int(datetime.datetime.now().month)
	toto=str(date)+'-'+str(month)
	t_add='vide'
	t_rem='vide'
	t_wri='vide'
	if len(c_added)>0:		
		t_add=ResAirbnb+toto+':'+str(c_added)
		t_add=t_add.replace("[","")
		t_add=t_add.replace("]","")
		#print (t_add)
	if c_remove!=['']:
		if c_remove!=[]:
			t_rem='/L'+toto+':'+str(c_remove)
			t_rem=t_rem.replace("[","")
			t_rem=t_rem.replace("]","")
			#print(t_rem)
	ca=ws.cell(row=j, column=c_write).value
	if ca==None:
		if t_add!='vide':
			t_wri=str(t_add)
	else:
		if t_add!='vide':
			if t_rem!='vide':
				t_wri=str(t_add)+';    '+str(t_rem)
			else:
				t_wri=str(t_add)
		else:
			if t_rem!='vide':
				t_wri=str(t_rem)
		if t_wri!='vide':
			t_wri=str(ca)+';    '+t_wri
	if t_wri!='vide':
		#print(t_wri)
		ws.cell(row=j, column=c_write).value=t_wri


def A_Statu_day5(c_write,j,ResAirbnb,new_mo,g):	
	month5=soup.findAll('div', attrs={"class":u"_1lds9wb"})[g]
	i=0
	li=[]
	ResAirbnb='/R'
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			the_tr= month5.findAll('td', attrs={"class": "_z39f86g"})[i]
			div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
			intdiv=int(div)
			li.append(intdiv)
			i=i+1
		except:
			break
	#print (li)
	back_li=ws.cell(row=j, column=c_write+1).value
	if back_li!=None:
		back_li=back_li.replace("[","")
		back_li=back_li.replace("]","")
		back_li=back_li.split(",")
		i=0
		bl=[]
		while i!=len(back_li):
			ivb=int(back_li[i])
			bl.append(ivb)
			i=i+1
		back_li=bl
		#print ("back_li="+str(back_li))
	else:
		back_li=[]
	ws.cell(row=j, column=c_write+1).value = str(li)
	#print(li)
	c_added=[]
	c_remove=[]
	c_added=[elem for elem in li if elem not in back_li ]
	c_remove=[elem for elem in back_li if elem not in li ]
	#print(c_added)
	#print(c_remove)
	date = int(datetime.datetime.now().day)
	month = int(datetime.datetime.now().month)
	toto=str(date)+'-'+str(month)
	t_add='vide'
	t_rem='vide'
	t_wri='vide'
	if len(c_added)>0:		
		t_add=ResAirbnb+toto+':'+str(c_added)
		t_add=t_add.replace("[","")
		t_add=t_add.replace("]","")
		#print (t_add)
	if c_remove!=['']:
		if c_remove!=[]:
			t_rem='/L'+toto+':'+str(c_remove)
			t_rem=t_rem.replace("[","")
			t_rem=t_rem.replace("]","")
			#print(t_rem)
	ca=ws.cell(row=j, column=c_write).value
	if ca==None:
		if t_add!='vide':
			t_wri=str(t_add)
	else:
		if t_add!='vide':
			if t_rem!='vide':
				t_wri=str(t_add)+';    '+str(t_rem)
			else:
				t_wri=str(t_add)
		else:
			if t_rem!='vide':
				t_wri=str(t_rem)
		if t_wri!='vide':
			t_wri=str(ca)+';    '+t_wri
	if t_wri!='vide':
		#print(t_wri)
		ws.cell(row=j, column=c_write).value=t_wri
	

def COMPUTE_M1(name_mois1):
	Dif_c=1
	if Dif_c==1:
		up=0
		i=1
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up==name_mois1:
				up=1
			else:
				i=i+1
		#print('Cmois='+str(i))
		Cmois=i

		up=0
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NB_COMMENT':
				up=1
			else:
				i=i+1
		#print('Ccomment1='+str(i))
		Ccomment1=i

		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='DIF_COMMENT':
				up=1
			else:
				i=i+1
		#print('DIF_Comment='+str(i))
		DIF_Comment=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NB_/A':
				up=1
			else:
				i=i+1
		C_nbA=i

		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='NB_NO/A':
				up=1
			else:
				i=i+1
		C_nbnoA=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SUM_NB':
				up=1
			else:
				i=i+1
		C_SUMnb=i

		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='nJ_/A':
				up=1
			else:
				i=i+1
		C_nJA=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='nJ_NO/A':
				up=1
			else:
				i=i+1
		C_NOnJA=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SUM_nJ':
				up=1
			else:
				i=i+1
		#print('SUM_nJ='+str(i))
		C_SUMnJ=i

		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='SUM_all_nJ':
				up=1
			else:
				i=i+1
		C_SUM_all_nJ=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='nb_/P':
				up=1
			else:
				i=i+1
		C_nb_P=i
		
		up=0
		i=Cmois
		while up==0:
			V_up=ws.cell(row=1, column=i).value
			if V_up=='nJ_/D':
				up=1
			else:
				i=i+1
		C_nJD=i
		
		up=0
		i=Cmois
		try:
			while up==0:
				V_up=ws.cell(row=1, column=i).value
				if V_up=='NB_COMMENT':
					up=1
				else:
					i=i-1
			#print('Ccommont2='+str(i))
			Ccomment2=i
			NOC2=0
		except:
			NOC2=1
			#print ('NOC2=====1')
	c=2
	while c<=nrow:
		if NOC2==0:
			V1=ws.cell(row=c, column=Ccomment1).value
			V2=ws.cell(row=c, column=Ccomment2).value
			try:
				DIF=int(V1)-int(V2)
				#print('ANNONCE:'+str(c)+('   DIF:')+str(DIF))
				ws.cell(row=c, column=DIF_Comment).value=DIF
			except:
				pass
	#--------COUNT NB/A and NB NO/A---------
		STR_NBA=ws.cell(row=c, column=Cmois).value
		continu=1
		if STR_NBA==None:
			continu=0
		if continu==1:
			count_AP=0
			count_D=0
			count_P=0
			count=0
			count_NBA=0
			count_R=0
			count_L=0
			count_RP=0
			count_RP=STR_NBA.count('/R/P')
			count_AP=STR_NBA.count('/A/P')
			count_NBA=STR_NBA.count('/A')
			real_NBA=count_NBA-count_AP
			count_P=STR_NBA.count('/P')
			count_D=STR_NBA.count('/D')
			count=STR_NBA.count(':')
			count_R=STR_NBA.count('/R')
			count_L=STR_NBA.count('/L')
			
			NBNOA=count-count_D-count_NBA-count_P-count_L
			#print (('NB_NO/A ===')+str(NBNOA))
			ws.cell(row=c, column=C_nbA).value=real_NBA
			ws.cell(row=c, column=C_nbnoA).value=NBNOA
			ws.cell(row=c, column=C_nb_P).value=count_P+count_AP+count_RP
			write=int(NBNOA)+int(real_NBA)
			ws.cell(row=c, column=C_SUMnb).value=write
		#---------COUNT nJ ---------
			list=STR_NBA.split(';')
			B=['/P', '/D', '/A/P', '/R', '/L', '/R/P']
			blacklist = re.compile('|'.join([re.escape(word) for word in B]))
			newL=[word for word in list if not blacklist.search(word)]
			D=['/D']
			blacklistD = re.compile('|'.join([re.escape(wordD) for wordD in D]))
			newLforD=[wordD for wordD in list if blacklistD.search(wordD)]
			rd=0
			lenD=len(newLforD)
			nbD=0
			while rd<lenD:
				pnlD=newLforD[rd].split(':')
				del pnlD[0]
				pld=pnlD[0].split(',')
				nbD=nbD+len(pld)
				rd=rd+1
			ws.cell(row=c, column=C_nJD).value=nbD
			#[x for x in list if not x.startswith('/A/P') and not x.startswith('/D') and not x.startswith('/P')]
			#[x for x in list if not any(bad in x for bad in B)]
			#-----/A--------
			BA=['/A']
			blacklistA = re.compile('|'.join([re.escape(wordA) for wordA in BA]))
			newLforA=[wordA for wordA in newL if blacklistA.search(wordA)] #-------Creation list AVEC que les lot /A
			newLfornoA=[wordA for wordA in newL if not blacklistA.search(wordA)] #-------Creation list SANS les lot /A
			nAlen=len(newLforA)
			rr=0
			nbA=0
			try: #---Recuperation nJ dans les lot /A
				while rr<nAlen:
					pnlA=newLforA[rr].split(':')
					del pnlA[0]
					pla=pnlA[0].split(',')
					nbA=nbA+len(pla)
					rr=rr+1
			except:
				pass
			ws.cell(row=c, column=C_nJA).value=nbA
			nAlen=len(newLfornoA)
			rr=0
			NnoJA=0
			try: #---Recuperation nJ dans les lot SANS /A
				while rr<nAlen:
					pnlA=newLfornoA[rr].split(':')
					del pnlA[0]
					pla=pnlA[0].split(',')
					NnoJA=NnoJA+len(pla)
					rr=rr+1
			except:
				pass
			ws.cell(row=c, column=C_NOnJA).value=NnoJA
			write=int(nbA)+int(NnoJA)
			ws.cell(row=c, column=C_SUMnJ).value=write
		c=c+1
	wbx.save(path_RESULT.filename)

		
#-----OPEN GOOGLE CHROME and AIRBNB PAGE---------

rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
#rootdriver.set_page_load_timeout(2)
rootdriver.set_window_size(2000, 1000)
wait = WebDriverWait(rootdriver, 5)

rootdriver2 = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
#rootdriver2 = webdriver.Chrome(chrome_options=chrome_options)
#rootdriver2.set_page_load_timeout(2)
rootdriver2.set_window_size(2000, 1000)
wait2 = WebDriverWait(rootdriver2, 5)

nrow=ws.max_row
print('NROW'+str(nrow))
j=2
z=0
end=0
ee=0
e=0
Tr=0
C_mois=0
C_mois5=0
drive=0
date = int(datetime.datetime.now().day)
f_mounth=1
fm=3
fff=0
f_xpathdate=0
w_month=0
c_month=0
while f_mounth==0:
	h=ws.cell(row=fm, column=2).value
	print(h)
	if fff==2:
		f_mounth=1
		f_xpathdate=1
		end=1
		run=emailfalde()
	fff=fff+1
	try:
		print('ici1')
		rootdriver.get(h)
		time.sleep(2)
		html = rootdriver.page_source
		soup = BeautifulSoup(html, 'html.parser')
		time.sleep(3)
		print('ici2')
		#---mois 1---
		name_mois1 = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1lds9wb'][1]//div[@class='_gucugi']/strong"))).text
		name_mois2 = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1lds9wb'][2]//div[@class='_gucugi']/strong"))).text
		#name_mois3 = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_gucugi'][3]/strong"))).text
		print('ici3')
		print(name_mois1)
		print(name_mois2)
		#print(name_mois3)
		mm3=0
		u=0
		while mm3==0:
			print('la')
			month311=soup.find('div', attrs={"class":u"_kuxo8ai"})
			month31=month311.find('div', attrs={"class":u"_gucugi"})
			name_mois3=month31.find('strong').text
			
			print ('name m3='+str(name_mois3))
			if name_mois3==None:
				mm3=0
				u=u+1
				print ('none')
			else:
				mm3=1
				print ('m3 yes')
			if u==3:
				mm3=1
			print ('go')
		
		print(name_mois1)
		Mname1=name_mois1.split(' ')
		MN1=Mname1[0]
		run_MN=MnumDay(MN1)
		print (MNumday)
		MNday1=MNumday
		run_c=A_Colonne_mois(name_mois1,k)
		m1_write=c_write
		m1_newmonth=new_month
		#---mois 2---
		
		print(name_mois2)
		Mname2=name_mois2.split(' ')
		MN2=Mname2[0]
		run_MN=MnumDay(MN2)
		print (MNumday)
		MNday2=MNumday
		run_c=A_Colonne_mois(name_mois2,k)
		m2_write=c_write
		m2_newmonth=new_month
		#---mois 3---
		#month31=soup.findAll('div', attrs={"class":u"_gucugi"})[3]
		#name_mois3=month31.find('strong').text

		run_c=A_Colonne_mois(name_mois3,k)
		m3_write=c_write
		m3_newmonth=new_month
		
		#---mois 4 et 5---
		print('TRY CLICK')
		ele=rootdriver.find_element_by_xpath("//div[@aria-label='Avancez pour passer au mois suivant.']")
		print('TRY FAILED')
		rootdriver.execute_script("arguments[0].scrollIntoView(true);", ele)
		rootdriver.execute_script("window.scrollBy(0,-500);")
		next_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Avancez pour passer au mois suivant.']")))
		next_calendar.click()
		time.sleep(2)
		next_calendar.click()
		time.sleep(2)
		next_calendar.click()
		time.sleep(2)
		html = rootdriver.page_source
		soup = BeautifulSoup(html, 'html.parser')
		time.sleep(1)
		#-----RECUPERATION CALANDAR MOIS 4--------
		if C_mois5==0:
			name_mois4 = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1lds9wb'][1]//div[@class='_gucugi']/strong"))).text
			print(name_mois4)
			Mname4=name_mois4.split(' ')
			MN4=Mname4[0]
			run_MN=MnumDay(MN4)
			print (MNumday)
			MNday4=MNumday
			run_c=A_Colonne_mois(name_mois4,k)
			m4_write=c_write
			m4_newmonth=new_month
		if C_mois5==0:
			name_mois5 = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1lds9wb'][2]//div[@class='_gucugi']/strong"))).text
			print(name_mois5)
			Mname5=name_mois5.split(' ')
			MN5=Mname5[0]
			run_MN=MnumDay(MN5)
			print (MNumday)
			MNday5=MNumday
			run_c=A_Colonne_mois(name_mois5,k)
			m5_write=c_write
			m5_newmonth=new_month
		f_mounth=1
		j=fm
	except:
		fm=fm+1



while w_month==0:
	run_month=whatmounth()
	print (name_mois1)
	print (name_mois2)
	print (name_mois3)
	print (name_mois4)
	print (name_mois5)
	w_month=1

while c_month==0:
#--mois1--
	print(name_mois1)
	Mname1=name_mois1.split(' ')
	MN1=Mname1[0]
	run_MN=MnumDay(MN1)
	MNday1=MNumday
	run_c=A_Colonne_mois(name_mois1,k)
	m1_write=c_write
	m1_newmonth=new_month
	print (m1_newmonth)
#--mois 2--
	print(name_mois2)
	Mname2=name_mois2.split(' ')
	MN2=Mname2[0]
	run_MN=MnumDay(MN2)
	MNday2=MNumday
	run_c=A_Colonne_mois(name_mois2,k)
	m2_write=c_write
	m2_newmonth=new_month
	print (m2_newmonth)
#--mois 3--
	print(name_mois3)
	Mname3=name_mois3.split(' ')
	MN3=Mname3[0]
	run_MN=MnumDay(MN3)
	MNday3=MNumday
	run_c=A_Colonne_mois(name_mois3,k)
	m3_write=c_write
	m3_newmonth=new_month
	print (m3_newmonth)
#--mois 4--
	print(name_mois4)
	Mname4=name_mois4.split(' ')
	MN4=Mname4[0]
	run_MN=MnumDay(MN4)
	MNday4=MNumday
	run_c=A_Colonne_mois(name_mois4,k)
	m4_write=c_write
	m4_newmonth=new_month
	print (m4_newmonth)
#--mois 5--
	print(name_mois5)
	Mname5=name_mois5.split(' ')
	MN5=Mname5[0]
	run_MN=MnumDay(MN5)
	MNday5=MNumday
	run_c=A_Colonne_mois(name_mois5,k)
	m5_write=c_write
	m5_newmonth=new_month
	print (m5_newmonth)
	time.sleep(5)
	c_month=1

while f_xpathdate==0:
	h=ws.cell(row=fm, column=2).value
	print(h)
	if fff==5:
		f_mounth=1
		f_xpathdate=1
		end=0
		#run=emailfalde2()
	fff=fff+1
	try:
		rootdriver.get(h)
		time.sleep(4)
		x_date = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_13m7kz7i']"))).text
		print("x date trouve")
		f_xpathdate=1
		b_cookie = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@class='optanon-allow-all accept-cookies-button']")))
		b_cookie.click()
	except:
		if fff!=5:
			rootdriver.quit()
			rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
			#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
			rootdriver.set_window_size(2000, 1000)
			wait = WebDriverWait(rootdriver, 5)

def f1(a):
	global hok
	print('F1')
	rootdriver.get(a)
	time.sleep(3)
	try:
		rootdriver.execute_script("window.scrollBy(0,1000);")
		time.sleep(1)
		print('DOWN 01')
		ele=rootdriver.find_element_by_xpath("//div[@aria-label='Avancez pour passer au mois suivant.']")
		print('DOWN 1')
		rootdriver.execute_script("arguments[0].scrollIntoView(true);", ele)
		print('DOWN 1.1')
		rootdriver.execute_script("window.scrollBy(0,-200);")
		time.sleep(2)
	except:
		time.sleep(2)
		print('DOWN KO')
	hok=1
def f2(a):
	global hok2
	print('F2')
	rootdriver2.get(a)
	time.sleep(3)
	try:
		rootdriver2.execute_script("window.scrollBy(0,1000);")
		time.sleep(1)
		print('DOWN 02')
		ele2=rootdriver2.find_element_by_xpath("//div[@aria-label='Avancez pour passer au mois suivant.']")
		print('DOWN 2')
		rootdriver2.execute_script("arguments[0].scrollIntoView(true);", ele2)
		print('DOWN 2.2')
		rootdriver2.execute_script("window.scrollBy(0,-200);")
		time.sleep(2)
	except:
		time.sleep(2)
		print('DOWN KO')
	hok2=1
hok=1
hok2=1
h1=ws.cell(row=2, column=2).value
h2=ws.cell(row=3, column=2).value
rootdriver2.get(h2)
rootdriver.get(h1)
time.sleep(5)
while end==0:
	try:
		while j<=nrow:
			h=ws.cell(row=j+1, column=2).value
			print('------'+str(j-1)+'------'+str(h))
			if h==None:
				j=j+1
				print('h=None')
			elif 'plus' in h:
				ResAirbnb=''
				rootdriver.get(h)
				rootdriver.execute_script("window.scrollBy(0,100);")
				time.sleep(2)
				try:
					b_add_date = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@class='_3uatz29']")))
					b_add_date.click()
					time.sleep(1)
					b_arrival = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@class='_153lip8'][1]")))
					b_arrival.click()
					time.sleep(3)
					#(date,c_write,page,j,g,ResAirbnb,new_mo,MNday,ONCOM)
					run_PLUS_1=A_Statu_PLUS(date,m1_write,2,j,0,ResAirbnb,m1_newmonth,500,1)
					ResAirbnb=''
					b_next = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1h5uiygl']")))
					b_next.click()
					time.sleep(1)
					run_PLUS_2=A_Statu_PLUS(1,m2_write,2,j,1,ResAirbnb,m2_newmonth,MNday1,0)
					b_next = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1h5uiygl']")))
					b_next.click()
					rootdriver.execute_script("window.scrollBy(0,-100);")
					time.sleep(1)
					run_PLUS_3=A_Statu_PLUS2(m3_write,j,ResAirbnb,m3_newmonth,2)
					b_next = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1h5uiygl']")))
					b_next.click()
					rootdriver.execute_script("window.scrollBy(0,-100);")
					time.sleep(1)
					run_PLUS_4=A_Statu_PLUS2(m4_write,j,ResAirbnb,m4_newmonth,2)
					b_next = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1h5uiygl']")))
					b_next.click()
					rootdriver.execute_script("window.scrollBy(0,-100);")
					time.sleep(1)
					run_PLUS_5=A_Statu_PLUS2(m5_write,j,ResAirbnb,m5_newmonth,2)
					#https://www.airbnb.fr/rooms/plus/21846063
				except:
					print('ECHEC PLUS')
				if (j/10).is_integer():
					wbx.save(path_RESULT.filename)
				j=j+1
			elif 'airbnb' in h:
				if j%2==0:
					threading.Thread(target=f1, args=(h,)).start()
					while hok2!=1:
						time.sleep(1)
						ee=ee+1
						if ee==20:
							hok2=1
					hok2=0
					html = rootdriver2.page_source
					soup = BeautifulSoup(html, 'html.parser')
				else:
					threading.Thread(target=f2, args=(h,)).start()
					while hok!=1:
						time.sleep(1)
						e=e+1
						if e==20:
							hok=1
					hok=0
					html = rootdriver.page_source
					soup = BeautifulSoup(html, 'html.parser')
				ResAirbnb=''
				#V_up=ws.cell(row=j, column=k).value
				#v_m=ws.cell(row=j, column=c_mouth).value
				#try:
				#	script=soup.find('script', attrs={"data-state":u"true"}).text
				#	p1=script.split("calendar_last")
				#	p2=p1[1].split("guest_controls")
				#	p3=p2[0].replace('_updated_at":"', '')
				#	p4=p3.replace('","', '')
					#print (p4)
				#	if p4==V_up:
						#ResAirbnb='/A'
				#		ResAirbnb=''
				#	else:
						#ws.cell(row=j, column=k).value=p4
				#		ResAirbnb=''
				#except:
				#	pass
				try:
				#-----RECUPERATION CALANDAR MOIS 1--------
					#print('le mois N est '+name_mois1)
					run_day=A_Statu_day2(date,m1_write,1,j,0,ResAirbnb,m1_newmonth,500,1)
				except:
					pass
				try:
				#-----RECUPERATION CALANDAR MOIS 2--------
					#print('le mois N+1 est '+name_mois2)
					run_day=A_Statu_day2(1,m2_write,2,j,1,ResAirbnb,m2_newmonth,MNday1,0)
				except:
					pass
				try:
				#-----RECUPERATION CALANDAR MOIS 3--------
					#print('le mois N+2 est '+name_mois3)
					run_resday=A_Statu_day4(m3_write,j,ResAirbnb,m3_newmonth)
				except:
					#print('PAS DE MOIS 3')
					pass
			#-----MOIS 4-5 -----
				try:
					if j%2==0:
						next_calendar = wait2.until(EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Avancez pour passer au mois suivant.']")))
						next_calendar.click()
						time.sleep(1)
						next_calendar.click()
						time.sleep(1)
						next_calendar.click()
						time.sleep(2)
						html = rootdriver2.page_source
						soup = BeautifulSoup(html, 'html.parser')
						time.sleep(1)
					else:
						next_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Avancez pour passer au mois suivant.']")))
						next_calendar.click()
						time.sleep(1)
						next_calendar.click()
						time.sleep(1)
						next_calendar.click()
						time.sleep(2)
						html = rootdriver.page_source
						soup = BeautifulSoup(html, 'html.parser')
						time.sleep(1)
					try:
					#-----RECUPERATION CALANDAR MOIS 4--------
						#print('le mois N est '+name_mois4)
						run_day=A_Statu_day5(m4_write,j,ResAirbnb,m4_newmonth,0)
					except:
						pass
				#-----RECUPERATION CALANDAR MOIS 5--------
					try:
						#print('le mois N+1 est '+name_mois5)
						RA4=ResAirbnb
						if date==1:
							RA4='/D'
						run_day=A_Statu_day5(m5_write,j,RA4,m5_newmonth,1)
					except:
						pass
				except:
					print('----click KO')
					pass
				if (j/10).is_integer():
					wbx.save(path_RESULT.filename)
				j=j+1
			elif 'abritel' in h:
				j=j+1
			else:
				j=j+1
		
		wbx.save(path_RESULT.filename)
		end=1
		now = str(datetime.datetime.now())[:19]
		now = now.replace(":","_")
		Tr=date
		print ('FIN')
		wbx = load_workbook(path_RESULT.filename)
		ws = wbx.active
		COMPUTE_M1(name_mois1)
		COMPUTE_M1(name_mois2)
		wbx.save(DIR2+NAMEFile+str(now)+".xlsx")
		try:
			run=email(DIR2,NAMEFile,now)
			print('sent email')
		except:
			print('rien')
		rootdriver.quit()
		wbx.close()
	except:
		j=j+1
		try:
			rootdriver.quit()
		except:
			pass
		# EXCEPT si Chrome se ferme tout seul, ici il va le réouvrir et relancer la boucle d'extraction
		rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
		#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
		rootdriver.set_window_size(2000, 1000)
		wait = WebDriverWait(rootdriver, 5)
		f_xpathdate=0
		fff=0
		while f_xpathdate==0:
			h=ws.cell(row=fm, column=2).value
			print(h)
			if fff==5:
				f_mounth=1
				f_xpathdate=1
				end=0
				#run=emailfalde2()
			fff=fff+1
			try:
				rootdriver.get(h)
				time.sleep(4)
				x_date = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_13m7kz7i']"))).text
				print("x date trouve")
				f_xpathdate=1
				b_cookie = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@class='optanon-allow-all accept-cookies-button']")))
				b_cookie.click()
			except:
				if fff!=5:
					rootdriver.quit()
					rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
					#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
					rootdriver.set_window_size(2000, 1000)
					wait = WebDriverWait(rootdriver, 5)

		
		
try:
	rootdriver.quit()
except:
	pass
print('FIN')
