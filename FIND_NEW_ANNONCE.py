import xlwt
import xlrd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from xlutils.copy import copy
from decimal import Decimal
from decimal import *
import decimal
#import xlwings as xw
#from xlwings.constants import DeleteShiftDirection
from xlrd import open_workbook,XL_CELL_TEXT
import datetime
from datetime import date
import datetime as dt
from tkinter import filedialog
from tkinter import *
import os
import re
import json
from urllib.request import urlopen
import pip
import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from shapely.geometry import Point
from shapely.geometry.polygon import Polygon
import datetime
import datetime as dt

#--------SELECTION DU FICHIER AVEC LES NOUVELLES ANNONCES--------
path_RESULT = Tk()
Label1 = Label(path_RESULT, text = "Sélectionner le fichier dont les nuitées doivent être calculées", fg = 'red')
Label1.pack()
path_RESULT.filename =  filedialog.askopenfilename(initialdir = "/",title = "Sélectionner le fichier dont les nuitées doivent être calculées",filetypes = (("Excel file","*.xlsx"),("all files","*.*")))
print (path_RESULT.filename)
NAMEFile=os.path.splitext(os.path.basename(path_RESULT.filename))[0]
print(NAMEFile)
DIR=os.path.dirname(path_RESULT.filename)
DIR2=DIR+'/'
print(os.path.dirname(path_RESULT.filename))

#now = str(datetime.date.now())[:19]
#now = now.replace(":","_")
now = str(date.today())
print(now)

Deccontext = Context(prec=10, rounding=ROUND_HALF_DOWN)
setcontext(Deccontext)

chrome_options = webdriver.ChromeOptions()
#prefs = {"profile.managed_default_content_settings.images": 2}
#chrome_options.add_experimental_option("prefs", prefs)

#on récupère la value dans le excel ORIGINE
#32460

workbook = xlwt.Workbook()
sheet_write = workbook.add_sheet('URL_PAGE')
sheet_write.write(0, 0, 'A')
sheet_write.write(0, 1, 'B')
sheet_write.write(0, 2, 'C')
sheet_write.write(0, 3, 'D')
sheet_write.write(0, 4, 'E')
sheet_write.write(0, 5, 'F')
sheet_write.write(1, 5, 'b')
sheet_write.write(0, 6, 'G')

fromINSEE=1
if fromINSEE==1:
	wb2 = xlrd.open_workbook('INSEE_FRANCE_XML.xls')
	sheet1 = wb2.sheet_by_index(0)
	v=0
	i=1
	ville='78158'   #LECHESNAY
	#ville='78646'  #VERSAILLES
	#ville='75102'
	#ville="64483"
	#ville="29046"
	#ISSY
	#ville=92040
	#st-malo
	#ville=35288
	#SABLE-OLONNE
	#ville=85194
	tv=str(ville)
	print(tv)
	while v==0:
		naming=sheet1.cell_value(i,0)
		try:
			tt=int(naming)
			t=str(tt)
		except:
			t=str(naming)
		if t==tv:
			v=1
			GPS_ORIGIN=sheet1.cell_value(i,6)
		i=i+1

	#on supprime les space
	GPS_sans_space=GPS_ORIGIN.replace(" ","")

	# split sur "["
	split_gps=GPS_sans_space.split("[")

	#on supprime le 1er element de la liste car il est vide
	split_gps.remove(split_gps[0])
	split_gps.remove(split_gps[0])
	split_gps.remove(split_gps[0])
	#longeur de la liste
	T_gps=len(split_gps)
	print (T_gps)

	#ligne excel initialisé à 0
	c=0
	l_max=[]
	l_min=[]
	for h in split_gps:
		crochet=h.split("]")
		double_gps=crochet[0].split(",")
		c=c+1
		l_max.append(double_gps[-1])
		l_min.append(double_gps[0])
	print(max(l_max)) #ne_lat3 N1
	print(min(l_min)) #sw_lat1 S3
	print(min(l_max)) #ne_lng2 w2
	print(max(l_min)) #sw_lng4 e4
elif fromINSEE==2:
	#Poly=[(2.922176, 48.928696), (2.626191, 48.935857), (2.628215, 48.783178), (2.953307, 48.785616), (2.922176, 48.928696)]
	#Poly=[(-1.7022587,46.556284), (-1.8665397,46.45638)]
	#Poly=[(-0.566461976479953, 44.54605534503501), (-0.601676214501757, 44.55285985236557), (-0.615502791796292, 44.58296692120781), (-0.609008046685042, 44.60364546556298), (-0.619038811778664, 44.61095042862514), (-0.6619057650858861, 44.6087626175587), (-0.694386280213617, 44.64370391840443), (-0.729966737111434, 44.67257549413891), (-0.721742144539515, 44.68114439646803), (-0.65662732785245, 44.68355975040168), (-0.66042725398573, 44.68993207437811), (-0.672078913543385, 44.69481824801064), (-0.64955519919815, 44.72100596663734), (-0.638113761690622, 44.7272407973585), (-0.635703284981555, 44.73024005567959), (-0.645128260959859, 44.73169655392201), (-0.638302988360768, 44.73661663347337), (-0.633985326942008, 44.73593830796983), (-0.632718880232667, 44.74217148515195), (-0.616642192434454, 44.74918679010368), (-0.607622410104385, 44.75723173048317), (-0.590136906201729, 44.76270268236247), (-0.585187455624579, 44.75312966997373), (-0.571356501419781, 44.74405964965744), (-0.562255748275608, 44.74192854372193), (-0.5658157884666259, 44.75200675728984), (-0.540660545025235, 44.76185305655517), (-0.53462259120344, 44.76721886057388), (-0.510893830266989, 44.76909999690599), (-0.5111915025439741, 44.76495405568519), (-0.511578560950593, 44.74583642722698), (-0.501591689732834, 44.73999758637085), (-0.483650238920068, 44.73612976395617), (-0.458048243215631, 44.72667717584467), (-0.448416777206444, 44.72194919998991), (-0.438030215293798, 44.70545692757102), (-0.436840151970893, 44.70417822912138), (-0.433892184161965, 44.70266572453448), (-0.436871442073003, 44.68875876065009), (-0.457059178510567, 44.66861994643612), (-0.441642519281372, 44.65480890650544), (-0.459694058384267, 44.63011315020232), (-0.488561495415893, 44.6269617627626), (-0.488719701862386, 44.61728753564925), (-0.488809427352535, 44.61241855986734), (-0.46708540617104, 44.60544291991868), (-0.481051634606019, 44.59159488660275), (-0.48790043592681, 44.57530270334252), (-0.511297994605287, 44.57001029560205), (-0.54576155958516, 44.56977378565707), (-0.550574420737808, 44.55971544008108), (-0.566461976479953, 44.54605534503501)]
	#perimeter=[-0.566461976479953,44.54605534503501;-0.601676214501757,44.55285985236557;-0.615502791796292,44.58296692120781;-0.609008046685042,44.60364546556298;-0.619038811778664,44.61095042862514;-0.6619057650858861,44.6087626175587;-0.694386280213617,44.64370391840443;-0.729966737111434,44.67257549413891;-0.721742144539515,44.68114439646803;-0.65662732785245,44.68355975040168;-0.66042725398573,44.68993207437811;-0.672078913543385,44.69481824801064;-0.64955519919815,44.72100596663734;-0.638113761690622,44.7272407973585;-0.635703284981555,44.73024005567959;-0.645128260959859,44.73169655392201;-0.638302988360768,44.73661663347337;-0.633985326942008,44.73593830796983;-0.632718880232667,44.74217148515195;-0.616642192434454,44.74918679010368;-0.607622410104385,44.75723173048317;-0.590136906201729,44.76270268236247;-0.585187455624579,44.75312966997373;-0.571356501419781,44.74405964965744;-0.562255748275608,44.74192854372193;-0.5658157884666259,44.75200675728984;-0.540660545025235,44.76185305655517;-0.53462259120344,44.76721886057388;-0.510893830266989,44.76909999690599;-0.5111915025439741,44.76495405568519;-0.511578560950593,44.74583642722698;-0.501591689732834,44.73999758637085;-0.483650238920068,44.73612976395617;-0.458048243215631,44.72667717584467;-0.448416777206444,44.72194919998991;-0.438030215293798,44.70545692757102;-0.436840151970893,44.70417822912138;-0.433892184161965,44.70266572453448;-0.436871442073003,44.68875876065009;-0.457059178510567,44.66861994643612;-0.441642519281372,44.65480890650544;-0.459694058384267,44.63011315020232;-0.488561495415893,44.6269617627626;-0.488719701862386,44.61728753564925;-0.488809427352535,44.61241855986734;-0.46708540617104,44.60544291991868;-0.481051634606019,44.59159488660275;-0.48790043592681,44.57530270334252;-0.511297994605287,44.57001029560205;-0.54576155958516,44.56977378565707;-0.550574420737808,44.55971544008108;-0.566461976479953,44.54605534503501]
	#perimeter=[-0.566461976479953,44.54605534503501; -0.601676214501757,44.55285985236557; -0.615502791796292,44.58296692120781; -0.609008046685042,44.60364546556298; -0.619038811778664,44.61095042862514; -0.6619057650858861,44.6087626175587; -0.694386280213617,44.64370391840443; -0.729966737111434,44.67257549413891; -0.721742144539515,44.68114439646803; -0.65662732785245,44.68355975040168; -0.66042725398573,44.68993207437811; -0.672078913543385,44.69481824801064; -0.64955519919815,44.72100596663734; -0.638113761690622,44.7272407973585; -0.635703284981555,44.73024005567959; -0.645128260959859,44.73169655392201; -0.638302988360768,44.73661663347337; -0.633985326942008,44.73593830796983; -0.632718880232667,44.74217148515195; -0.616642192434454,44.74918679010368; -0.607622410104385,44.75723173048317; -0.590136906201729,44.76270268236247; -0.585187455624579,44.75312966997373; -0.571356501419781,44.74405964965744; -0.562255748275608,44.74192854372193; -0.5658157884666259,44.75200675728984; -0.540660545025235,44.76185305655517; -0.53462259120344,44.76721886057388; -0.510893830266989,44.76909999690599; -0.5111915025439741,44.76495405568519; -0.511578560950593,44.74583642722698; -0.501591689732834,44.73999758637085; -0.483650238920068,44.73612976395617; -0.458048243215631,44.72667717584467; -0.448416777206444,44.72194919998991; -0.438030215293798,44.70545692757102; -0.436840151970893,44.70417822912138; -0.433892184161965,44.70266572453448; -0.436871442073003,44.68875876065009; -0.457059178510567,44.66861994643612; -0.441642519281372,44.65480890650544; -0.459694058384267,44.63011315020232; -0.488561495415893,44.6269617627626; -0.488719701862386,44.61728753564925; -0.488809427352535,44.61241855986734; -0.46708540617104,44.60544291991868; -0.481051634606019,44.59159488660275; -0.48790043592681,44.57530270334252; -0.511297994605287,44.57001029560205; -0.54576155958516,44.56977378565707; -0.550574420737808,44.55971544008108; -0.566461976479953,44.54605534503501]
	#perimeter=[(-0.566461976479953, 44.54605534503501), (-0.601676214501757, 44.55285985236557), (-0.615502791796292, 44.58296692120781), (-0.609008046685042, 44.60364546556298), (-0.619038811778664, 44.61095042862514), (-0.6619057650858861, 44.6087626175587), (-0.694386280213617, 44.64370391840443), (-0.729966737111434, 44.67257549413891), (-0.721742144539515, 44.68114439646803), (-0.65662732785245, 44.68355975040168), (-0.66042725398573, 44.68993207437811), (-0.672078913543385, 44.69481824801064), (-0.64955519919815, 44.72100596663734), (-0.638113761690622, 44.7272407973585), (-0.635703284981555, 44.73024005567959), (-0.645128260959859, 44.73169655392201), (-0.638302988360768, 44.73661663347337), (-0.633985326942008, 44.73593830796983), (-0.632718880232667, 44.74217148515195), (-0.616642192434454, 44.74918679010368), (-0.607622410104385, 44.75723173048317), (-0.590136906201729, 44.76270268236247), (-0.585187455624579, 44.75312966997373), (-0.571356501419781, 44.74405964965744), (-0.562255748275608, 44.74192854372193), (-0.5658157884666259, 44.75200675728984), (-0.540660545025235, 44.76185305655517), (-0.53462259120344, 44.76721886057388), (-0.510893830266989, 44.76909999690599), (-0.5111915025439741, 44.76495405568519), (-0.511578560950593, 44.74583642722698), (-0.501591689732834, 44.73999758637085), (-0.483650238920068, 44.73612976395617), (-0.458048243215631, 44.72667717584467), (-0.448416777206444, 44.72194919998991), (-0.438030215293798, 44.70545692757102), (-0.436840151970893, 44.70417822912138), (-0.433892184161965, 44.70266572453448), (-0.436871442073003, 44.68875876065009), (-0.457059178510567, 44.66861994643612), (-0.441642519281372, 44.65480890650544), (-0.459694058384267, 44.63011315020232), (-0.488561495415893, 44.6269617627626), (-0.488719701862386, 44.61728753564925), (-0.488809427352535, 44.61241855986734), (-0.46708540617104, 44.60544291991868), (-0.481051634606019, 44.59159488660275), (-0.48790043592681, 44.57530270334252), (-0.511297994605287, 44.57001029560205), (-0.54576155958516, 44.56977378565707), (-0.550574420737808, 44.55971544008108), (-0.566461976479953, 44.54605534503501)]
	#GPS_ORIGIN='{"type": "Polygon", "coordinates": [[[-0.566461976479953, 44.54605534503501], [-0.601676214501757, 44.55285985236557], [-0.615502791796292, 44.58296692120781], [-0.609008046685042, 44.60364546556298], [-0.619038811778664, 44.61095042862514], [-0.6619057650858861, 44.6087626175587], [-0.694386280213617, 44.64370391840443], [-0.729966737111434, 44.67257549413891], [-0.721742144539515, 44.68114439646803], [-0.65662732785245, 44.68355975040168], [-0.66042725398573, 44.68993207437811], [-0.672078913543385, 44.69481824801064], [-0.64955519919815, 44.72100596663734], [-0.638113761690622, 44.7272407973585], [-0.635703284981555, 44.73024005567959], [-0.645128260959859, 44.73169655392201], [-0.638302988360768, 44.73661663347337], [-0.633985326942008, 44.73593830796983], [-0.632718880232667, 44.74217148515195], [-0.616642192434454, 44.74918679010368], [-0.607622410104385, 44.75723173048317], [-0.590136906201729, 44.76270268236247], [-0.585187455624579, 44.75312966997373], [-0.571356501419781, 44.74405964965744], [-0.562255748275608, 44.74192854372193], [-0.5658157884666259, 44.75200675728984], [-0.540660545025235, 44.76185305655517], [-0.53462259120344, 44.76721886057388], [-0.510893830266989, 44.76909999690599], [-0.5111915025439741, 44.76495405568519], [-0.511578560950593, 44.74583642722698], [-0.501591689732834, 44.73999758637085], [-0.483650238920068, 44.73612976395617], [-0.458048243215631, 44.72667717584467], [-0.448416777206444, 44.72194919998991], [-0.438030215293798, 44.70545692757102], [-0.436840151970893, 44.70417822912138], [-0.433892184161965, 44.70266572453448], [-0.436871442073003, 44.68875876065009], [-0.457059178510567, 44.66861994643612], [-0.441642519281372, 44.65480890650544], [-0.459694058384267, 44.63011315020232], [-0.488561495415893, 44.6269617627626], [-0.488719701862386, 44.61728753564925], [-0.488809427352535, 44.61241855986734], [-0.46708540617104, 44.60544291991868], [-0.481051634606019, 44.59159488660275], [-0.48790043592681, 44.57530270334252], [-0.511297994605287, 44.57001029560205], [-0.54576155958516, 44.56977378565707], [-0.550574420737808, 44.55971544008108], [-0.566461976479953, 44.54605534503501]]]}'
	#GPS_ORIGIN='{"type": "Polygon", "coordinates": [[[2.922176, 48.928696], [2.626191, 48.935857], [2.628215, 48.783178], [2.953307, 48.785616], [2.922176, 48.928696]]]}'
	#olonne
	#GPS_ORIGIN='{"type": "Polygon", "coordinates": [[[-1.7022587,46.556284], [-1.7022587,46.45638], [-1.8665397,46.45638], [-1.8665397,46.556284]]]}'
	#angers
	#GPS_ORIGIN='{"type": "Polygon", "coordinates": [[[-0.3343564,47.5825096], [-0.3343564,47.267604], [-0.9328407,47.267604], [-0.9328407,47.5825096]]]}'
	GPS_ORIGIN='{"type": "Polygon", "coordinates": [[[5.9192223,50.240395], [5.9192223,48.917916], [3.583670,48.917916], [3.583670,50.240395]]]}'
	
	#on supprime les space
	GPS_sans_space=GPS_ORIGIN.replace(" ","")

	# split sur "["
	split_gps=GPS_sans_space.split("[")

	#on supprime le 1er element de la liste car il est vide
	split_gps.remove(split_gps[0])
	split_gps.remove(split_gps[0])
	split_gps.remove(split_gps[0])
	#longeur de la liste
	T_gps=len(split_gps)
	print (T_gps)

	#ligne excel initialisé à 0
	c=0
	l_max=[]
	l_min=[]
	for h in split_gps:
		crochet=h.split("]")
		double_gps=crochet[0].split(",")
		c=c+1
		l_max.append(double_gps[-1])
		l_min.append(double_gps[0])
	print(max(l_max)) #ne_lat N
	print(max(l_min)) #ne_lng E
	print(min(l_max)) #sw_lat S
	print(min(l_min)) #sw_lng W
else:
	l_max=[48.108438, 46.170246]
	l_min=[2.1249163, -0.8501037]
	print(max(l_max)) #ne_lat N
	print(max(l_min)) #ne_lng E
	print(min(l_max)) #sw_lat S
	print(min(l_min)) #sw_lng W

#l_max=[48.863344374598334, 48.869863809746434, 48.870630685567946]
#l_min=[2.350834505477619, 2.327877416924118, 2.347826239446091]
#https://www.airbnb.fr/s/dinard/homes?refinement_paths%5B%5D=%2Fhomes&query=dinard&search_type=unknown&ne_lat=48.6435799942263&ne_lng=-2.0336458067321246&sw_lat=48.623414896166715&sw_lng=-2.074629960479683&zoom=15&search_by_map=true&place_id=ChIJVVqqTcCADkgR48BvH6Pvhd8&s_tag=Y6jz9TTh
#https://www.airbnb.fr/s/blois/homes?refinement_paths%5B%5D=%2Fhomes&current_tab_id=home_tab&selected_tab_id=home_tab&source=mc_search_bar&click_referer=t%3ASEE_ALL%7Csid%3Aba7e066c-7524-4d15-8ebc-24f988bad2e4%7Cst%3ALANDING_PAGE_MARQUEE&screen_size=large&zoom=14&search_by_map=true&sw_lat=47.56476024379833&sw_lng=1.2938918252564235&ne_lat=47.611939549842894&ne_lng=1.3847865243531032&search_type=unknown&adults=1
#https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat=47.54160887370493&sw_lng=1.355879995800566&ne_lat=47.620619985628096&ne_lng=1.254120619005437&zoom=14&search_by_map=true&map_toggle=true
#url_start='https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&current_tab_id=home_tab&selected_tab_id=home_tab&allow_override%5B%5D=&sw_lat='+str(min(l_max))+'&sw_lng='+str(min(l_min))+'&ne_lat='+str(max(l_max))+'&ne_lng='+str(max(l_min))
#url_start='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat='+str(min(l_max))+'&sw_lng='+str(min(l_min))+'&ne_lat='+str(max(l_max))+'&ne_lng='+str(max(l_min))+'&zoom=14&search_by_map=true&map_toggle=true'
url_start='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=unknown&ne_lat='+str(max(l_max))+'&ne_lng='+str(max(l_min))+'& sw_lat='+str(min(l_max))+'&sw_lng=-'+str(min(l_min))+'&zoom=17&search_by_map=true'

print(url_start)

typ=1

if typ==1:
	sheet_write.write(1, 0, url_start)
	sheet_write.write(1, 1, max(l_max))#3 N
	sheet_write.write(1, 2, max(l_min))#4 E
	sheet_write.write(1, 3, min(l_max))#1 S
	sheet_write.write(1, 4, min(l_min))#2 W
else:
	sheet_write.write(1, 0, url_start)
	sheet_write.write(1, 1, max(l_max))#3 N
	sheet_write.write(1, 2, min(l_min))#4 E
	sheet_write.write(1, 3, min(l_max))#1 S
	sheet_write.write(1, 4, max(l_min))#2 W
workbook.save('myFile'+str(now)+'.xls')

book = xlrd.open_workbook('myFile'+str(now)+'.xls')
copy_book =copy(book)
sheet_write=copy_book.get_sheet(0)
r=2
nrow=2
h=1

#driver = webdriver.Chrome(chrome_options=chrome_options)
driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
driver.set_window_size(1500, 2000)
wait = WebDriverWait(driver, 3)
time.sleep(5)
x_p="//div[@class='_1h559tl']"
while h < nrow:

	print (h)
	book_dup = xlrd.open_workbook('myFile'+str(now)+'.xls')
	sheet_read = book_dup.sheet_by_index(0)
	URL=sheet_read.cell(h,0).value
	print (URL)
	driver.get(URL)
	time.sleep(3)
	wait = WebDriverWait(driver, 3)
	locations='aucune annonce'
	try:
		#print ('ici')
		locations = wait.until(EC.presence_of_element_located((By.XPATH, x_p))).text
		#print (locations)
		n_loc=locations.split(" ")
		#print (n_loc)
		#nb_loc=n_loc[4].replace("+","")
		n_loc.reverse()
		nb_loc=n_loc[1]
		f_loc=float(nb_loc)
		print (nb_loc)
		sheet_write.write(h, 6, nb_loc)
		copy_book.save('myFile'+str(now)+'.xls')
	except:
		pass
	#print (locations)
	if locations=='aucune annonce':
		sheet_write.write(h, 5, 'O')
		sheet_write.write(h, 6, 0)
		copy_book.save('myFile'+str(now)+'.xls')
		f_loc=0
	if f_loc>=300:
	#MARQUAGE X SUR URL 300+
		sheet_write.write(h, 5, "X")
	#FORMULE POUR DIVISION 4 CELLS
		n_t=sheet_read.cell(h,1).value
		e_t=sheet_read.cell(h,2).value
		s_t=sheet_read.cell(h,3).value
		w_t=sheet_read.cell(h,4).value
		#print (n_t)
		#print (e_t)
		#print (s_t)
		#print (w_t)
		div_lat=(Decimal(n_t)-Decimal(s_t))/2
		div_long=(Decimal(e_t)-Decimal(w_t))/2
		#print (div_lat)
		#print (div_long)
	#DIVISION 1
		n_g1=Decimal(n_t)-(div_lat)
		e_g1=Decimal(e_t)-(div_long)
		s_g1=Decimal(s_t)
		w_g1=Decimal(w_t)
		n_t1=str(n_g1)
		e_t1=str(e_g1)
		s_t1=str(s_g1)
		w_t1=str(w_g1)
		sheet_write.write(r, 1, n_t1)
		sheet_write.write(r, 2, e_t1)
		sheet_write.write(r, 3, s_t1)
		sheet_write.write(r, 4, w_t1)
	#DIVISION 2
		r=r+1
		n_g2=Decimal(n_t)
		e_g2=Decimal(w_t)+(div_long)
		s_g2=Decimal(s_t)+(div_lat)
		w_g2=Decimal(w_t)
		n_t2=str(n_g2)
		e_t2=str(e_g2)
		s_t2=str(s_g2)
		w_t2=str(w_g2)
		sheet_write.write(r, 1, n_t2)
		sheet_write.write(r, 2, e_t2)
		sheet_write.write(r, 3, s_t2)
		sheet_write.write(r, 4, w_t2)
	#DIVISION 3
		r=r+1
		n_g3=Decimal(n_t)-(div_lat)
		e_g3=Decimal(e_t)
		s_g3=Decimal(s_t)
		w_g3=Decimal(w_t)+(div_long)
		n_t3=str(n_g3)
		e_t3=str(e_g3)
		s_t3=str(s_g3)
		w_t3=str(w_g3)
		sheet_write.write(r, 1, n_t3)
		sheet_write.write(r, 2, e_t3)
		sheet_write.write(r, 3, s_t3)
		sheet_write.write(r, 4, w_t3)
	#DIVISION 4
		r=r+1
		n_g4=Decimal(n_t)
		e_g4=Decimal(e_t)
		s_g4=Decimal(s_t)+(div_lat)
		w_g4=Decimal(w_t)+(div_long)
		n_t4=str(n_g4)
		e_t4=str(e_g4)
		s_t4=str(s_g4)
		w_t4=str(w_g4)
		sheet_write.write(r, 1, n_t4)
		sheet_write.write(r, 2, e_t4)
		sheet_write.write(r, 3, s_t4)
		sheet_write.write(r, 4, w_t4)
	#CREATION URL DES DIVISIONS
		#url1='https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat='+str(s_g1)+'&sw_lng='+str(w_g1)+'&ne_lat='+str(n_g1)+'&ne_lng='+str(e_g1)+'&zoom=14&search_by_map=true&map_toggle=true'
		#url2='https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat='+str(s_g2)+'&sw_lng='+str(w_g2)+'&ne_lat='+str(n_g2)+'&ne_lng='+str(e_g2)+'&zoom=14&search_by_map=true&map_toggle=true'
		#url3='https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat='+str(s_g3)+'&sw_lng='+str(w_g3)+'&ne_lat='+str(n_g3)+'&ne_lng='+str(e_g3)+'&zoom=14&search_by_map=true&map_toggle=true'
		#url4='https://www.airbnb.fr/s/toto/homes?refinement_paths%5B%5D=%2Fhomes&allow_override%5B%5D=&sw_lat='+str(s_g4)+'&sw_lng='+str(w_g4)+'&ne_lat='+str(n_g4)+'&ne_lng='+str(e_g4)+'&zoom=14&search_by_map=true&map_toggle=true'
		url1='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=pagination&ne_lat='+str(n_g1)+'&ne_lng='+str(e_g1)+'&sw_lat='+str(s_g1)+'&sw_lng='+str(w_g1)+'&zoom=17&search_by_map=true'
		url2='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=pagination&ne_lat='+str(n_g2)+'&ne_lng='+str(e_g2)+'&sw_lat='+str(s_g2)+'&sw_lng='+str(w_g2)+'&zoom=17&search_by_map=true'
		url3='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=pagination&ne_lat='+str(n_g3)+'&ne_lng='+str(e_g3)+'&sw_lat='+str(s_g3)+'&sw_lng='+str(w_g3)+'&zoom=17&search_by_map=true'
		url4='https://www.airbnb.fr/s/Olonne~sur~Mer/homes?tab_id=home_tab&search_type=pagination&ne_lat='+str(n_g4)+'&ne_lng='+str(e_g4)+'&sw_lat='+str(s_g4)+'&sw_lng='+str(w_g4)+'&zoom=17&search_by_map=true'

		sheet_write.write((r-3), 0, url1)
		sheet_write.write((r-2), 0, url2)
		sheet_write.write((r-1), 0, url3)
		sheet_write.write(r, 0, url4)
	#SAVE EXCEL
		copy_book.save('myFile'+str(now)+'.xls')
		nrow=nrow+4
		r=r+1
	h=h+1

print ('_______    ___    ___     ___')
print ('|      |   |  |   |  \    |  |')
print ('|  |__     |  |   |   \   |  |')
print ('|     |    |  |   |    \  |  |')
print ('|  |       |  |   |  |\ \ |  |')
print ('|  |       |  |   |  | \ \|  |')
print ('|__|       |__|   |__|  \____|')
print ('le découpage des zones est terminé, vous pouvez exécuter le .exe N°2 qui va extraire la liste des URL de la totalité des annonces de votre zone.')

xl = pd.ExcelFile('myFile'+str(now)+'.xls')
df = xl.parse("URL_PAGE")
#SUPPRIME les lignes contenant X dans la colonne F
df2 = df.loc[df['F'] != 'X']
list_URL=df2['A'].tolist()
#print (list_URL)
#print (len(list_URL))
time.sleep(1)
#OUVERTURE DES PAGES CHROME
#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
wait = WebDriverWait(driver, 10)
wait2 = WebDriverWait(driver, 5)

#c = ligne 2 du xls
c=1

test='OK'
z=0
while test=='KO':
	driver.get(list_URL[0])
	time.sleep(2)
	try:
		link = wait.until(EC.presence_of_element_located((By.XPATH, '//span[@class="_qlq27g"]['+'1'+']/a')))
		url=link.get_attribute("href")
		test='OK'
	except:
		driver.quit()
		driver = webdriver.Chrome(chrome_options=chrome_options)
		wait = WebDriverWait(driver, 5)
		wait2 = WebDriverWait(driver, 5)
		z=z+1
	if z==10:
		test='OK'

dec=1
list_ann=[]


for h in list_URL:
	print ("DECOMPTE = "+str(dec)+"<<<<<<<<")
	dec=dec+1
	driver.get(h)
	time.sleep(3)
	html = driver.page_source
	soup = BeautifulSoup(html, 'html.parser')
	j=1
	time.sleep(2)
	print(len(list_ann))
	while (j<=17):
		i=0
		try:
			while (i<=20):
				try:
					try:
						#print(c)
					#URL ANNONCE
						#link = wait.until(EC.presence_of_element_located((By.XPATH, XPATH_ann1+str(i)+XPATH_ann2)))
						#link = wait.until(EC.presence_of_element_located((By.XPATH, "//span[@class='_qlq27g']["+str(i)+"]/a")))
						the_tr= soup.findAll('div', attrs={"class": "_8ssblpx"})[i]
						div=the_tr.find('a').attrs
						#link = rootdriver.find_element_by_xpath(XPATH_ann1+str(i)+XPATH_ann2)
						link=div['href']
						#url=link.get_attribute("href")
						v_url=link.split("?")
						#print (v_url[0])
						i=i+1
						
						#sheet_write2.write(c, 1, 'https://www.airbnb.fr'+str(v_url[0]))
						#sheet_write2.cell(row=c+1, column=2).value = 'https://www.airbnb.fr'+str(v_url[0])
						list_ann.append('https://www.airbnb.fr'+str(v_url[0]))
					except:
						#print('NO LINK')
						zzz=1
						#print(str(i)+"--------------"+str(c))
						break
					c=c+1
					#wb.save('RESULT.xls')
					
				except:
					print("dernière annonce")
					
					break
			#print(list_ann)
			nextpage='0'
			#nextpage = wait2.until(EC.presence_of_element_located((By.XPATH, "//li[@class='_r4n1gzb']/a")))
			
			nextpage = wait2.until(EC.presence_of_element_located((By.XPATH, "//a[@aria-label='Suivant']")))

			#nextpage=rootdriver.find_element_by_xpath(XPATH_Next)
			a_nextpage=nextpage.get_attribute("href")
			label=nextpage.get_attribute("aria-label")
			#print (label)
			#if label is None:
			if label=='Suivant':
			#le bouton NextPage n'a pas d'attribut Label, quand il n'y a plus de bouton nextpage, l'attribut Label est présent dans //li[@class='_b8vexar']/a
				driver.get(a_nextpage)
				time.sleep(1)
				html = driver.page_source
				soup = BeautifulSoup(html, 'html.parser')
				time.sleep(1)
				j=j+1
			else:
			#Si label pas NULL, arreter la boucle While j
				j=20
				print ("---------URL suivant-----------")
		except:
			print("FIN")
			#rootdriver.close()
			#driver.close()
			break
	else:
		continue
#print(list_ann)
#workbook2.save('RESULT_AIRBNB.xls')
print ('_______    ___    ___     ___')
print ('|      |   |  |   |  \    |  |')
print ('|  |__     |  |   |   \   |  |')
print ('|     |    |  |   |    \  |  |')
print ('|  |       |  |   |  |\ \ |  |')
print ('|  |       |  |   |  | \ \|  |')
print ('|__|       |__|   |__|  \____|')
print('Exécuter à présent le exe N°3 pour obtenir le détail des annonces ici extraites')


wb = load_workbook(path_RESULT.filename)
ws=wb.active
nrow=ws.max_row
#--------------------------------------------------------------------------------------------------------------
#-------------------CREATE LIST URL
#-------------------

try:
	ws2 = wb.get_sheet_by_name('LIST_BACKUP')
	nrow2=ws2.max_row
	c=nrow2+1
	list_BACKUP=[]
	for col in ws2['A']:
		list_BACKUP.append(col.value)
	for ann in list_ann:
		if not ann in list_BACKUP:
			ws2.cell(row=c, column=1).value = ann
			ws2.cell(row=c, column=2).value = now
			c=c+1
except:
	ws2=wb.create_sheet('LIST_BACKUP')
	i=1
	while i<len(list_ann):
		ws2.cell(row=i, column=1).value = list_ann[i]
		ws2.cell(row=i, column=2).value = now
		i=i+1
		



#list_URL=sheet_read.col_values(1)
list_URL=[]
for col in ws['B']:
    list_URL.append(col.value)



del list_URL[0]
c_list=[elem for elem in list_ann if elem not in list_URL ]
print(len(list_URL))
print(len(list_ann))
print(len(c_list))
print("start extraction details")
c=nrow+1
for g in c_list:
	ws.cell(row=c, column=2).value = g
	c=c+1
wb.save(path_RESULT.filename)
nrow2=ws.max_row
#-------------------------------------------------------------------------------------------------------------
#-------------------POLYGONE
#-------------------


#EXAMPLE OF POLY
#poly27=Polygon([(3.434049833168874, 43.82205305876419), (3.425479726783461, 43.814518621478754), (3.38892811812125, 43.81604096360177), (3.377387088079953, 43.824252644937545), (3.369846342135982, 43.851779842951125), (3.387068364134024, 43.860755354858405), (3.395221036634612, 43.86162606223585), (3.421593517026773, 43.87231756200473), (3.43518153631202, 43.862958915752216), (3.429726395191789, 43.8537292666638), (3.436344397967047, 43.84096009534653), (3.440458969865034, 43.839233714400095), (3.437984929930104, 43.82951558615399), (3.432345765972657, 43.825173513947725), (3.434049833168874, 43.82205305876419), ])

#PARIS 2eme
#polygon=Polygon([(2.350834505477619, 48.863344374598334), (2.327877416924118, 48.869863809746434), (2.347826239446091, 48.870630685567946), (2.354114163249875, 48.86927979700135), (2.350834505477619, 48.863344374598334)])

#VERSAILLES
polygon=Polygon([(2.146395117801802, 48.7908464891847), (2.142014643567137, 48.79233514912969), (2.125014463768882, 48.78083407054695), (2.105472174907891, 48.78074332208332), (2.070251387007021, 48.78981062262139), (2.083161019829898, 48.818952748452425), (2.089120120222795, 48.826337463396264), (2.091738802551524, 48.8238030685563), (2.100586756530543, 48.824139766534394), (2.112676124617275, 48.82851123738547), (2.113987690376982, 48.82187437468609), (2.131007668278255, 48.81486698444451), (2.136072645252615, 48.814939667169654), (2.140437002227598, 48.82446882876733), (2.148475927847977, 48.82849268523683), (2.151465535500457, 48.821408687854955), (2.150744478965478, 48.81884926644504), (2.161274834107939, 48.81278066375281), (2.160154468443745, 48.800324937247375), (2.168168048617256, 48.787870032204715), (2.161929188683613, 48.78223199211193), (2.146395117801802, 48.7908464891847)])

#MARTILLAC
#polygon=Polygon([(-0.566461976479953, 44.54605534503501), (-0.601676214501757, 44.55285985236557), (-0.615502791796292, 44.58296692120781), (-0.609008046685042, 44.60364546556298), (-0.619038811778664, 44.61095042862514), (-0.6619057650858861, 44.6087626175587), (-0.694386280213617, 44.64370391840443), (-0.729966737111434, 44.67257549413891), (-0.721742144539515, 44.68114439646803), (-0.65662732785245, 44.68355975040168), (-0.66042725398573, 44.68993207437811), (-0.672078913543385, 44.69481824801064), (-0.64955519919815, 44.72100596663734), (-0.638113761690622, 44.7272407973585), (-0.635703284981555, 44.73024005567959), (-0.645128260959859, 44.73169655392201), (-0.638302988360768, 44.73661663347337), (-0.633985326942008, 44.73593830796983), (-0.632718880232667, 44.74217148515195), (-0.616642192434454, 44.74918679010368), (-0.607622410104385, 44.75723173048317), (-0.590136906201729, 44.76270268236247), (-0.585187455624579, 44.75312966997373), (-0.571356501419781, 44.74405964965744), (-0.562255748275608, 44.74192854372193), (-0.5658157884666259, 44.75200675728984), (-0.540660545025235, 44.76185305655517), (-0.53462259120344, 44.76721886057388), (-0.510893830266989, 44.76909999690599), (-0.5111915025439741, 44.76495405568519), (-0.511578560950593, 44.74583642722698), (-0.501591689732834, 44.73999758637085), (-0.483650238920068, 44.73612976395617), (-0.458048243215631, 44.72667717584467), (-0.448416777206444, 44.72194919998991), (-0.438030215293798, 44.70545692757102), (-0.436840151970893, 44.70417822912138), (-0.433892184161965, 44.70266572453448), (-0.436871442073003, 44.68875876065009), (-0.457059178510567, 44.66861994643612), (-0.441642519281372, 44.65480890650544), (-0.459694058384267, 44.63011315020232), (-0.488561495415893, 44.6269617627626), (-0.488719701862386, 44.61728753564925), (-0.488809427352535, 44.61241855986734), (-0.46708540617104, 44.60544291991868), (-0.481051634606019, 44.59159488660275), (-0.48790043592681, 44.57530270334252), (-0.511297994605287, 44.57001029560205), (-0.54576155958516, 44.56977378565707), (-0.550574420737808, 44.55971544008108), (-0.566461976479953, 44.54605534503501)])

#ST-MALO
#polygon=Polygon([(-1.996983832359921, 48.60369217345382), (-2.012033542725769, 48.598074355232804), (-2.011566084333277, 48.605250644339314), (-2.001868800020467, 48.614554990395284), (-2.013709752148175, 48.61343447211897), (-2.017582832284507, 48.61994093687032), (-2.013265006240889, 48.626233947865124), (-2.021376665440134, 48.633255585484356), (-2.030868930655294, 48.63382799310378), (-2.030794076180643, 48.63841888342676), (-2.020660427530679, 48.63995553331221), (-2.02899195392072, 48.64696742767445), (-2.027833399467906, 48.651324640428676), (-2.015518506481623, 48.65198846427292), (-1.993153298280507, 48.66036440844067), (-1.981560436471955, 48.67404459470891), (-1.989582189435102, 48.6833702132521), (-1.975695892619096, 48.683049948845124), (-1.963803038350252, 48.687088096279716), (-1.961394217722459, 48.68078969663222), (-1.962340932240303, 48.67070140434514), (-1.9525954220265, 48.668376424159334), (-1.937680867701459, 48.658857029799734), (-1.952302376473395, 48.64591147140626), (-1.948667547385805, 48.632949352866156), (-1.939868211308484, 48.62170677422457), (-1.948157267855236, 48.6060134728909), (-1.97735980690807, 48.613546319108785), (-1.987714093015666, 48.612229231260905), (-1.996983832359921, 48.60369217345382)])

#ISSY-LES-MOULINEAUX
#polygon=Polygon([(2.272566370808877, 48.81440000991327), (2.251865120621872, 48.81638584250767), (2.236100452595979, 48.821547611630635), (2.249565274017956, 48.82478073110864), (2.262784637049641, 48.833928727333785), (2.267617156346577, 48.83420124774352), (2.271928340063522, 48.82888525520508), (2.280857987861738, 48.83133163356634), (2.289399238300263, 48.82835177916151), (2.281272073214195, 48.82501147379042), (2.279395187834523, 48.818893346034436), (2.272566370808877, 48.81440000991327)])
#Val-d'Europe
#polygon=Polygon([(48.928696, 2.922176), (48.935857, 2.626191), (48.783178, 2.628215), (48.785616, 2.953307), (48.928696, 2.922176)])
#LES-SABLE-D-OLONNE
#polygon=Polygon([(-1.761022476910707, 46.48064000858759), (-1.769978107450098, 46.48896240693245), (-1.782750103494504, 46.49430085250431), (-1.792481048970734, 46.49340573992643), (-1.799923114351193, 46.489069720101845), (-1.812565543107437, 46.49438334113363), (-1.818474294112867, 46.518058603953484), (-1.804265332550544, 46.51845854663043), (-1.794115710279393, 46.506931951583034), (-1.78975972377008, 46.505238314171), (-1.780236484299396, 46.50733322626675), (-1.773210765274535, 46.50290004984825), (-1.756355404739619, 46.486265111677554), (-1.761022476910707, 46.48064000858759)])
#polygon=Polygon([(-1.7022587,46.556284), (-1.7022587,46.45638), (-1.8665397,46.45638), (-1.8665397,46.556284)])
#ANGERS
#polygon=Polygon([(-0.3343564,47.5825096), (-0.3343564,47.267604), (-0.9328407,47.267604), (-0.9328407,47.5825096)])
#polygon=Polygon([(5.9192223,50.240395), (5.9192223,48.917916), (3.583670,48.917916), (3.583670,50.240395)])
#ZJ
#polygon=Polygon([(2.1249163,48.108438), (2.1249163,46.170246), (-0.8501037,46.170246), (-0.8501037,48.108438)])
#-------------------------------------------------------------------------------------------------------------
#-------------------SCRAP DETAILS
#-------------------

#c = ligne 2 du xls resultant
c=nrow+1
wait2 = WebDriverWait(driver, 2)
wait3 = WebDriverWait(driver, 3)	


end=0
while end==0:
	try:
		while c<=nrow2:
			print (str(c-nrow)+'/'+str(c))
			h=ws.cell(row=c, column=2).value
			print (h)
			if (c/2000).is_integer():
				driver.quit()
				time.sleep(5)
				driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver')
				#driver = webdriver.Chrome()
				driver.set_window_size(1500, 2000)
				wait2 = WebDriverWait(driver, 2)
				wait3 = WebDriverWait(driver, 3)
				time.sleep(5)
				driver.get(h)
				time.sleep(5)
			#do=sheet_read.cell(i,0).value
			do=True
			if do is True:
				driver.get(h)
				time.sleep(5)
				f_ele=5
				while f_ele<=3:
					try:
						#ele=driver.find_element_by_xpath("//div[@class='_1cvivhm']")
						ele=driver.find_element_by_xpath("//div[@class='_cg8a3u']")
						driver.execute_script("arguments[0].scrollIntoView(true);", ele)
						driver.execute_script("window.scrollBy(0,-800);")
						#driver.execute_script("window.scrollBy(0,500);")
						f_ele=6
						time.sleep(3)
					except:
						driver.execute_script("window.scrollBy(0,1000);")
						f_ele=f_ele+1
						time.sleep(1)
				driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
				time.sleep(2)
				driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
				time.sleep(1)
			#PROFILE
				html = driver.page_source
				time.sleep(1)
				soup = BeautifulSoup(html, 'html.parser')
				time.sleep(1)
				try:
					#GPS
					try:
						tp_c=soup.find('div', attrs={"class": "gm-style"})
						tt=tp_c.find('div', attrs={"style": "margin-left: 5px; margin-right: 5px; z-index: 1000000; position: absolute; left: 0px; bottom: 0px;"})
						tt1=tt.find('a', attrs={"target": "_blank"})
						tt2=tt1['href']
						#----------Create translation table----------
						table = str.maketrans('=&', '++')
						result_gps = tt2.translate(table)
						split_gps=result_gps.split("+")
						#https://www.google.com/maps?ll+46.23657,3.92341+z=14&t=m&hl=fr&gl=FR&mapclient=apiv3
						#https://maps.google.com/maps?ll=49.19054,-2.11715&z=14&t=m&hl=fr&gl=FR&mapclient=apiv3
						coor=split_gps[1]
						long_lat=coor.split(',')
						#--------------Write results--------------
						print(long_lat[0])
						print(long_lat[1])
						point=Point(float(long_lat[1]),float(long_lat[0]))
						question=polygon.contains(point)
						print(question)
						#sheet.write(c, 12, long_lat[0])
						#sheet.write(c, 13, long_lat[1])
						#ws.cell(row=c+1, column=13+1).value = long_lat[0]
						#ws.cell(row=c+1, column=14+1).value = long_lat[1]
					except:
						question=True
						print('NO GPS')
						ws.cell(row=c, column=70).value = 'NOGPS'
					if question is True:
						#sheet.write(c, 12, long_lat[0])
						#sheet.write(c, 13, long_lat[1])
						try:
							ws.cell(row=c, column=2).value = h
							ws.cell(row=c, column=14).value = long_lat[0]
							ws.cell(row=c, column=15).value = long_lat[1]
						except:
							ee=1
					#TITLE
						try:
							div1=soup.find('div', attrs={"class": "_mbmcsn"})
							ws.cell(row=c, column=1).value = div1.h1.text
						except:
							zzzz=1
							#print('NO TITLE')
					#URL HOTE
						try:
							div=soup.findAll('a', attrs={"class": "_105023be"})[-1]
							div1=div['href']  #.attrs['href']
							ws.cell(row=c, column=5).value = "https://www.airbnb.fr"+str(div1)
						except:
							zzzz=1
							#print('NO PROFILE')
					#COMMENTAIRE
						COMMENT='NO COMMENT'
						#run_price=extract("//span[@class='_wfad8t']",6,COMMENT,c,YN_comment)
						try:
							p_c=[]
							tp_c=soup.findAll('span', attrs={"class": "_bq6krt"})[1].text
							p_c=tp_c.replace("(","")
							cc=p_c.replace(")","")
							try:
								pp=cc.split(' ')
								cc=pp[0]
								ws.cell(row=c, column=7).value = cc
							except:
								pass
							ws.cell(row=c, column=7).value = cc
							#print ("COMMENT ===")
							print(cc)
							#p_c=tp_c.split("(")
							#print('ici1')
							#table_c = p_c[1].replace(")"," ")
							#print (table_c)
							#ws.cell(row=c+1, column=6+1).value = table_c
						except:
							zzzz=1
							#print('NOCOMMENT')
					#VOYAGEUR
						try:
							the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
							tt=the_tr.find_all('div')[1]
							tt1=tt.find_all('span')[0]
							tt2=tt1.text
							p_tp=tt2.split(" ")
							ws.cell(row=c, column=9).value = p_tp[0]
						except:
							zzzz=1
							#print('NO VOYAGER')

					#LITS
						try:
							the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
							tt=the_tr.find_all('div')[1]
							tt1=tt.find_all('span')[4]
							tt2=tt1.text
							p_tp=tt2.split(" ")
							ws.cell(row=c, column=12).value = p_tp[0]
						except:
							zzzz=1
							#print('NO LIT')
					#SdB
						try:
							the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
							tt=the_tr.find_all('div')[1]
							tt1=tt.find_all('span')[6]
							tt2=tt1.text
							p_tp=tt2.split(" ")
							ws.cell(row=c, column=11).value = p_tp[0]
						except:
							zzzz=1
							#print('NO SdB')
					#CHAMBRE
						try:
							the_tr= soup.find('div', attrs = {'class' : '_tqmy57'})
							tt=the_tr.find_all('div')[1]
							tt1=tt.find_all('span')[2]
							tt2=tt1.text
							p_tp=tt2.split(" ")
							ws.cell(row=c, column=10).value = p_tp[0]
						except:
							zzzz=1
							#print('NO CHAMBRE')
					#VILLE
						try:
							tp_c=soup.find('a', attrs={"class": "_5twioja"}).text
							ws.cell(row=c, column=13).value = tp_c
						except:
							zzzz=1
							#print('NO VILLE')


					#NAME_HOTE
						try:
							tp_c=soup.find('div', attrs={"class": "_f47qa6"})
							tt=tp_c.find('div', attrs={"class": "_svr7sj"})
							tt1=tt.h2.get_text()
							pp=tt1.split('par ')
							ws.cell(row=c, column=3).value = pp[1]
						except:
							zzzz=1
							#print ('NO_NAME')
					#TYPE_HOME
						try:
							the_tr= soup.find('div', attrs={"class": "_1b3ij9t"}).text
							pp=the_tr.split('.')
							ws.cell(row=c, column=8).value = pp[0]
						except:
							try:
								the_tr= soup.find('div', attrs={"class": "_xcsyj0"}).text
								pp=the_tr.split('.')
								ws.cell(row=c, column=8).value = pp[0]
							except:
								zzzz=1
								#print('NOTYPE')
						
					#ANCIENNETE
						try:
							tp_c=soup.find('div', attrs={"class": "_f47qa6"})
							tt=tp_c.find('div', attrs={"class": "_svr7sj"})
							tt1=tt.div.get_text()
							ws.cell(row=c, column=4).value = tt1
						except:
							zzzz=1
							#print ('NOOLD')

					#SUPER HOTE
						try:
							#the_tr= soup.find('span', text=re.compile(r'\bSuperhost\b'),attrs = {'aria-hidden' : 'false'})
							tp_c=soup.find('div', attrs={"class": "_1ft6jxp"}).text
							ws.cell(row=c, column=16).value = 'X'
						except:
							zzzz=1
							#print('no superhote')
					#COMMENT PROFIL
						try:
							the_tr= soup.findAll('span', attrs = {'class' : '_pog3hg'})[0]
							ccc=the_tr.text
							pp=ccc.split('c')
							cc=pp[0]
							#div2=the_tr.findNextSibling('div')
							#print(the_tr.section.span.div.span.text)
							if cc=='Identité':
								cc=0
							ws.cell(row=c, column=17).value = cc
							#print(div2.text)
						except:
							zzzz=1
							#print('No Comment profil')
					#IDENTIFIE CHECK
						try:
							the_tr= soup.find('span', text=re.compile(r"\bIdentité vérifiée\b"))
							ws.cell(row=c, column=20).value = 'YES'
							#print(div2.text)
						except:
							ws.cell(row=c, column=20).value = 'NO'
							#print('No CHECK ID')
					#CO HOTE
						try:
							the_tr= soup.find('ul', attrs = {'class' : '_1omtyzc'})
							the_tr1= the_tr.findAll('li', attrs = {'class' : '_108byt5'})[0]
							tt11= the_tr1.find('a', attrs = {'target' : '_blank'})
							tt12= the_tr1.find('span', attrs = {'class' : '_1kfl0pr'})
							tt13= the_tr1.find('img', attrs = {'class' : '_6tbg2q'})
							div1=tt11['href']  #.attrs['href']
							the_tr2= the_tr.findAll('li', attrs = {'class' : '_108byt5'})[1]
							tt21= the_tr2.find('a', attrs = {'target' : '_blank'})
							tt22= the_tr2.find('span', attrs = {'class' : '_1kfl0pr'})
							tt23= the_tr2.find('img', attrs = {'class' : '_6tbg2q'})
							div2=tt21['href']  #.attrs['href']
							ws.cell(row=c, column=50).value = "https://www.airbnb.fr"+str(div1)
							ws.cell(row=c, column=51).value = tt12.text
							ws.cell(row=c, column=52).value = tt13['src']
							ws.cell(row=c, column=53).value = "https://www.airbnb.fr"+str(div2)
							ws.cell(row=c, column=54).value = tt22.text
							ws.cell(row=c, column=55).value = tt23['src']
							ws.cell(row=c, column=18).value = 2
						except:
							try:
								the_tr= soup.find('ul', attrs = {'class' : '_1omtyzc'})
								the_tr1= the_tr.find('li', attrs = {'class' : '_108byt5'})
								tt= the_tr1.find('a', attrs = {'target' : '_blank'})
								tt2= the_tr1.find('span', attrs = {'class' : '_1kfl0pr'})
								tt3= the_tr1.find('img', attrs = {'class' : '_6tbg2q'})
								div1=tt['href']  #.attrs['href']
								ws.cell(row=c, column=50).value = "https://www.airbnb.fr"+str(div1)
								ws.cell(row=c, column=51).value = tt2.text
								ws.cell(row=c, column=52).value = tt3['src']
								ws.cell(row=c, column=18).value = 1
							except:
								ws.cell(row=c, column=18).value = 0
								#print('no co hote')
					#PROPRETE
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[0]
							#print(tt.text)
							ws.cell(row=c, column=21).value = tt.text
						except:
							zzzz=1
							#print('no proprete')
					#PRECISION
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[1]
							#print(tt.text)
							ws.cell(row=c, column=22).value = tt.text
						except:
							zzzz=1
							#print('no Precision')
					#COMMUNICATION
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[2]
							#print(tt.text)
							ws.cell(row=c, column=23).value = tt.text
						except:
							zzzz=1
							#print('no communication')
					#EMPLACEMENT
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[3]
							#print(tt.text)
							ws.cell(row=c, column=24).value = tt.text
						except:
							zzzz=1
							#print('no emplacement')
					#ARRIVEE
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[4]
							#print(tt.text)
							ws.cell(row=c, column=25).value = tt.text
						except:
							zzzz=1
							#print('no arrivee')
					#QUALITY PRICE
						try:
							tt= soup.findAll('span', attrs={"class": "_4oybiu"})[5]
							#print(tt.text)
							ws.cell(row=c, column=26).value = tt.text
						except:
							zzzz=1
							#print('no price quality')
				#N° ENREGISTREMENT
						try:
							the_tr= soup.find('li', text=re.compile(r'\bNuméro\b'), attrs = {'class' : '_1q2lt74'})
							pp=the_tr.text
							#print(pp)
							sp=pp.split(' ')
							ws.cell(row=c, column=27).value = sp[-1]
						except:
							a=1
							#print('no N° enregistrement')				
				#TAUX REPONSE
						try:
							the_tr=soup.find('li', text=re.compile(r'\bTaux\b'))
							pp=the_tr.text
							pp=pp.replace(" ","")
							#print(pp)
							sp=pp.split(':')
							#print(sp[-1])
							ws.cell(row=c, column=28).value = sp[-1]
						except:
							zzzz=1
							#print('no taux réponse')
				#DELAI REPONSE
						try:
							the_tr=soup.find('li', text=re.compile(r'\bDélai\b'))
							pp=the_tr.text
							sp=pp.split(':')
							ws.cell(row=c, column=29).value = sp[-1]
						except:
							zzzz=1
							#print('no DELAI REPONSE')
				#DURING SEJOUR
						try:
							the_tr=soup.findAll('div', attrs={"class": "_1byskwn"})[-1]
							try:
								tt= the_tr.find('span', text=re.compile(r'\bArrivée\b'))
								ws.cell(row=c, column=31).value = tt.text
							except:
								zzzz=1
								#print('no ARRIVE')
							try:
								tt= the_tr.find('span', text=re.compile(r'\bDépart\b'))
								ws.cell(row=c, column=32).value = tt.text
							except:
								zzzz=1
								#print('no DEPART')
							try:
								tt= the_tr.find('span', text=re.compile(r'\bNon fumeur\b'))
								ws.cell(row=c, column=33).value = tt.text
							except:
								zzzz=1
								#print('no FUMEUR')
							try:
								tt= the_tr.find('span', text=re.compile(r'\bNe convient pas aux\b'))
								ws.cell(row=c, column=34).value = tt.text
							except:
								zzzz=1
								#print('no CHILD')
							try:
								tt= the_tr.find('span', text=re.compile(r"\bArrivée autonome\b"))
								ws.cell(row=c, column=35).value = tt.text
							except:
								zzzz=1
								#print('no AUTOMATIC')
							try:
								tt= the_tr.find('span', text=re.compile(r"\bPas d'animaux\b"))
								ws.cell(row=c, column=36).value = tt.text
							except:
								try:
									tt= the_tr.find('span', text=re.compile(r"\bAnimaux de compagnie\b"))
									ws.cell(row=c, column=36).value = tt.text
								except:
									zzzz=1
									#print('no ANIMAL')
							try:
								tt= the_tr.find('span', text=re.compile(r"\bCaution\b"))
								ws.cell(row=c, column=37).value = tt.text
							except:
								zzzz=1
								#print('no Caution')
							try:
								tt= the_tr.find('span', text=re.compile(r"\bDétecteur de fumée\b"))
								ws.cell(row=c, column=38).value = tt.text
							except:
								zzzz=1
								#print('no detecteur fumee')
							try:
								tt= the_tr.find('span', text=re.compile(r"\bDétecteur de monoxyde de carbone\b"))
								ws.cell(row=c, column=39).value = tt.text
							except:
								zzzz=1
								#print('no detecteur monoxyde')
							try:
								tt= the_tr.find('span', text=re.compile(r"\bPas de fête ni de soirée\b"))
								ws.cell(row=c, column=40).value = tt.text
							except:
								zzzz=1
								#print('no detecteur monoxyde')
							try:
								tt= the_tr.find('span', text=re.compile(r"\bmatière de distanciation sociale\b"))
								ws.cell(row=c, column=30).value = 'Y'
							except:
								ws.cell(row=c, column=30).value = 'N'
								#print('no distanciation sociale')
						except:
							print('no INSIDE RULE')
				#LANGUE
						try:
							the_tr= soup.find('li', text=re.compile(r'\bLangues\b'))
							#print(the_tr)
							pp=the_tr.text
							#print(pp)
							sp=pp.split(':')
							ws.cell(row=c, column=41).value = sp[-1]
						except:
							zzzz=1
							#print('no LANGUAGE')
				#IMAGE
						try:
							the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[0]
							#print(the_tr)
							tt=the_tr['src']
							ws.cell(row=c, column=42).value = tt
						except:
							try:
								the_tr= soup.find('img', attrs={"class": "_6tbg2q"})
								#print(the_tr)
								tt=the_tr['data-original-uri']
								ws.cell(row=c, column=42).value = tt
							except:
								zzzz=1
								#print('no IMAGE 0')
						try:
							the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[1]
							#print(the_tr)
							tt=the_tr['src']
							ws.cell(row=c, column=44).value = tt
						except:
							try:
								the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[1]
								#print(the_tr)
								tt=the_tr['data-original-uri']
								ws.cell(row=c, column=44).value = tt
							except:
								zzzz=1
								#print('no IMAGE 1')
						try:
							the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[2]
							#print(the_tr)
							tt=the_tr['src']
							ws.cell(row=c, column=45).value = tt
						except:
							try:
								the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[2]
								#print(the_tr)
								tt=the_tr['data-original-uri']
								ws.cell(row=c, column=45).value = tt
							except:
								zzzz=1
								#print('no IMAGE 2')
						try:
							the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[3]
							#print(the_tr)
							tt=the_tr['src']
							ws.cell(row=c, column=46).value = tt
						except:
							try:
								the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[3]
								#print(the_tr)
								tt=the_tr['data-original-uri']
								ws.cell(row=c, column=46).value = tt
							except:
								zzzz=1
								#print('no IMAGE 3')
						try:
							the_tr= soup.findAll('img', attrs={"class": "_9ofhsl"})[4]
							#print(the_tr)
							tt=the_tr['src']
							print(tt)
							ws.cell(row=c, column=47).value = tt
						except:
							try:
								the_tr= soup.findAll('img', attrs={"class": "_6tbg2q"})[4]
								#print(the_tr)
								tt=the_tr['data-original-uri']
								ws.cell(row=c, column=47).value = tt
							except:
								zzzz=1
								#print('no IMAGE 4')
				#IMAGE_HOTE
						try:
							the_tr= soup.find('div', attrs={"class": "_5kripx"})
							t= the_tr.find('img', attrs={"class": "_6tbg2q"})
							tt=t['src']
							ws.cell(row=c, column=43).value = tt
						except:
							zzzz=1
							#print('no IMAGE_HOTE')
						if (c/200).is_integer():
							wb.save(path_RESULT.filename)
						ws.cell(row=c, column=56).value = 'YES'

#------------------------
				except:
					try:
						#page sans annonce, je dois trouver quelque soit le type
						wait3.until(EC.presence_of_element_located((By.XPATH, "//h3[@class='_jmmm34f']")))
					except:
						# rien tourvé précédent, donc c'est que je suis sur mauvais design, clos et reopen chrome
						driver.quit()
						f_xpathdate=0
						fff=0
						fm=2
						driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver')
						#driver = webdriver.Chrome()
						driver.set_window_size(1500, 2000)
						wait3 = WebDriverWait(driver, 3)
						while f_xpathdate==0:
							h=ws.cell(row=fm, column=2).value
							print(h)
							if fff==10:
								f_mounth=1
								f_xpathdate=1
								end=0
							fff=fff+1
							try:
								driver.get(h)
								time.sleep(4)
								x_date = wait3.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_13m7kz7i']"))).text
								print("x date trouve")
								f_xpathdate=1
							except:
								if fff!=10:
									driver.quit()
									driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver')
									#driver = webdriver.Chrome()
									driver.set_window_size(1500, 2000)
									wait3 = WebDriverWait(driver, 3)
			c=c+1
	except:
		print("END")
		# EXCEPT si Chrome se ferme tout seul, ici il va le réouvrir et relancer la boucle d'extraction
		#driver = webdriver.Chrome()
		#driver.set_window_size(800, 1500)
		wb.save(path_RESULT.filename)
	wb.save(path_RESULT.filename)
	print ('_______    ___    ___     ___')
	print ('|      |   |  |   |  \    |  |')
	print ('|  |__     |  |   |   \   |  |')
	print ('|     |    |  |   |    \  |  |')
	print ('|  |       |  |   |  |\ \ |  |')
	print ('|  |       |  |   |  | \ \|  |')
	print ('|__|       |__|   |__|  \____|')
	print ('Votre fichier RESULT est à présent terminé, renommer le et concerver le.')
	print ('Vous pouvez maintenant:')
	print ('  1- exécuter le .exe N°4 qui calcule le nombre de logement détenus par hôte')
	print ('  2- exétuter le .exe N°5 qui lui calcule les nuitées de chaque annonces, à exécuter hebdomadairement')
	print ('  3- N°6 qui est à utiliser si votre fichier RESULT contient plus de 2000 annonces, il va découper le fichier en lot de moins de 2000 annonces si vous voulez l importer sur GoogleMap')
	print ('  4- N°7 vous permet de comparer le fichier RESULT avec un autre fichier de liste d annonce que vous possédez déjà, il va alors ajouter dans votre autre fichier les annonces manquante')
	end=1
try:
	driver.quit()
except:
	print('fin')
