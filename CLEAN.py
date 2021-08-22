import xlwt
import xlrd
import time
from xlutils.copy import copy
from decimal import Decimal
from decimal import *
import decimal
#import xlwings as xw
#from xlwings.constants import DeleteShiftDirection
from xlrd import open_workbook,XL_CELL_TEXT
import datetime
import datetime as dt
from tkinter import filedialog
from tkinter import *
import os
import re
import json
from urllib.request import urlopen
import pip
#import pandas as pd
import openpyxl
from openpyxl import load_workbook


wb = load_workbook(path_RESULT.filename)
ws=wb.active
nrow=ws.max_row


#-------------------------------------------------------------------------------------------------------------
#-------------------SCRAP DETAILS
#-------------------

#c = ligne 2 du xls resultant

def Clean(testmois):
	up=0
	i=1
	while up==0:
		V_up=ws.cell(row=1, column=i).value
		if V_up==testmois:
			up=1
		else:
			i=i+1
	#print('DIF_Comment='+str(i))
	mois=i
	c=2
	while c<=nrow:
		newlist=''
		vlist=[]
		data=ws.cell(row=c, column=mois).value
		if data is not None:
			vlist=data.split(';')
			#print (len(vlist))
			i=0
			while i<len(vlist):
				#if '20-6' not in vlist[i] and '21-6' not in vlist[i] and'22-6' not in vlist[i] and '23-6' not in vlist[i] and '24-6' not in vlist[i] and '25-6' not in vlist[i] and '26-6' not in vlist[i] and '30-6' not in vlist[i]:
				if 'R21-08' not in vlist[i] and 'L21-08' not in vlist[i] and 'C21-08' not in vlist[i]:
					if newlist=='':
						newlist=newlist+vlist[i]
					else:
						newlist=newlist+';'+vlist[i]
				i=i+1
			#print('----------LIST')
			#print(vlist)
			#print('----------NEWLIST')
			#print (newlist)
			ws.cell(row=c, column=mois).value=newlist
		c=c+1
run_mars=Clean('août 2021')
run_mars=Clean('juillet 2021')
run_mars=Clean('septembre 2021')
run_mars=Clean('octobre 2021')
run_mars=Clean('novembre 2021')
wb.save(path_RESULT.filename)

