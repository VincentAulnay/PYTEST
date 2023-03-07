print('start')
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import xlwt
import xlrd
import time
from xlutils.copy import copy
import datetime
import datetime as dt
from tkinter import filedialog
from tkinter import *
from bs4 import BeautifulSoup
from urllib.request import urlopen
import os
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
import threading
import sys



chrome_options = webdriver.ChromeOptions()
prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)
chrome_options.add_argument("window-size=2000,1000")
#chrome_options.add_argument("-headless")
#chrome_options.add_argument("-disable-gpu")
print ('▀▄▀▄▀▄ STOPBNB ▄▀▄▀▄▀')

now = str(datetime.datetime.now())[:19]
now = now.replace(":","_")
print(now)
print('tes')
#-----EXCEL RESULT OPEN AND READ-----

import re
import json
import csv
from google.oauth2 import service_account
import pygsheets
import pandas as pd

#wbx = load_workbook(path_RESULT.filename)
#ws = wbx.active
print('ici1')
try:
	client = pygsheets.authorize(service_account_file='/home/vincent/Desktop/raspbian-364809-be26e1ee6573.json')
except:
	client = pygsheets.authorize(service_account_file='/home/pi/Desktop/raspbian-364809-be26e1ee6573.json')
print('ici1')
#spreadsheet_url = "https://docs.google.com/spreadsheets/d/1vx34zctZXc2eQSFFe4I7zY6bjKJz9MtO7pgAIaQix4c/edit?usp=sharing"
spreadsheet_url = "https://docs.google.com/spreadsheets/d/14fiETLENGjJU3LMIybT-LAKw2DgXnol8ZAHgn46FkPs/edit?usp=sharing"
print('ici2')	

sheet_data = client.sheet.get('14fiETLENGjJU3LMIybT-LAKw2DgXnol8ZAHgn46FkPs')
print('ici3')
sheet = client.open('testoct')
print('ici4')
ws = sheet.worksheet_by_title('Sheet1')
print('sheet')
#print(ws)
#-------FIND COLUMN UPDATE------
up=0
k=1
while up==0:
	#V_up=sheet_read.cell(0,i).value
	#V_up=ws.cell(row=1, column=k).value
	V_up=ws.cell((1, k)).value
	print(V_up)
	if V_up=='ACTIVE_YES/NO':
		up=1
	else:
		k=k+1
		
c_mouth=k

#V_mouth=ws.cell(row=1, column=c_mouth).value
#if V_mouth!='3/5_mois':
#	ws.insert_cols(c_mouth)
#	ws.cell(row=1, column=c_mouth).value = '3/5_mois'

#-----RECUP INFO XPATH FROM EXCEL------
#book_GMAIL = xlrd.open_workbook('/home/pi/Desktop/GMAIL_ACCOUNT.xls')
#book_GMAIL = xlrd.open_workbook('/home/ubuntu/Desktop/GMAIL_ACCOUNT.xls')
#sheet_GMAIL = book_GMAIL.sheet_by_index(0)
#ADRESS_GMAIL=sheet_GMAIL.cell(0,1).value
#PSW_GMAIL=sheet_GMAIL.cell(1,1).value
#RECEIVER=sheet_GMAIL.cell(2,1).value

#-------DATE DU JOUR-------
date = int(datetime.datetime.now().day)
month = int(datetime.datetime.now().month)
Hr=dt.datetime.now().hour

#------RECUP INFO CALANDAR------

def email(DIR2,NAMEFile,now,total_R,total_L,total_P,total_C):
	#sender = ADRESS_GMAIL
	#sender_password = PSW_GMAIL
	#sender = 'stopbnb33650@gmail.com'
	#sender_password = '@stop$n$33650'
	print('ici1')
	sender = 'vincent.aulnay@gmx.fr'
	sender_password = '@Vincent94'
	receivers = ['vincent.aulnay@gmail.com', 'vincent.aulnay@gmx.fr']
	print('ici2')

	#s = smtplib.SMTP('smtp.gmail.com', 587)
	s = smtplib.SMTP('mail.gmx.com', 587)
	s.starttls()
	s.login(sender, sender_password)
	print('ici3')
	msg = MIMEMultipart()
	msg['From'] = sender
	msg['To'] = ", ".join(receivers)
	#msg['Subject'] = "Subject of the Mail- image -2"
	body = "Body_of_the_mail"
	msg.attach(MIMEText(body, 'plain'))
	msg['Subject'] = "STOP AIRBNB - "+str(now)+"-R"+str(total_R)+"-L"+str(total_L)+"-P"+str(total_P)+"-C"+str(total_C)
	# path along with extension of file to be attachmented 
	filename = DIR2+NAMEFile+str(now)+".xlsx"
	attachmentment = open(filename, "rb")
	 
	# instance of MIMEBase and named as p
	attachment = MIMEBase('application', 'octet-stream')
	# To change the payload into encoded form
	attachment.set_payload((attachmentment).read())
	# encode into base64
	encoders.encode_base64(attachment)
	attachment.add_header('Content-Disposition', "attachmentment; filename= %s" % filename)
	# attachment the instance  to instance 'msg'
	msg.attach(attachment)
	text = msg.as_string()
	print('ici4')
	s.sendmail(sender, receivers, text)
	print('*** email sent ***') 
	time.sleep(10)
	del filename
	del attachmentment
	del attachment
	del text
	del msg

def emailfalde():
	sender = ADRESS_GMAIL
	sender_password = PSW_GMAIL
	receivers = 'vincent.aulnay@gmail.com'

	s = smtplib.SMTP('smtp.gmail.com', 587)
	s.starttls()
	s.login(sender, sender_password)
	text = "echec de capture xpath mois"
	s.sendmail(sender, receivers, text)
	print('*** email sent ***') 
	time.sleep(10)
	del text

def emailfalde2():
	sender = ADRESS_GMAIL
	sender_password = PSW_GMAIL
	receivers = 'vincent.aulnay@gmail.com'

	s = smtplib.SMTP('smtp.gmail.com', 587)
	s.starttls()
	s.login(sender, sender_password)
	text = "echec de xpathdate"
	s.sendmail(sender, receivers, text)
	print('*** email sent ***') 
	time.sleep(10)
	del text

def whatmounth():
	month = int(datetime.datetime.now().month)
	global name_mois1
	global name_mois2
	global name_mois3
	global name_mois4
	global name_mois5
	if month==1:
		name_mois1='janvier 2023'
		name_mois2='février 2023'
		name_mois3='mars 2023'
		name_mois4='avril 2023'
		name_mois5='mai 2023'
	elif month==2:
		name_mois1='février 2023'
		name_mois2='mars 2023'
		name_mois3='avril 2023'
		name_mois4='mai 2023'
		name_mois5='juin 2023'
	elif month==3:
		name_mois1='mars 2023'
		name_mois2='avril 2023'
		name_mois3='mai 2023'
		name_mois4='juin 2023'
		name_mois5='juillet 2023'
	elif month==4:
		name_mois1='avril 2023'
		name_mois2='mai 2023'
		name_mois3='juin 2023'
		name_mois4='juillet 2023'
		name_mois5='août 2023'
	elif month==5:
		name_mois1='mai 2023'
		name_mois2='juin 2023'
		name_mois3='juillet 2023'
		name_mois4='août 2023'
		name_mois5='septembre 2023'
	elif month==6:
		name_mois1='juin 2023'
		name_mois2='juillet 2023'
		name_mois3='août 2023'
		name_mois4='septembre 2023'
		name_mois5='octobre 2023'
	elif month==7:
		name_mois1='juillet 2023'
		name_mois2='août 2023'
		name_mois3='septembre 2023'
		name_mois4='octobre 2023'
		name_mois5='novembre 2023'
	elif month==8:
		name_mois1='août 2023'
		name_mois2='septembre 2023'
		name_mois3='octobre 2023'
		name_mois4='novembre 2023'
		name_mois5='décembre 2023'
	elif month==9:
		name_mois1='septembre 2023'
		name_mois2='octobre 2023'
		name_mois3='novembre 2023'
		name_mois4='décembre 2023'
		name_mois5='janvier 2024'
	elif month==10:
		name_mois1='octobre 2023'
		name_mois2='novembre 2023'
		name_mois3='décembre 2023'
		name_mois4='janvier 2024'
		name_mois5='février 2024'
	elif month==11:
		name_mois1='novembre 2022'
		name_mois2='décembre 2022'
		name_mois3='janvier 2023'
		name_mois4='février 2023'
		name_mois5='mars 2023'
	elif month==12:
		name_mois1='décembre 2022'
		name_mois2='janvier 2023'
		name_mois3='février 2023'
		name_mois4='mars 2023'
		name_mois5='avril 2023'

def MnumDay (Mmois):
	global MNumday
	if Mmois=='janvier':
		MNumday=31
	elif Mmois=='février':
		MNumday=28
	elif Mmois=='mars':
		MNumday=31	
	elif Mmois=='avril':
		MNumday=30
	elif Mmois=='mai':
		MNumday=31
	elif Mmois=='juin':
		MNumday=30
	elif Mmois=='juillet':
		MNumday=31
	elif Mmois=='août':
		MNumday=30
	elif Mmois=='septembre':
		MNumday=31
	elif Mmois=='octobre':
		MNumday=30
	elif Mmois=='novembre':
		MNumday=31
	elif Mmois=='décembre':
		MNumday=30
		
def A_Colonne_mois(name_mois,c):
#1- récupération book Result qui évolue au court du script
#2- compter le nombre de colonne
#3- déterminer si colonne == name_mois de airbnb
#4- si condition alors c_write=c pour définir la colonne où écrire
	global c_write
	global new_month
	global k
	#book_mois = xlrd.open_workbook(path_RESULT.filename, on_demand = True)
	#sheet_mois = book_mois.sheet_by_index(0)
	#nc=sheet_mois.ncols
	#book_mois.release_resources()
	#del book_mois
	
	new_month=0
	
	find_month=0
	while find_month==0:
		this_month=ws.cell((1, c+1)).value
		print("le mois dans GS est : "+str(this_month))
		if this_month==name_mois:
			print('EQUAL MOUNTH')
			c_write=c+1
			c_stat='rien'
			c_stat=ws.cell((1, c+2)).value
			if c_stat!="STATE":
				ws.insert_cols(c+2)
				ws.cell(row=1, column=c+2).value = 'STATE'
			break
		#this_month=ws.cell(row=1, column=c+1).value
		#this_month=ws.cell((1, c+1)).value
		#if this_month==name_mois:
		#	c_write=c+1
		#	c_stat='rien'
		#	#c_stat=ws.cell(row=1, column=c+2).value
		#	c_stat=ws.cell((1, c+2)).value
		#	if c_stat!="STATE":
		#		ws.insert_cols(c+2)
				#ws.cell(row=1, column=c+2).value = 'STATE'
		#		ws.update_value((1,c+2), 'STATE')
				#wbx.save(path_RESULT.filename)
		#	break
		elif this_month=='':
			c_write=c+1
			ws.add_cols(1)
			ws.update_value((1,c+1), name_mois)
			ws.add_cols(1)
			ws.update_value((1,c+2), 'STATE')
			ws.add_cols(1)
			ws.update_value((1,c+3), 'NB_COMMENT')
			ws.add_cols(1)
			ws.update_value((1,c+4), 'DIF_COMMENT')
			ws.add_cols(1)
			ws.update_value((1,c+5), 'Sum_Nuitee')
			ws.add_cols(1)
			ws.update_value((1,c+6), 'Nuitee_bloquee')
			ws.add_cols(1)
			ws.update_value((1,c+7), 'R5_ACT')
			ws.add_cols(1)
			ws.update_value((1,c+8), 'R15_ACT')
			ws.add_cols(1)
			ws.update_value((1,c+9), 'R30_ACT')
			ws.add_cols(1)
			ws.update_value((1,c+10), 'R5_Jours')
			ws.add_cols(1)
			ws.update_value((1,c+11), 'R15_Jours')
			ws.add_cols(1)
			ws.update_value((1,c+12), 'R30_Jours')
			ws.add_cols(1)
			ws.update_value((1,c+13), 'L_ACT')
			ws.add_cols(1)
			ws.update_value((1,c+14), 'P_NB')
			ws.add_cols(1)
			ws.update_value((1,c+15), 'D_Jours')
			ws.add_cols(1)
			#c_write=c+1
			find_month=1
			new_month=1
			print ('plus une colonne')
			break
		else:
			c=c+1
	k=c

def A_Statu_PLUS(date,c_write,page,j,g,ResAirbnb,new_mo,MNday,ONCOM):	
	int_timeday=int(date)
	month=soup.findAll('div', attrs={"class":u"_1lds9wb"})[g]
	i=0
	li=[]
	#SEMAINE
	tr=1
	#JOUR
	td=1
	#x_ = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_14676s3']/div[2]/div/div["+str(page)+"]//tr[2]/td[6]/span/div/div/div"))).text
	#print (x_)
	ResAirbnb='/R'
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			try:
				the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\buniquement\b')})[i]
				#div=the_tr.span.div.div.div.get_text()
				div=the_tr.div.get_text()
				#_1lds9wb
				intdiv=int(div)
				if intdiv>=int_timeday:
					li.append(intdiv)
			except:
				z=0
			try:
				the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\bNon\b')})[i]
		#div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
				#div=the_tr.span.div.div.div.get_text()
				div=the_tr.div.get_text()
				intdiv=int(div)
				if intdiv>=int_timeday:
					li.append(intdiv)
			except:
				z=0
			i=i+1
		except:
			break
	li.sort()
	#back_li=ws.cell(row=j, column=c_write+1).value
	back_li=ws.cell((j, c_write+1)).value
	try:
		if back_li!='[]':
			back_li=back_li.replace("[","")
			back_li=back_li.replace("]","")
			back_li=back_li.split(",")
			i=0
			bl=[]
			while i!=len(back_li):
				ivb=int(back_li[i])
				if ivb>=int_timeday:
					bl.append(ivb)
				i=i+1
			back_li=bl
			#print ("back_li="+str(back_li))
		else:
			back_li=[]
	except:
		back_li=[]
	#ws.cell(row=j, column=c_write+1).value = str(li)
	ws.update_value((j,c_write+1), str(li))
	#print(li)
	c_added=[]
	c_remove=[]
	c_added=[elem for elem in li if elem not in back_li ]
	c_remove=[elem for elem in back_li if elem not in li ]
	#print(c_added)
	#print(c_remove)
	date = int(datetime.datetime.now().day)
	month = int(datetime.datetime.now().month)
	toto=str(date).zfill(2)+'-'+str(month).zfill(2)
	t_add='vide'
	t_rem='vide'
	t_wri='vide'
	if len(c_added)>0:
		if len(c_added)==1:
			dif=c_added[0]-date
			if dif==0 or dif==1 or dif==2 or dif==6:
				ResAirbnb='/P'
			elif dif<0:
				difP=MNday-date+lie[0]
				if difP==0 or difP==1 or difP==2 or difP==6:
					ResAirbnb='/P'		
		t_add=ResAirbnb+toto+':'+str(c_added)
		t_add=t_add.replace("[","")
		t_add=t_add.replace("]","")
		#print (t_add)
	if c_remove!=['']:
		if c_remove!=[]:
			t_rem='/L'+toto+':'+str(c_remove)
			t_rem=t_rem.replace("[","")
			t_rem=t_rem.replace("]","")
			#print(t_rem)
	#ca=ws.cell(row=j, column=c_write).value
	ca=ws.cell((j, c_write)).value
	if ca==None:
		if t_add!='vide':
			t_wri=str(t_add)
	else:
		if t_add!='vide':
			if t_rem!='vide':
				t_wri=str(t_add)+';    '+str(t_rem)
			else:
				t_wri=str(t_add)
		else:
			if t_rem!='vide':
				t_wri=str(t_rem)
		if t_wri!='vide':
			t_wri=str(ca)+';    '+t_wri
	if t_wri!='vide':
		print(t_wri)
		#ws.cell(row=j, column=c_write).value=t_wri
		ws.update_value((j,c_write), t_wri)
	#COMMENTAIRE
	ONC=ONCOM
	if ONC==1:
		try:
			#//span[@class='_so3dpm2']
			#Bcomment=soup.find('button', attrs={"class": "_ff6jfq"})
			#Scomment=Bcomment.find('span', attrs={"class": "_so3dpm2"}).text
			Lcomment=[]
			Icomment=0
			Scomment=soup.find('span', attrs={"class": "_bq6krt"}).text
			Lcomment=Scomment.split("(")
			Tcomment=Lcomment[1].replace(")","")
			Icomment=int(Tcomment)
			#ws.cell(row=j, column=c_write+2).value=Icomment
			ws.update_value((j,c_write+2), Icomment)
		except:
			pass

def A_Statu_PLUS2(c_write,j,ResAirbnb,new_mo,page):	
	month=soup.find('div', attrs={"class":u"_kuxo8ai"})
	i=0
	li=[]
	#SEMAINE
	tr=1
	#JOUR
	td=1
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			try:
				the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\buniquement\b')})[i]
				#div=the_tr.span.div.div.div.get_text()
				div=the_tr.div.get_text()
				#_1lds9wb
				intdiv=int(div)
				if intdiv>=int_timeday:
					li.append(intdiv)
			except:
				z=0
			try:
				the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\bNon\b')})[i]
		#div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
				#div=the_tr.span.div.div.div.get_text()
				div=the_tr.div.get_text()
				intdiv=int(div)
				if intdiv>=int_timeday:
					li.append(intdiv)
			except:
				z=0
			i=i+1
		except:
			break
	li.sort()
	#back_li=ws.cell(row=j, column=c_write+1).value
	back_li=ws.cell((j, c_write+1)).value
	if back_li!=None:
		back_li=back_li.replace("[","")
		back_li=back_li.replace("]","")
		back_li=back_li.split(",")
		i=0
		bl=[]
		while i!=len(back_li):
			ivb=int(back_li[i])
			bl.append(ivb)
			i=i+1
		back_li=bl
		#print ("back_li="+str(back_li))
	else:
		back_li=[]
	#ws.cell(row=j, column=c_write+1).value = str(li)
	ws.update_value((j,c_write+1), str(li))
	print(li)
	c_added=[]
	c_remove=[]
	c_added=[elem for elem in li if elem not in back_li ]
	c_remove=[elem for elem in back_li if elem not in li ]
	#print(c_added)
	#print(c_remove)
	date = int(datetime.datetime.now().day)
	month = int(datetime.datetime.now().month)
	toto=str(date).zfill(2)+'-'+str(month).zfill(2)
	t_add='vide'
	t_rem='vide'
	t_wri='vide'
	if len(c_added)>0:		
		t_add=ResAirbnb+toto+':'+str(c_added)
		t_add=t_add.replace("[","")
		t_add=t_add.replace("]","")
		#print (t_add)
	if c_remove!=['']:
		if c_remove!=[]:
			t_rem='/L'+toto+':'+str(c_remove)
			t_rem=t_rem.replace("[","")
			t_rem=t_rem.replace("]","")
			#print(t_rem)
	#ca=ws.cell(row=j, column=c_write).value
	ca=ws.cell((j, c_write)).value
	if ca==None:
		if t_add!='vide':
			t_wri=str(t_add)
	else:
		if t_add!='vide':
			if t_rem!='vide':
				t_wri=str(t_add)+';    '+str(t_rem)
			else:
				t_wri=str(t_add)
		else:
			if t_rem!='vide':
				t_wri=str(t_rem)
		if t_wri!='vide':
			t_wri=str(ca)+';    '+t_wri
	if t_wri!='vide':
		#print(t_wri)
		#ws.cell(row=j, column=c_write).value=t_wri
		ws.update_value((j,c_write+2), t_wri)
			
def A_Statu_day2(date,c_write,page,j,g,ResAirbnb,new_mo,MNday,ONCOM,des):	
	global R1
	global R2
	global L1
	global L2
	global P1
	global P2
	global C1
	P=0
	R=0
	L=0
	C=0
	int_timeday=int(date)
	month=soup.findAll('div', attrs={"class":u"_1lds9wb"})[g]
	i=0
	li=[]
	fakeli=[]
	ResAirbnb='/R'
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			if des==0:
				the_tr= month.findAll('td', attrs={"class": "_l9wspk2"})[i]
				#div=the_tr.span.div.div.div.get_text()
				div=the_tr.div.get_text()
				#_1lds9wb
				intdiv=int(div)
				if intdiv>=int_timeday:
					li.append(intdiv)
			if des==1:
				try:
					#the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\buniquement\b')})[i]
					the_tr= month.find_all(attrs={'aria-label':re.compile('uniquement')})[i]
					#div=the_tr.span.div.div.div.get_text()
					div=the_tr.div.get_text()
					#_1lds9wb
					intdiv=int(div)
					if intdiv>=int_timeday:
						li.append(intdiv)
				except:
					z=0
				try:
					#the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\bNon\b')})[i]
					the_tr= month.find_all(attrs={'aria-label':re.compile('Non')})[i]
					#div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
					#div=the_tr.span.div.div.div.get_text()
					div=the_tr.div.get_text()
					intdiv=int(div)
					if intdiv>=int_timeday:
						li.append(intdiv)
				except:
					z=0
				try:
					#the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\pas\b')})[i]
					the_tr= month.find_all(attrs={'aria-label':re.compile('pas')})[i]
					div=the_tr.div.get_text()
					intdiv=int(div)
					if intdiv>=int_timeday:
						#li.append(intdiv)
						fakeli.append(intdiv)
				except:
					z=0
		except:
			break
		i=i+1
	li.sort()
	#print(li)
	#back_li=ws.cell(row=j, column=c_write+1).value
	back_li=ws.cell((j, c_write+1)).value
	#print(back_li)
	try:
		if back_li!=[]:
			back_li=back_li.replace("[","")
			back_li=back_li.replace("]","")
			back_li=back_li.split(",")
			i=0
			bl=[]
			while i!=len(back_li):
				ivb=int(back_li[i])
				if ivb>=int_timeday:
					bl.append(ivb)
				i=i+1
			back_li=bl
			#print ("back_li="+str(back_li))
		else:
			back_li=[]
	except:
		back_li=[]
	#ws.cell(row=j, column=c_write+1).value = str(li)
	ws.update_value((j,c_write+1), str(li))
	#print(str(li))
	c_added=[]
	c_remove=[]
	c_added=[elem for elem in li if elem not in back_li ]
	c_remove=[elem for elem in back_li if elem not in li ]
	#print(c_added)
	#print(c_remove)
	date = int(datetime.datetime.now().day)
	month = int(datetime.datetime.now().month)
	toto=str(date).zfill(2)+'-'+str(month).zfill(2)
	t_add='vide'
	t_rem='vide'
	t_wri='vide'
	ONC=ONCOM
	write_c=0
	if ONC==1:
		try:
			#oldcc=ws.cell(row=j, column=c_write+2).value
			oldcc=ws.cell((j, c_write+2)).value
			#try:
			#	tp_c=soup.findAll('span', attrs={"class": "_bq6krt"})[1].text
			#except:
			#	tp_c=soup.find('span', attrs={"class": "_142pbzop"}).text
			#print(tp_c)
			try:
				tp_c=soup.find('span', attrs={"class": "_s65ijh7"}).text
			except:
				aaa=1
			p_c=tp_c.replace("(","")
			cc=p_c.replace(")","")
			try:
				pp=cc.split(' ')
				#print(pp)
				cc=pp[0]
			except:
				pass
			#ws.cell(row=j, column=c_write+2).value=cc
			ws.update_value((j,c_write+2), cc)
			if oldcc!=cc:
				C1=1
				write_c=1
				v_com=int(cc)-int(oldcc)
				t_com='/C'+toto+':'+str(v_com)
		except:
			pass
	
	if len(c_added)>0:
		R=1
		if len(c_added)==1:
			dif=c_added[0]-date
			if dif==0 or dif==1 or dif==2 or dif==3 or dif==6 or dif==7:
				ResAirbnb='/P'
				P=1
				R=0
			elif dif<0:
				difP=MNday-date+lie[0]
				if difP==0 or difP==1 or difP==2 or dif==3 or difP==6 or dif==7:
					ResAirbnb='/P'
					P=1
					R=0
		t_add=ResAirbnb+toto+':'+str(c_added)
		t_add=t_add.replace("[","")
		t_add=t_add.replace("]","")
		#print (t_add)
	if c_remove!=['']:
		if c_remove!=[]:
			t_rem='/L'+toto+':'+str(c_remove)
			t_rem=t_rem.replace("[","")
			t_rem=t_rem.replace("]","")
			L=1
			#print(t_rem)
	#ca=ws.cell(row=j, column=c_write).value
	ca=ws.cell((j, c_write)).value
	if ca=='':
		if t_add!='vide':
			t_wri=str(t_add)
	else:
		if t_add!='vide':
			if t_rem!='vide':
				t_wri=str(t_add)+';    '+str(t_rem)
			else:
				t_wri=str(t_add)
		else:
			if t_rem!='vide':
				t_wri=str(t_rem)
		if write_c==1:
			if t_wri=='vide':
				t_wri=str(t_com)
			else:
				t_wri=t_wri+';    '+str(t_com)
			C1=1
		if t_wri!='vide':
			t_wri=str(ca)+';    '+t_wri

	
	if t_wri!='vide':
		#ws.cell(row=j, column=c_write).value=t_wri
		ws.update_value((j,c_write), t_wri)
	if g==0:
		R1=R
		L1=L
		P1=P
	else:
		R2=R
		L2=L
		P2=P


def A_Statu_day4(c_write,j,ResAirbnb,new_mo,des):
	global R3
	global L3
	R3=0
	L3=0
	month=soup.find('div', attrs={"class":u"_kuxo8ai"})
	#print('fevrier')
	i=0
	li=[]
	ResAirbnb='/R'
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			if des==0:
				the_tr= month.findAll('td', attrs={"class": "_l9wspk2"})[i]
				div=the_tr.span.div.div.div.get_text()
				#_1lds9wb
				intdiv=int(div)
				li.append(intdiv)
			if des==1:
				try:
					#the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\buniquement\b')})[i]
					the_tr= month.find_all(attrs={'aria-label':re.compile('uniquement')})[i]
					#div=the_tr.span.div.div.div.get_text()
					div=the_tr.div.get_text()
					#_1lds9wb
					intdiv=int(div)
					li.append(intdiv)
				except:
					z=0
				try:
					#the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\bNon\b')})[i]
					the_tr= month.find_all(attrs={'aria-label':re.compile('Non')})[i]
					#div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
					#div=the_tr.span.div.div.div.get_text()
					div=the_tr.div.get_text()
					intdiv=int(div)
					li.append(intdiv)
				except:
					z=0
				try:
					#the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\pas\b')})[i]
					the_tr= month.find_all(attrs={'aria-label':re.compile('pas')})[i]
					div=the_tr.div.get_text()
					intdiv=int(div)
					#if intdiv>=int_timeday:
					#	li.append(intdiv)
				except:
					z=0
		except:
			break
		i=i+1
	li.sort()
	#back_li=ws.cell(row=j, column=c_write+1).value
	back_li=ws.cell((j, c_write+1)).value
	#print(back_li)
	try:
		if back_li!=[]:
			back_li=back_li.replace("[","")
			back_li=back_li.replace("]","")
			back_li=back_li.split(",")
			i=0
			bl=[]
			while i!=len(back_li):
				ivb=int(back_li[i])
				bl.append(ivb)
				i=i+1
			back_li=bl
			#print ("back_li="+str(back_li))
		else:
			back_li=[]
	except:
		back_li=[]
	#ws.cell(row=j, column=c_write+1).value = str(li)
	ws.update_value((j,c_write+1), str(li))
	#print(li)
	c_added=[]
	c_remove=[]
	c_added=[elem for elem in li if elem not in back_li ]
	c_remove=[elem for elem in back_li if elem not in li ]
	#print(c_added)
	#print(c_remove)
	date = int(datetime.datetime.now().day)
	month = int(datetime.datetime.now().month)
	toto=str(date).zfill(2)+'-'+str(month).zfill(2)
	t_add='vide'
	t_rem='vide'
	t_wri='vide'
	if len(c_added)>0:
		R3=1
		t_add=ResAirbnb+toto+':'+str(c_added)
		t_add=t_add.replace("[","")
		t_add=t_add.replace("]","")
		#print (t_add)
	if c_remove!=['']:
		if c_remove!=[]:
			prefix='/L'
			if len(c_remove)==1:
				prefix='/X'
			else:
				L3=1
			t_rem=prefix+toto+':'+str(c_remove)
			t_rem=t_rem.replace("[","")
			t_rem=t_rem.replace("]","")
			#print(t_rem)
	#ca=ws.cell(row=j, column=c_write).value
	ca=ws.cell((j, c_write)).value
	print(ca)
	if ca=='':
		if t_add!='vide':
			t_wri=str(t_add)
	else:
		if t_add!='vide':
			if t_rem!='vide':
				t_wri=str(t_add)+';    '+str(t_rem)
			else:
				t_wri=str(t_add)
		else:
			if t_rem!='vide':
				t_wri=str(t_rem)
		if t_wri!='vide':
			t_wri=str(ca)+';    '+t_wri
	if t_wri!='vide':
		#print(t_wri)
		#ws.cell(row=j, column=c_write).value=t_wri
		ws.update_value((j,c_write), t_wri)
		
def A_Statu_day5(c_write,j,ResAirbnb,new_mo,g,des):	
	global R4
	global R5
	global L4
	global L5
	R=0
	L=0
	month=soup.findAll('div', attrs={"class":u"_1lds9wb"})[g]
	i=0
	li=[]
	ResAirbnb='/R'
	if new_mo==1:
		ResAirbnb='/D'
	while i<=31:
		try:
			if des==0:
				the_tr= month.findAll('td', attrs={"class": "_l9wspk2"})[i]
				#div=the_tr.span.div.div.div.get_text()
				div=the_tr.div.get_text()
				#_1lds9wb
				intdiv=int(div)
				li.append(intdiv)
			if des==1:
				try:
					the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\buniquement\b')})[i]
					#div=the_tr.span.div.div.div.get_text()
					div=the_tr.div.get_text()
					#_1lds9wb
					intdiv=int(div)
					li.append(intdiv)
				except:
					z=0
				try:
					the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\bNon\b')})[i]
					#div=the_tr.span.div.div.div.get_text()
					div=the_tr.div.get_text()
					#div=the_tr.find('div', attrs={"class": "_13m7kz7i"}).text
					intdiv=int(div)
					li.append(intdiv)
				except:
					z=0
				try:
					the_tr= month.findAll('td', attrs={'aria-label':re.compile(r'\pas\b')})[i]
					div=the_tr.div.get_text()
					intdiv=int(div)
				#	if intdiv>=int_timeday:
				#		li.append(intdiv)
				except:
					z=0
		except:
			break
		i=i+1
	li.sort()
	#print (li)
	#back_li=ws.cell(row=j, column=c_write+1).value
	back_li=ws.cell((j, c_write+1)).value
	#print(back_li)
	try:
		if back_li!=[]:
			back_li=back_li.replace("[","")
			back_li=back_li.replace("]","")
			back_li=back_li.split(",")
			i=0
			bl=[]
			while i!=len(back_li):
				ivb=int(back_li[i])
				bl.append(ivb)
				i=i+1
			back_li=bl
			#print ("back_li="+str(back_li))
		else:
			back_li=[]
	except:
		back_li=[]
	#ws.cell(row=j, column=c_write+1).value = str(li)
	ws.update_value((j,c_write+1), str(li))
	#print(li)
	c_added=[]
	c_remove=[]
	c_added=[elem for elem in li if elem not in back_li ]
	c_remove=[elem for elem in back_li if elem not in li ]
	#print(c_added)
	#print(c_remove)
	date = int(datetime.datetime.now().day)
	month = int(datetime.datetime.now().month)
	toto=str(date).zfill(2)+'-'+str(month).zfill(2)
	t_add='vide'
	t_rem='vide'
	t_wri='vide'
	if len(c_added)>0:		
		t_add=ResAirbnb+toto+':'+str(c_added)
		t_add=t_add.replace("[","")
		t_add=t_add.replace("]","")
		R=1
		#print (t_add)
	if c_remove!=['']:
		if c_remove!=[]:
			prefix='/L'
			if len(c_remove)==1:
				prefix='/X'
			else:
				L=1
			t_rem=prefix+toto+':'+str(c_remove)
			t_rem=t_rem.replace("[","")
			t_rem=t_rem.replace("]","")
			#print(t_rem)
	#ca=ws.cell(row=j, column=c_write).value
	ca=ws.cell((j, c_write)).value
	if ca=='':
		if t_add!='vide':
			t_wri=str(t_add)
	else:
		if t_add!='vide':
			if t_rem!='vide':
				t_wri=str(t_add)+';    '+str(t_rem)
			else:
				t_wri=str(t_add)
		else:
			if t_rem!='vide':
				t_wri=str(t_rem)
		if t_wri!='vide':
			t_wri=str(ca)+';    '+t_wri
	if t_wri!='vide':
		#print(t_wri)
		#ws.cell(row=j, column=c_write).value=t_wri
		ws.update_value((j,c_write), t_wri)
	if g==0:
		R4=R
		L4=L
	else:
		R5=R
		L5=L
	

def COMPUTE_M1(name_mois1):
	Dif_c=1
	if Dif_c==1:
		up=0
		i=1
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up==name_mois1:
				up=1
			else:
				i=i+1
		print('Cmois='+str(i))
		Cmois=i

		up=0
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='NB_COMMENT':
				up=1
			else:
				i=i+1
		print('Ccomment1='+str(i))
		Ccomment1=i

		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='DIF_COMMENT':
				up=1
			else:
				i=i+1
		print('DIF_Comment='+str(i))
		DIF_Comment=i
		
		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='Sum_Nuitee':
				up=1
			else:
				i=i-1
		C_N1SumNuitee=i
		
		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='Sum_Nuitee':
				up=1
			else:
				i=i+1
		C_SumNuitee=i

		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='Nuitee_bloquee':
				up=1
			else:
				i=i+1
		C_NuitBloquee=i
		
		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='R5_ACT':
				up=1
			else:
				i=i+1
		C_R5A=i

		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='R15_ACT':
				up=1
			else:
				i=i+1
		C_R15A=i
		
		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='R30_ACT':
				up=1
			else:
				i=i+1
		C_R30A=i
		
		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='R5_Jours':
				up=1
			else:
				i=i+1
		print('SUM_nJ='+str(i))
		C_R5J=i

		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='R15_Jours':
				up=1
			else:
				i=i+1
		C_R15J=i
		
		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='R30_Jours':
				up=1
			else:
				i=i+1
		C_R30J=i
		
		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='L_ACT':
				up=1
			else:
				i=i+1
		C_L=i
		
		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='P_NB':
				up=1
			else:
				i=i+1
		C_P=i
		
		up=0
		i=Cmois
		while up==0:
			#V_up=ws.cell(row=1, column=i).value
			V_up=ws.cell((1, i)).value
			if V_up=='D_Jours':
				up=1
			else:
				i=i+1
		C_D=i
		
		up=0
		i=Cmois
		try:
			while up==0:
				#V_up=ws.cell(row=1, column=i).value
				V_up=ws.cell((1, i)).value
				if V_up=='NB_COMMENT':
					up=1
				else:
					i=i-1
			print('Ccommont2='+str(i))
			Ccomment2=i
			NOC2=0
		except:
			NOC2=1
			print ('NOC2=====1')
	c=2
	while c<=nrow:
		if NOC2==0:
			#V1=ws.cell(row=c, column=Ccomment1).value
			#V2=ws.cell(row=c, column=Ccomment2).value
			V1=ws.cell((c, Ccomment1)).value
			V2=ws.cell((c, Ccomment2)).value
			try:
				DIF=int(V1)-int(V2)
				print('ANNONCE:'+str(c)+('   DIF:')+str(DIF))
				#ws.cell(row=c, column=DIF_Comment).value=DIF
				ws.update_value((c,DIF_Comment), DIF)
			except:
				pass
	#--------COUNT NB/A and NB NO/A---------
		print('start 01')
		#STR_NBA=ws.cell(row=c, column=Cmois).value
		STR_NBA=ws.cell((c, Cmois)).value
		print(STR_NBA)
		#if STR_NBA==None:
		#	STR_NBA="/X1-7:3"
		continu=1
		if STR_NBA=='':
			continu=0
			#ws.cell(row=c, column=C_L).value=0
			#ws.cell(row=c, column=C_P).value=0
			#ws.cell(row=c, column=C_D).value=0
			#ws.cell(row=c, column=C_R5A).value=0
			#ws.cell(row=c, column=C_R15A).value=0
			#ws.cell(row=c, column=C_R30A).value=0
			#ws.cell(row=c, column=C_R5J).value=0
			#ws.cell(row=c, column=C_R15J).value=0
			#ws.cell(row=c, column=C_R30J).value=0
			#ws.cell(row=c, column=C_NuitBloquee).value=0
			#ws.update_value((c,C_L), 0)
			ws.cell((c, C_L)).value = 0
			#ws.update_value((c,C_P), 0)
			ws.cell((c, C_P)).value = 0
			#ws.update_value((c,C_D), 0)
			ws.cell((c, C_D)).value = 0
			#ws.update_value((c,C_R5A), 0)
			ws.cell((c, C_R5A)).value = 0
			#ws.update_value((c,C_R15A), 0)
			ws.cell((c, C_R15A)).value = 0
			#ws.update_value((c,C_R30A), 0)
			ws.cell((c, C_R30A)).value = 0
			#ws.update_value((c,C_R5J), 0)
			ws.cell((c, C_R5J)).value = 0
			#ws.update_value((c,C_R15J), 0)
			ws.cell((c, C_R15J)).value = 0
			#ws.update_value((c,C_R30J), 0)
			ws.cell((c, C_R30J)).value = 0
			ws.update_value((c,C_NuitBloquee), 0)
			#N1nuit=ws.cell(row=c, column=C_N1SumNuitee).value
			N1nuit=ws.cell((c, C_N1SumNuitee)).value
			if N1nuit is None:
				N1nuit=0
			#ws.cell(row=c, column=C_SumNuitee).value=N1nuit
			#ws.update_value((c,C_SumNuitee), N1nuit)
			ws.cell((c, C_SumNuitee)).value = N1nuit
		if continu==1:
			print('start 02')
			count_R=STR_NBA.count('/R')
			count_L=STR_NBA.count('/L')
			count_P=STR_NBA.count('/P')
			#ws.cell(row=c, column=C_L).value=count_L
			#ws.cell(row=c, column=C_P).value=count_P
			ws.update_value((c,C_L), count_L)
			ws.update_value((c,C_P), count_P)
					#---------COUNT nJ ---------
			print('start 03')
			list=STR_NBA.split(';')
			B=['/P', '/D', '/R', '/L', '/X']
			blacklist = re.compile('|'.join([re.escape(word) for word in B]))
			newL=[word for word in list if not blacklist.search(word)]
			D=['/D']
			blacklistD = re.compile('|'.join([re.escape(wordD) for wordD in D]))
			newLforD=[wordD for wordD in list if blacklistD.search(wordD)]
			rd=0
			lenD=len(newLforD)
			nbD=0
			print('start 04')
			while rd<lenD:
				pnlD=newLforD[rd].split(':')
				del pnlD[0]
				pld=pnlD[0].split(',')
				nbD=nbD+len(pld)
				rd=rd+1
			#ws.cell(row=c, column=C_D).value=nbD
			ws.update_value((c,C_D), nbD)
			#[x for x in list if not x.startswith('/A/P') and not x.startswith('/D') and not x.startswith('/P')]
			#[x for x in list if not any(bad in x for bad in B)]
			#-----/A--------
			print('start 05')
			BA=['/R']
			blacklistA = re.compile('|'.join([re.escape(wordA) for wordA in BA]))
			newLforA=[wordA for wordA in list if blacklistA.search(wordA)] #-------Creation list AVEC que les lot /R
			newLfornoA=[wordA for wordA in list if not blacklistA.search(wordA)] #-------Creation list SANS les lot /R
			nAlen=len(newLforA)
			rr=0
			nbA=0
			R5=0
			R15=0
			R30=0
			print('start 06')
			try: #---Recuperation nJ dans les lot /R
				while rr<nAlen:
					pnlA=newLforA[rr].split(':')
					del pnlA[0]
					pla=pnlA[0].split(',')
					if len(pla)<=5:
						R5=R5+1
					elif len(pla)<=15:
						R15=R15+1
					else:
						R30=R30+1
					nbA=nbA+len(pla)
					rr=rr+1
				ws.update_value((c,C_R5A), R5)
				ws.update_value((c,C_R15A), R15)
				ws.update_value((c,C_R30A), R30)
			except:
				pass
			print('start 07')
			list.reverse()
			print(list)
			s=0
			listL=[]
			listR=[]
			LR1=[]
			LR2=[]
			LR11=[]
			LR22=[]
			NR5=0
			NR15=0
			NR30=0
			sp3=[]
			while s<len(list):
				if "/R" in list[s]:
					sp1=list[s].split(":")
					sp2=sp1[1].split(",")
					ffloat=0
					while ffloat<len(sp2):
						fsp2=int(sp2[ffloat])
						sp3.append(fsp2)
						ffloat=ffloat+1
					LR1=[elem for elem in sp3 if elem not in listR ]
					LR2=[elem for elem in LR1 if elem not in listL ]
					if len(LR2)<=5:
						NR5=NR5+len(LR2)
					elif len(LR2)<=15:
						NR15=NR15+len(LR2)
					else:
						NR30=NR30+len(LR2)
					listR.extend(LR2)
				elif "/L" in list[s]:
					sp1=list[s].split(":")
					sp2=sp1[1].split(",")
					ffloat=0
					while ffloat<len(sp2):
						fsp2=int(sp2[ffloat])
						sp3.append(fsp2)
						ffloat=ffloat+1
					listL.extend(sp3)
				elif "/X" in list[s]:
					sp1=list[s].split(":")
					sp2=sp1[1].split(",")
					ffloat=0
					while ffloat<len(sp2):
						fsp2=int(sp2[ffloat])
						sp3.append(fsp2)
						ffloat=ffloat+1
					listL.extend(sp3)
				s=s+1
			#ws.cell(row=c, column=C_R5J).value=NR5
			#ws.cell(row=c, column=C_R15J).value=NR15
			#ws.cell(row=c, column=C_R30J).value=NR30
			#ws.cell(row=c, column=C_NuitBloquee).value=NR30+NR15+NR5
			print('start 071')
			#ws.update_value((c,C_R5J), NR5)
			ws.cell((c, C_R5J)).value = NR5
			#ws.update_value((c,C_R15J), NR15)
			ws.cell((c, C_R15J)).value = NR15
			#ws.update_value((c,C_30J), NR30)
			ws.cell((c, C_R30J)).value = NR30
			#ws.update_value((c,C_NuitBloquee), NR30+NR15+NR5)
			ws.cell((c, C_NuitBloquee)).value = NR30+NR15+NR5
			#N1nuit=ws.cell(row=c, column=C_N1SumNuitee).value
			print('start 072')
			N1nuit=ws.cell((c, C_N1SumNuitee)).value
			print(N1nuit)
			print('start 08')
			if N1nuit is None:
				N1nuit=0
			print('start 081')
			#ws.cell(row=c, column=C_SumNuitee).value=N1nuit+NR30+NR15+NR5
			#ws.update_value((c,C_SumNuitee), N1nuit+NR30+NR15+NR5)
			ws.cell((c, C_SumNuitee)).value = N1nuit+NR30+NR15+NR5
			print('start 09')
		c=c+1
	#wbx.save(path_RESULT.filename)
	time.sleep(5)

		
#-----OPEN GOOGLE CHROME and AIRBNB PAGE---------

rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
#rootdriver.set_page_load_timeout(2)
rootdriver.set_window_size(2000, 1000)
rootdriver.get('https://www.google.com/')
rootdriver.get('chrome://settings/')
rootdriver.execute_script('chrome.settingsPrivate.setDefaultZoom(0.5);')
rootdriver.implicitly_wait(10)
wait = WebDriverWait(rootdriver, 5)
nrow=ws.rows
print('NROW'+str(nrow))
j=2
z=0
end=0
EE=0
Tr=0
C_mois=0
C_mois5=0
drive=0
date = int(datetime.datetime.now().day)
f_mounth=1
fm=2
fff=0
f_xpathdate=0
w_month=0
c_month=0
bouton_mois_suivant=0
R1=0
R2=0
R3=0
R4=0
R5=0
L1=0
L2=0
L3=0
L4=0
L5=0
P1=0
P2=0
PLUS=0
C1=0
total_R=0
total_L=0
total_P=0
total_PLUS=0
total_C=0
i=1
cACTIVE=0
while cACTIVE==0:
	#g=ws.cell(row=1, column=i).value
	g=ws.cell((1, i)).value
	if g=='ACTIVE_YES/NO':
		cACTIVE=i
	i=i+1
i=1
cANNONCE=0
while cANNONCE==0:
	#g=ws.cell(row=1, column=i).value
	g=ws.cell((1, i)).value
	if g=='ANNONCE':
		cANNONCE=i
	i=i+1	
while f_mounth==0:
	#h=ws.cell(row=fm, column=cANNONCE).value
	h=ws.cell((fl, cANNONCE)).value
	print(h)
	#if fff==2:
	#	f_mounth=1
	#	f_xpathdate=1
	#	end=1
	#	run=emailfalde()
	fff=fff+1
	try:
		print('ici1')
		rootdriver.get(h)
		time.sleep(2)
		html = rootdriver.page_source
		soup = BeautifulSoup(html, 'html.parser')
		time.sleep(3)
		print('ici2')
		#---mois 1---
		name_mois1 = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1lds9wb'][1]//div[@class='_gucugi']/strong"))).text
		name_mois2 = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1lds9wb'][2]//div[@class='_gucugi']/strong"))).text
		#name_mois3 = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_gucugi'][3]/strong"))).text
		print('ici3')
		print(name_mois1)
		print(name_mois2)
		#print(name_mois3)
		mm3=0
		u=0
		while mm3==0:
			print('la')
			month311=soup.find('div', attrs={"class":u"_kuxo8ai"})
			month31=month311.find('div', attrs={"class":u"_gucugi"})
			name_mois3=month31.find('strong').text
			
			print ('name m3='+str(name_mois3))
			if name_mois3==None:
				mm3=0
				u=u+1
				print ('none')
			else:
				mm3=1
				print ('m3 yes')
			if u==3:
				mm3=1
			print ('go')
		
		print(name_mois1)
		Mname1=name_mois1.split(' ')
		MN1=Mname1[0]
		run_MN=MnumDay(MN1)
		print (MNumday)
		MNday1=MNumday
		run_c=A_Colonne_mois(name_mois1,k)
		m1_write=c_write
		m1_newmonth=new_month
		#---mois 2---
		
		print(name_mois2)
		Mname2=name_mois2.split(' ')
		MN2=Mname2[0]
		run_MN=MnumDay(MN2)
		print (MNumday)
		MNday2=MNumday
		run_c=A_Colonne_mois(name_mois2,k)
		m2_write=c_write
		m2_newmonth=new_month
		#---mois 3---
		#month31=soup.findAll('div', attrs={"class":u"_gucugi"})[3]
		#name_mois3=month31.find('strong').text

		run_c=A_Colonne_mois(name_mois3,k)
		m3_write=c_write
		m3_newmonth=new_month
		
		#---mois 4 et 5---
		print('TRY CLICK')
		ele=rootdriver.find_element_by_xpath("//div[@aria-label='Avancez pour passer au mois suivant.']")
		print('TRY FAILED')
		rootdriver.execute_script("arguments[0].scrollIntoView(true);", ele)
		rootdriver.execute_script("window.scrollBy(0,-500);")
		next_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Avancez pour passer au mois suivant.']")))
		next_calendar.click()
		time.sleep(2)
		next_calendar.click()
		time.sleep(2)
		next_calendar.click()
		time.sleep(2)
		html = rootdriver.page_source
		soup = BeautifulSoup(html, 'html.parser')
		time.sleep(1)
		#-----RECUPERATION CALANDAR MOIS 4--------
		if C_mois5==0:
			name_mois4 = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1lds9wb'][1]//div[@class='_gucugi']/strong"))).text
			print(name_mois4)
			Mname4=name_mois4.split(' ')
			MN4=Mname4[0]
			run_MN=MnumDay(MN4)
			print (MNumday)
			MNday4=MNumday
			run_c=A_Colonne_mois(name_mois4,k)
			m4_write=c_write
			m4_newmonth=new_month
		if C_mois5==0:
			name_mois5 = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1lds9wb'][2]//div[@class='_gucugi']/strong"))).text
			print(name_mois5)
			Mname5=name_mois5.split(' ')
			MN5=Mname5[0]
			run_MN=MnumDay(MN5)
			print (MNumday)
			MNday5=MNumday
			run_c=A_Colonne_mois(name_mois5,k)
			m5_write=c_write
			m5_newmonth=new_month
		f_mounth=1
		j=fm
	except:
		fm=fm+1



while w_month==0:
	run_month=whatmounth()
	print (name_mois1)
	print (name_mois2)
	print (name_mois3)
	print (name_mois4)
	print (name_mois5)
	w_month=1

while c_month==0:
	k=k+10
#--mois1--
	print(name_mois1)
	Mname1=name_mois1.split(' ')
	MN1=Mname1[0]
	run_MN=MnumDay(MN1)
	MNday1=MNumday
	run_c=A_Colonne_mois(name_mois1,k)
	m1_write=c_write
	m1_newmonth=new_month
	print (m1_newmonth)
#--mois 2--
	print(name_mois2)
	Mname2=name_mois2.split(' ')
	MN2=Mname2[0]
	run_MN=MnumDay(MN2)
	MNday2=MNumday
	run_c=A_Colonne_mois(name_mois2,k)
	m2_write=c_write
	m2_newmonth=new_month
	print (m2_newmonth)
#--mois 3--
	print(name_mois3)
	Mname3=name_mois3.split(' ')
	MN3=Mname3[0]
	run_MN=MnumDay(MN3)
	MNday3=MNumday
	run_c=A_Colonne_mois(name_mois3,k)
	m3_write=c_write
	m3_newmonth=new_month
	print (m3_newmonth)
#--mois 4--
	print(name_mois4)
	Mname4=name_mois4.split(' ')
	MN4=Mname4[0]
	run_MN=MnumDay(MN4)
	MNday4=MNumday
	run_c=A_Colonne_mois(name_mois4,k)
	m4_write=c_write
	m4_newmonth=new_month
	print (m4_newmonth)
#--mois 5--
	print(name_mois5)
	Mname5=name_mois5.split(' ')
	MN5=Mname5[0]
	run_MN=MnumDay(MN5)
	MNday5=MNumday
	run_c=A_Colonne_mois(name_mois5,k)
	m5_write=c_write
	m5_newmonth=new_month
	print (m5_newmonth)
	time.sleep(5)
	c_month=1

f_xpathdate=1
fm=2
fff=0
while f_xpathdate==0:
	h=ws.cell(row=fm, column=cANNONCE).value
	fm=fm+1
	print(h)
	if fff==8:
		f_mounth=1
		f_xpathdate=1
		end=0
		#run=emailfalde2()
	fff=fff+1
	try:
		rootdriver.get(h)
		print('test1')
		time.sleep(8)
		#x_date = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_13m7kz7i']"))).text
		#x_date = wait.until(EC.presence_of_element_located((By.XPATH, "//td[@class='_l9wspk2']")))
		x_date = wait.until(EC.presence_of_element_located((By.XPATH, "//td[contains(@aria-label,'non')]")))
		print('date find')
		#x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_18hrqvin']")))
		#x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_5z4v7g']")))
		#x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//h1[@class='_14i3z6h']")))
		#x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//h1[@class='_fecoyn4']")))
		print('title trouve')
		ele=rootdriver.find_element_by_xpath("//button[@aria-label='Avancez pour passer au mois suivant.']")
		rootdriver.execute_script("arguments[0].scrollIntoView(true);", ele)
		rootdriver.execute_script("window.scrollBy(0,-200);")
		next_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Avancez pour passer au mois suivant.']")))
		next_calendar.click()
		print('next find')
		f_xpathdate=1
		try:
			b_cookie = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@class='optanon-allow-all accept-cookies-button']")))
			b_cookie.click()
		except:
			aaa=1
	except:
		if fff!=5:
			rootdriver.quit()
			rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
			#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
			rootdriver.set_window_size(2000, 1000)
			wait = WebDriverWait(rootdriver, 5)
def f1(a):
	global des
	try:
		#x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_5z4v7g']")))
		x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//h1[@class='_14i3z6h']")))
		#x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//span[@class='_18hrqvin']")))
		des=1
	except:
		des=0
def f2(bouton_mois_suivant):
	global next_calendar
	try:
		if bouton_mois_suivant==0:
			next_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Avancez pour passer au mois suivant.']")))
		else:
			next_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Avancez pour passer au mois suivant.']")))
	except:
		a=1

def checkmounth(name_mois1,bouton_mois_suivant):
	try:
		which_mount = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1lds9wb'][1]//h1[@class='_14i3z6h']"))).text
	except:
		which_mount = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1lds9wb'][1]//h3[@class='_14i3z6h']"))).text
	if name_mois1=='septembre 2021':
		nmount=9
	elif name_mois1=='octobre 2021':
		nmount=10	
	elif name_mois1=='novembre 2021':
		nmount=11
	elif name_mois1=='décembre 2021':
		nmount=12
	elif name_mois1=='janvier 2022':
		nmount=1
	elif name_mois1=='février 2022':
		nmount=2
	elif name_mois1=='mars 2022':
		nmount=3
	elif name_mois1=='avril 2022':
		nmount=4
	elif name_mois1=='mai 2022':
		nmount=5
	elif name_mois1=='juin 2022':
		nmount=6
	elif name_mois1=='juillet 2022':
		nmount=7
	elif name_mois1=='août 2022':
		nmount=8
	elif name_mois1=='septembre 2022':
		nmount=9
	elif name_mois1=='octobre 2022':
		nmount=10
	elif name_mois1=='novembre 2022':
		nmount=11
	elif name_mois1=='decembre 2022':
		nmount=12

		
	if which_mount!=name_mois1:
		#print('check on')
		if which_mount=='septembre 2021':
			goback=9
		elif which_mount=='octobre 2021':
			goback=10
		elif which_mount=='novembre 2021':
			goback=11
		elif which_mount=='décembre 2021':
			goback=12
		elif which_mount=='janvier 2022':
			goback=1
		elif which_mount=='février 2022':
			goback=2
		elif which_mount=='mars 2022':
			goback=3
		elif which_mount=='avril 2022':
			goback=4
		elif which_mount=='mai 2022':
			goback=5
		elif which_mount=='juin 2022':
			goback=6
		elif which_mount=='juillet 2022':
			goback=7
		elif which_mount=='août 2022':
			goback=8
		elif which_mount=='septembre 2022':
			goback=9
		elif which_mount=='octobre 2022':
			goback=10
		elif which_mount=='novembre 2022':
			goback=11
		elif which_mount=='decembre 2022':
			goback=12
		
		if goback<nmount:
			backnumber=12-nmount+goback
		else:
			backnumber=goback-nmount
		
		try:
			if bouton_mois_suivant==0:
				previous_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Reculez pour passer au mois précédent.']")))
			else:
				previous_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Reculez pour passer au mois précédent.']")))
			loop=0
			while loop<backnumber:
				previous_calendar.click()
				time.sleep(1)
				loop=loop+1
			time.sleep(1)
		except:
			a=1

def getdrive(h):
	global v_get
	rootdriver.get(h)
	v_get=1

while end==0:
	try:
		j=nrow
		while j<=nrow:
			#----COMPUTE REPORT----
			v_get=0
			total_R=total_R+R1+R2+R3+R4+R5
			total_L=total_L+L1+L2+L3+L4+L5
			total_P=total_P+P1+P2
			total_PLUS=total_PLUS+PLUS
			total_C=total_C+C1
			R1=0
			R2=0
			R3=0
			R4=0
			R5=0
			L1=0
			L2=0
			L3=0
			L4=0
			L5=0
			P1=0
			P2=0
			PLUS=0
			C1=0
			#----START TRAQUING----
			#h=ws.cell(row=j, column=cANNONCE).value
			#vACTIVE=ws.cell(row=j, column=cACTIVE).value
			h=ws.cell((j, cANNONCE)).value
			vACTIVE=ws.cell((j, cACTIVE)).value
			checker=0
			print('------'+str(j-1)+'------'+str(h)+'-----'+str(vACTIVE))
			if vACTIVE=='YES':
				if h==None:
					j=j+1
					print('h=None')
				elif 'plus' in h:
					ResAirbnb=''
					run_getdrive=getdrive(h)
					gettimer=0
					while gettimer<=70:
						gettimer=gettimer+1
						time.sleep(1)
						if v_get==1:
							gettimer=100
						elif gettimer==60:
							rootdriver.quit()
							rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
							#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
							rootdriver.set_window_size(2000, 1000)
							wait = WebDriverWait(rootdriver, 5)
					#rootdriver.get(h)
					time.sleep(5)
					des=1
					try:
						#print('search ele')
						#ele=rootdriver.find_element_by_xpath("//button[@aria-label='Avancez pour passer au mois suivant.']")
						ele = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Avancez pour passer au mois suivant.']")))
						rootdriver.execute_script("arguments[0].scrollIntoView(true);", ele)
						rootdriver.execute_script("window.scrollBy(0,-200);")
						time.sleep(2)
						try:
							run_checkmounth=checkmounth(name_mois1,bouton_mois_suivant)
						except:
							zzzz=1
						html = rootdriver.page_source
						soup = BeautifulSoup(html, 'html.parser')
						time.sleep(2)
						#print('search bouton')
						next_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Avancez pour passer au mois suivant.']")))
					#b_add_date = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@class='_3uatz29']")))
						#b_add_date.click()
						#b_arrival = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@class='_153lip8'][1]")))
						#b_arrival.click
						time.sleep(1)
						#print('1')
						run_day=A_Statu_day2(date,m1_write,1,j,0,ResAirbnb,m1_newmonth,500,0,des)
						#print('2')
						run_day=A_Statu_day2(1,m2_write,2,j,1,ResAirbnb,m2_newmonth,MNday1,0,des)
						#print('3')
						run_resday=A_Statu_day4(m3_write,j,ResAirbnb,m3_newmonth,des)
						#(date,c_write,page,j,g,ResAirbnb,new_mo,MNday,ONCOM)
						#run_PLUS_1=A_Statu_PLUS(date,m1_write,2,j,0,ResAirbnb,m1_newmonth,500,1)
						#run_PLUS_2=A_Statu_PLUS(1,m2_write,2,j,1,ResAirbnb,m2_newmonth,MNday1,0)
						#run_PLUS_3=A_Statu_PLUS2(m3_write,j,ResAirbnb,m3_newmonth,2)
						#print('4')
						next_calendar.click()
						time.sleep(1)
						next_calendar.click()
						time.sleep(1)
						next_calendar.click()
						time.sleep(1)
						html = rootdriver.page_source
						soup = BeautifulSoup(html, 'html.parser')
						time.sleep(2)
						PLUS=1
						try:
						#-----RECUPERATION CALANDAR MOIS 4--------
							run_day=A_Statu_day5(m4_write,j,ResAirbnb,m4_newmonth,0,des)
						except:
							pass
					#-----RECUPERATION CALANDAR MOIS 5--------
						try:
							run_day=A_Statu_day5(m5_write,j,ResAirbnb,m5_newmonth,1,des)
						except:
							pass
						#print('6')
						#https://www.airbnb.fr/rooms/plus/21846063
						try:
							#//span[@class='_so3dpm2']
							#Bcomment=soup.find('button', attrs={"class": "_ff6jfq"})
							#Scomment=Bcomment.find('span', attrs={"class": "_so3dpm2"}).text
							Lcomment=[]
							Icomment=0
							Scomment=soup.find('span', attrs={"class": "_bq6krt"}).text
							Lcomment=Scomment.split("(")
							Tcomment=Lcomment[1].replace(")","")
							Icomment=int(Tcomment)
							print(Icomment)
							#ws.cell(row=j, column=c_write+2).value=Icomment
							ws.update_value((j,c_write+2), Icomment)
						except:
							pass
					except:
						print('ECHEC PLUS')
					#if (j/20).is_integer():
					#	wbx.save(path_RESULT.filename)
					j=j+1
				elif 'airbnb' in h:
					run_getdrive=getdrive(h)
					gettimer=0
					while gettimer<=70:
						gettimer=gettimer+1
						time.sleep(1)
						if v_get==1:
							gettimer=100
						elif gettimer==60:
							rootdriver.quit()
							rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
							#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
							rootdriver.set_window_size(2000, 1000)
							wait = WebDriverWait(rootdriver, 5)
					#rootdriver.get(h)
					time.sleep(5)
					#x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//span[@class='_18hrqvin']"))).text
					#threading.Thread(target=scroll, args=(1,)).start()
					#try:
					#	rootdriver.execute_script("window.scrollBy(0,1200);")
					#except:
					#	time.sleep(1)
					#	pass
					f_ele=0
					#threading.Thread(target=f1, args=(h,)).start()
					des=1
					b_scrolldown=0
					rootdriver.execute_script("window.scrollBy(0,2000);")
					while f_ele<=3:
						try:
							#print("try ele")
							#ele=rootdriver.find_element_by_xpath("//div[@aria-label='Avancez pour passer au mois suivant.']")
							ele = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Avancez pour passer au mois suivant.']")))
							rootdriver.execute_script("arguments[0].scrollIntoView(true);", ele)
							rootdriver.execute_script("window.scrollBy(0,-150);")
							print("try ele")
							f_ele=6
							bouton_mois_suivant=0
							b_scrolldown=1
						except:
							try:
								ele=rootdriver.find_element_by_xpath("//button[@aria-label='Avancez pour passer au mois suivant.']")
								rootdriver.execute_script("arguments[0].scrollIntoView(true);", ele)
								rootdriver.execute_script("window.scrollBy(0,-150);")
								bouton_mois_suivant=1
								f_ele=6
								b_scrolldown=1
							except:
								#print('DOWN KO')
								time.sleep(2)
							#rootdriver.execute_script("window.scrollBy(0,2000);")
							#print('DOWN KO')

						f_ele=f_ele+1
					#time.sleep(3)
					#threading.Thread(target=f2, args=(bouton_mois_suivant,)).start()
					if b_scrolldown==1:
						td28=0
						while td28<28:
							#print("try td V4")
							try:
								#v_td = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_kuxo8ai']//div[last()]")))
								#v_td=rootdriver.find_element_by_xpath("//div[@class='_kuxo8ai']//div[last()]")
								#v_td=rootdriver.find_element_by_xpath("//div[@class='_kuxo8ai']//tr[2]/td[2]/div").text
								vv=wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class='_kuxo8ai']//table//tr//div)[last()]")))
								v_td=int(vv.get_attribute("textContent"))
								#print(v_td)
								if v_td>=28:
									td28=28
							except:
								aaa=1
							time.sleep(1)
							td28=td28+7
						#try:
						#	run_checkmounth=checkmounth(name_mois1,bouton_mois_suivant)
						#except:
					#		zzzz=1
						html = rootdriver.page_source
						time.sleep(2)
						soup = BeautifulSoup(html, 'html.parser')
						time.sleep(2)
						ResAirbnb=''
						#V_up=ws.cell(row=j, column=k).value
						v_m='8'

						#try:
						#	script=soup.find('script', attrs={"data-state":u"true"}).text
						#	p1=script.split("calendar_last")
						#	p2=p1[1].split("guest_controls")
						#	p3=p2[0].replace('_updated_at":"', '')
						#	p4=p3.replace('","', '')
							#print (p4)
						#	if p4==V_up:
								#ResAirbnb='/A'
						#		ResAirbnb=''
						#	else:
								#ws.cell(row=j, column=k).value=p4
						#		ResAirbnb=''
						#except:
						#	pass
						try:
						#-----RECUPERATION CALANDAR MOIS 1--------
							#print('le mois N est '+name_mois1)
							run_day=A_Statu_day2(date,m1_write,1,j,0,ResAirbnb,m1_newmonth,500,1,des)
						except:
							pass
						try:
						#-----RECUPERATION CALANDAR MOIS 2--------
							#print('le mois N+1 est '+name_mois2)
							run_day=A_Statu_day2(1,m2_write,2,j,1,ResAirbnb,m2_newmonth,MNday1,0,des)
						except:
							pass
						try:
						#-----RECUPERATION CALANDAR MOIS 3--------
							#print('le mois N+2 est '+name_mois3)
							RA4=ResAirbnb
							if v_m=='X' and date==1:
								RA4='/D'
							run_resday=A_Statu_day4(m3_write,j,RA4,m3_newmonth,des)
						except:
							#print('PAS DE MOIS 3')
							pass
					#-----MOIS 4-5 -----
						if v_m!='z':
							try:
								ele.click()
								time.sleep(1)
								ele.click()
								time.sleep(1)
								ele.click()
								time.sleep(2)
								print("try click")
								html = rootdriver.page_source
								soup = BeautifulSoup(html, 'html.parser')
								time.sleep(1)
								try:
								#-----RECUPERATION CALANDAR MOIS 4--------
									#print('le mois N est '+name_mois4)
									run_day=A_Statu_day5(m4_write,j,ResAirbnb,m4_newmonth,0,des)
								except:
									pass
							#-----RECUPERATION CALANDAR MOIS 5--------
								try:
									#print('le mois N+1 est '+name_mois5)
									run_day=A_Statu_day5(m5_write,j,ResAirbnb,m5_newmonth,1,des)
									#run_resday=A_Statu_day4(m5_write,j,RA4,m5_newmonth,des)
								except:
									pass
							except:
								#print('----click KO')
								zzz=1
								pass
							C_mois5=1
							checker=1
					#if (j/100).is_integer():
					#	wbx.save(path_RESULT.filename)
						#if checker==1:
						#	x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_1b3ij9t']")))
					if (j/500).is_integer():
						rootdriver.quit()
						time.sleep(5)
						rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
						rootdriver.set_window_size(2000, 1000)
						rootdriver.get('https://www.google.com/')
						rootdriver.get('chrome://settings/')
						rootdriver.execute_script('chrome.settingsPrivate.setDefaultZoom(0.5);')
						rootdriver.implicitly_wait(10)
						wait = WebDriverWait(rootdriver, 5)
						time.sleep(5)
						rootdriver.get(h)
						time.sleep(5)
						
					C_mois=1
					j=j+1
				elif 'abritel' in h:
					j=j+1
				else:
					j=j+1
			else:
				j=j+1
		
		#wbx.save(path_RESULT.filename)
		end=1
		now = str(datetime.datetime.now())[:19]
		now = now.replace(":","_")
		print(now)
		Tr=date
		print ('FIN')
		#wbx = load_workbook(path_RESULT.filename)
		#ws = wbx.active
		try:
			print("compute")
			COMPUTE_M1('octobre 2022')
			#COMPUTE_M1(name_mois2)
		except:
			print('no compute')
		#wbx.save(DIR2+NAMEFile+str(now)+".xlsx")
		try:
			print('///REPORT///')
			print('Total R = '+str(total_R))
			print('Total L = '+str(total_L))
			print('Total P = '+str(total_P))
			print('Total PLUS = '+str(total_PLUS))
			print('Total C = '+str(total_C))
		except:
			print('NO REPORT')
		try:
			run=email(DIR2,NAMEFile,now,total_R,total_L,total_P,total_C)
			print('sent email')
		except:
			print('rien')
		#rootdriver.quit()
		#wbx.close()
	except:
		print(j)
		j=j+1
		try:
			rootdriver.quit()
		except:
			aaa=1
		# EXCEPT si Chrome se ferme tout seul, ici il va le réouvrir et relancer la boucle d'extraction
		rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
		#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
		rootdriver.set_window_size(2000, 1000)
		rootdriver.set_window_size(2000, 1000)
		rootdriver.get('https://www.google.com/')
		rootdriver.get('chrome://settings/')
		rootdriver.execute_script('chrome.settingsPrivate.setDefaultZoom(0.5);')
		rootdriver.implicitly_wait(10)
		wait = WebDriverWait(rootdriver, 5)
		f_xpathdate=1
		fm=2
		fff=0
		while f_xpathdate==0:
			print('test2')
			#h=ws.cell(row=fm, column=cANNONCE).value
			h=ws.cell((fm, cANNONCE)).value
			fm=fm+1
			print(h)
			if fff==8:
				f_mounth=1
				f_xpathdate=1
				end=0
				#run=emailfalde2()
			fff=fff+1
			try:
				rootdriver.get(h)
				time.sleep(8)
				#x_date = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_13m7kz7i']"))).text
				#x_date = wait.until(EC.presence_of_element_located((By.XPATH, "//td[@class='_l9wspk2']")))
				x_date = wait.until(EC.presence_of_element_located((By.XPATH, "//td[contains(@aria-label,'non')]")))
				print('date find')
				#x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_18hrqvin']")))
				#x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_5z4v7g']")))
				#x_title = wait.until(EC.presence_of_element_located((By.XPATH, "//h1[@class='_14i3z6h']")))
				print('title trouve')
				ele=rootdriver.find_element_by_xpath("//button[@aria-label='Avancez pour passer au mois suivant.']")
				rootdriver.execute_script("arguments[0].scrollIntoView(true);", ele)
				rootdriver.execute_script("window.scrollBy(0,-200);")
				next_calendar = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Avancez pour passer au mois suivant.']")))
				next_calendar.click()
				print('next find')
				f_xpathdate=1
				try:
					b_cookie = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@class='optanon-allow-all accept-cookies-button']")))
					b_cookie.click()
				except:
					aaa=1
			except:
				if fff!=5:
					rootdriver.quit()
					rootdriver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver',chrome_options=chrome_options)
					#rootdriver = webdriver.Chrome(chrome_options=chrome_options)
					rootdriver.set_window_size(2000, 1000)
					rootdriver.set_window_size(2000, 1000)
					rootdriver.get('https://www.google.com/')
					rootdriver.get('chrome://settings/')
					rootdriver.execute_script('chrome.settingsPrivate.setDefaultZoom(0.5);')
					rootdriver.implicitly_wait(10)
					wait = WebDriverWait(rootdriver, 5)

		
		

print('FIN')

#code='https://raw.githubusercontent.com/VincentAulnay/PYTEST/master/FIND_NEW_ANNONCE.py'

go=0
if go==1:
	try:
		#code='https://raw.githubusercontent.com/VincentAulnay/PYTEST/master/ADD_COLUMN.py'
		code='https://raw.githubusercontent.com/VincentAulnay/PYTEST/master/FIND_NEW_ANNONCE.py'
		response=urllib.request.urlopen(code)
		data=response.read()
		exec(data)
		run=email(DIR2,NAMEFile,now,total_R,total_L,total_P)
		print('sent email')
	except:
		print('rien')
clean=0
if clean==1:
	try:
		code='https://raw.githubusercontent.com/VincentAulnay/PYTEST/master/CLEAN.py'
		response=urllib.request.urlopen(code)
		data=response.read()
		exec(data)
		print('CLEAN DONE')
	except:
		print('rien')
active=0
if active==1:
	try:
		code='https://raw.githubusercontent.com/VincentAulnay/PYTEST/master/ACTIVE_YESNO.py'
		response=urllib.request.urlopen(code)
		data=response.read()
		exec(data)
		print('ACTIVE DONE')
	except:
		print('rien')
launch_check=0
if launch_check==1:
	j=2
	while j<=nrow:
		#----START TRAQUING----
		h=ws.cell(row=j, column=cANNONCE).value
		rootdriver.get(h)
		time.sleep(5)
		while f_ele<6:
			try:
				ele=rootdriver.find_element_by_xpath("//div[@aria-label='Avancez pour passer au mois suivant.']")
				rootdriver.execute_script("arguments[0].scrollIntoView(true);", ele)
				f_ele=7
				ws.cell(row=j, column=cACTIVE).value='YES'
			except:
				try:
					ele=rootdriver.find_element_by_xpath("//button[@aria-label='Avancez pour passer au mois suivant.']")
					rootdriver.execute_script("arguments[0].scrollIntoView(true);", ele)
					f_ele=7
					ws.cell(row=j, column=cACTIVE).value='YES'
				except:
					time.sleep(1)

			f_ele=f_ele+1
			if f_ele==6:
				ws.cell(row=j, column=cACTIVE).value='NO'
		j=j+1
	wbx = load_workbook(path_RESULT.filename)
	response=urllib.request.urlopen(code)
	data=response.read()
	exec(data)
	run=email(DIR2,NAMEFile,now,total_R,total_L,total_P)
	print('sent email')
try:
	rootdriver.quit()
	wbx.close()
except:
	pass
